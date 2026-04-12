import io
import platform
import subprocess
import sys
import os
import shutil
import random
import traceback
import importlib.util
import cv2
import socket
import datetime
import matplotlib
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from PIL import Image
from openpyxl.styles import PatternFill
from segmenter import RockSegmenter
from cv2_io_utils import cv2_imread, cv2_imwrite
from matplotlib.figure import Figure
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg, NavigationToolbar2QT as NavigationToolbar
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QLabel, QFileDialog, QMessageBox,
                             QFrame, QProgressBar, QTextEdit, QTextBrowser, QScrollArea,
                             QPushButton, QLineEdit, QSizePolicy, QInputDialog, QToolBar, QMenu,
                             QGroupBox, QComboBox, QSlider, QDialog, QProgressDialog, QDoubleSpinBox, QSpinBox,
                             QCheckBox,
                             QRadioButton, QButtonGroup, QTabWidget, QTreeWidget, QTreeWidgetItem, QTableWidget,
                             QTableWidgetItem, QHeaderView, QGridLayout, QListWidget, QListWidgetItem,
                             QAbstractItemView, QFormLayout)
from PyQt6.QtGui import QPixmap, QImage, QAction, QDesktopServices, QIcon, QCloseEvent, QPainter, QPen, QColor, QFont, \
    QBrush
from PyQt6.QtCore import Qt, QSize, QThread, pyqtSignal, QUrl, QBuffer, QIODevice, QSettings, QRect, QPoint, QTimer, \
    QEvent

# ========== 论文用图导出：统一格式选项（多格式 + 高DPI）==========
IMAGE_EXPORT_FILTER = (
    "PNG 图像 (*.png);;"
    "JPEG 图像 (*.jpg *.jpeg);;"
    "TIFF 图像 (*.tif *.tiff);;"
    "PDF 矢量/高清 (*.pdf)"
)


def get_export_format_from_path(save_path):
    """根据保存路径返回 (matplotlib 格式名, 扩展名)，用于统一导出逻辑"""
    if not save_path:
        return 'png', '.png'
    path_lower = save_path.lower()
    if path_lower.endswith('.pdf'):
        return 'pdf', '.pdf'
    if path_lower.endswith(('.tif', '.tiff')):
        return 'tif', '.tif'
    if path_lower.endswith(('.jpg', '.jpeg')):
        return 'jpg', '.jpg'
    return 'png', '.png'


# ========== 单张图像分割后台线程 ==========
class SingleImageSegmentationWorker(QThread):
    """单张图像分割后台线程，避免主界面卡顿"""
    progress_signal = pyqtSignal(str, int)  # 消息, 进度0-100
    finished_signal = pyqtSignal(dict)
    error_signal = pyqtSignal(str)

    def __init__(self, image, image_name, selected_methods, segmenter, parent=None):
        super().__init__(parent)
        self.image = image
        self.image_name = image_name
        self.selected_methods = selected_methods
        self.segmenter = segmenter
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    def run(self):
        try:
            total = len(self.selected_methods)
            current = 0

            def progress_cb(msg):
                nonlocal current
                if self._cancelled:
                    raise InterruptedError("用户取消")
                if "完成" in msg or "成功" in msg or "失败" in msg or "出错" in msg:
                    current += 1
                pct = 10 + int((current / total) * 80) if total else 10
                self.progress_signal.emit(msg, min(99, pct))

            results = self.segmenter.segment_by_methods(
                self.image, self.image_name, self.selected_methods, log_callback=progress_cb
            )
            self.progress_signal.emit("完成", 100)
            self.finished_signal.emit(results)
        except InterruptedError:
            self.progress_signal.emit("已取消", 0)
        except Exception as e:
            self.error_signal.emit(str(e))


# ========== 单个单色识别线程 ==========
class ColorAnalysisWorker(QThread):
    progress_updated = pyqtSignal(int)
    result_ready = pyqtSignal(str, list, int)
    error_occurred = pyqtSignal(str)

    def __init__(self, image, grid_size, standard_vectors, color_names, color_codes):
        super().__init__()
        self.image = image
        self.grid_size = grid_size
        self.standard_vectors = standard_vectors
        self.color_names = color_names
        self.color_codes = color_codes

    def run(self):
        try:
            # 1. 图像预处理
            self.progress_updated.emit(0)

            # 亮度归一化
            img_normalized = self.normalize_brightness(self.image)  # 修复：使用实例方法
            self.progress_updated.emit(10)

            # 缩放处理
            scale_factor = 0.5
            img_resized = cv2.resize(
                img_normalized,
                (int(img_normalized.shape[1] * scale_factor),
                 int(img_normalized.shape[0] * scale_factor))
            )
            self.progress_updated.emit(20)

            height, width, _ = img_resized.shape
            half_grid = self.grid_size // 2
            x_range = width - half_grid
            y_range = height - half_grid

            # 2. 提取颜色向量
            color_vectors = []
            total_steps = (x_range - half_grid) * (y_range - half_grid)
            current_step = 0

            for i in range(half_grid, x_range):
                for j in range(half_grid, y_range):
                    grid_pixels = img_resized[j - half_grid:j + half_grid + 1,
                    i - half_grid:i + half_grid + 1]

                    if self.grid_size < 15:
                        color_values = self.arithmetic_mean(grid_pixels)  # 修复：使用实例方法
                    else:
                        color_values = self.weighted_mean(grid_pixels, self.grid_size)  # 修复：使用实例方法

                    color_vectors.append(color_values)

                    current_step += 1
                    # 每处理100个点更新一次进度（25-65%）
                    if current_step % 100 == 0:
                        progress = 25 + int((current_step / total_steps) * 40)
                        self.progress_updated.emit(progress)

            self.progress_updated.emit(65)

            # 3. 颜色匹配
            color_frequency, color_matches = self.find_closest_color(color_vectors)
            total_vectors = len(color_vectors)

            # 4. 构建统计结果
            structured_stats = []
            sorted_colors = sorted(color_frequency.items(), key=lambda x: x[1], reverse=True)

            for idx, ((color_name, color_code), count) in enumerate(sorted_colors):
                percentage = (count / total_vectors) * 100

                # 获取该颜色下的所有目标向量
                relevant_matches = [m for m in color_matches
                                    if m["color_name"] == color_name and m["color_code"] == color_code]

                if relevant_matches:
                    vectors_arr = np.array([m["target_vector"] for m in relevant_matches])
                    avg_target_vec = np.mean(vectors_arr, axis=0).astype(int)
                    std_vec = relevant_matches[0]["standard_vector"].astype(int)
                else:
                    avg_target_vec = np.array([0, 0, 0])
                    std_vec = np.array([0, 0, 0])

                structured_stats.append({
                    "name": color_name,
                    "code": color_code,
                    "count": count,
                    "percent": percentage,
                    "std_vec": std_vec,
                    "target_vec": avg_target_vec
                })

                # 每处理10种颜色更新一次进度（65-95%）
                if idx % 10 == 0:
                    progress = 65 + int((idx / len(sorted_colors)) * 30)
                    self.progress_updated.emit(progress)

            self.progress_updated.emit(95)

            # 5. 生成文本报告
            report_text = "=== 颜色分析报告 ===\n\n"
            for stat in structured_stats:
                report_text += f"{stat['name']} ({stat['code']}): {stat['percent']:.2f}%\n"

            # 6. 发送结果信号
            self.result_ready.emit(report_text, structured_stats, total_vectors)
            self.progress_updated.emit(100)

        except Exception as e:

            error_msg = f"分析过程中发生错误:\n{str(e)}\n\n详细追踪:\n{traceback.format_exc()}"
            self.error_occurred.emit(error_msg)

    def normalize_brightness(self, img):
        """对图像进行亮度均衡化，增强颜色识别稳定性"""
        img_yuv = cv2.cvtColor(img, cv2.COLOR_BGR2YUV)
        img_yuv[:, :, 0] = cv2.equalizeHist(img_yuv[:, :, 0])  # 均衡化亮度通道
        return cv2.cvtColor(img_yuv, cv2.COLOR_YUV2BGR)

    @staticmethod
    def arithmetic_mean(grid_pixels):
        """计算算术平均法"""
        mean_r = np.mean(grid_pixels[:, :, 2])
        mean_g = np.mean(grid_pixels[:, :, 1])
        mean_b = np.mean(grid_pixels[:, :, 0])
        return [int(round(mean_r)), int(round(mean_g)), int(round(mean_b))]

    @staticmethod
    def weighted_mean(grid_pixels, grid_size):
        """计算加权平均法"""
        half_grid = grid_size // 2
        height, width, _ = grid_pixels.shape
        y, x = np.meshgrid(np.arange(height), np.arange(width), indexing='ij')
        dist = np.sqrt((y - half_grid) ** 2 + (x - half_grid) ** 2)
        weights = np.exp(-dist ** 2 / (2 * (half_grid ** 2)))

        mean_r = np.round(np.sum(grid_pixels[:, :, 2] * weights) / np.sum(weights))
        mean_g = np.round(np.sum(grid_pixels[:, :, 1] * weights) / np.sum(weights))
        mean_b = np.round(np.sum(grid_pixels[:, :, 0] * weights) / np.sum(weights))
        return [int(mean_r), int(mean_g), int(mean_b)]

    def find_closest_color(self, color_vectors):
        """精确的颜色匹配算法，特别处理黑白等极端颜色"""
        color_matches = []
        color_frequency = {}

        # 将输入转换为numpy数组
        color_vectors = np.array(color_vectors)

        # 特别处理黑色和白色的阈值
        black_threshold = 50  # RGB值都低于这个阈值认为是黑色
        white_threshold = 200  # RGB值都高于这个阈值认为是白色

        for target_color in color_vectors:
            # 提取RGB分量（后三个元素）
            r, g, b = target_color[0], target_color[1], target_color[2]

            # 首先检查是否为黑色
            if r <= black_threshold and g <= black_threshold and b <= black_threshold:
                # 直接匹配为黑色
                best_index = 8  # 黑色在数据中的索引是8
            # 然后检查是否为白色
            elif r >= white_threshold and g >= white_threshold and b >= white_threshold:
                # 直接匹配为白色
                best_index = 4  # 白色在数据中的索引是4
            else:
                # 对于其他颜色，使用改进的匹配算法
                target_rgb = np.array([r, g, b])

                # 计算与所有标准颜色的距离（只使用RGB分量）
                distances = np.linalg.norm(self.standard_vectors - target_rgb, axis=1)

                # 找到最小距离的索引
                best_index = np.argmin(distances)

            color_name = self.color_names[best_index]
            color_code = self.color_codes[best_index]

            # 记录匹配详情
            color_matches.append({
                "target_vector": target_color,
                "color_name": color_name,
                "color_code": color_code,
                "standard_vector": self.standard_vectors[best_index],
                "distance": np.linalg.norm(self.standard_vectors[best_index] - np.array([r, g, b]))
            })

            color_key = (color_name, color_code)
            color_frequency[color_key] = color_frequency.get(color_key, 0) + 1

        return color_frequency, color_matches


# ========== 颜色识别模型推理线程 ==========
class ColorModelAnalysisWorker(QThread):
    """使用训练好的分类模型进行颜色识别"""
    progress_updated = pyqtSignal(int)
    result_ready = pyqtSignal(str, list, int)
    error_occurred = pyqtSignal(str)

    def __init__(self, image, grid_size, model_path, color_names_csv, color_codes_csv):
        super().__init__()
        self.image = image
        self.grid_size = grid_size
        self.model_path = model_path
        self.color_names_csv = color_names_csv  # 用于查找 color_code
        self.color_codes_csv = color_codes_csv

    def run(self):
        try:
            import torch
            from PIL import Image as PILImage
        except ImportError:
            self.error_occurred.emit("未安装 PyTorch，无法使用模型识别。请使用颜色数据库匹配。")
            return

        try:
            from torchvision import transforms
            from torchvision import models
        except (ImportError, RuntimeError, Exception) as e:
            err_msg = str(e).strip()
            if "nms" in err_msg or "does not exist" in err_msg or "torchvision" in err_msg:
                self.error_occurred.emit(
                    "PyTorch 与 torchvision 版本不兼容，无法使用模型识别。\n\n"
                    "请使用「颜色数据库匹配」进行识别，或在本机重装与 PyTorch 版本匹配的 torchvision。"
                )
            else:
                self.error_occurred.emit(f"加载 torchvision 失败: {err_msg}\n请使用颜色数据库匹配或检查环境。")
            return

        try:
            import torch.nn as nn
        except ImportError:
            self.error_occurred.emit("未安装 PyTorch，无法使用模型识别。请使用颜色数据库匹配。")
            return

        try:
            self.progress_updated.emit(0)
            # 加载模型
            ckpt = torch.load(self.model_path, map_location='cpu')
            if not isinstance(ckpt, dict):
                self.error_occurred.emit("无效的模型文件格式。请指定「颜色识别模型」（分类模型 .pth），非 U-Net 分割模型。")
                return
            # 区分：U-Net 分割模型存的是 model_state；颜色识别模型存的是 model_state_dict + num_classes + class_names
            if ckpt.get('model_state') is not None and 'num_classes' not in ckpt:
                self.error_occurred.emit(
                    "您选择的是「图像分割模型」（U-Net），不是「颜色识别模型」。\n\n"
                    "请到 设置→路径设置 中，将「颜色识别模型路径」指定为颜色分类模型（.pth），"
                    "将「分割模型路径」指定为 U-Net 分割模型，二者不要混用。"
                )
                return
            if 'num_classes' not in ckpt or 'model_state_dict' not in ckpt:
                self.error_occurred.emit(
                    "当前文件不是颜色识别模型。请在 设置→路径设置 中指定「颜色识别模型路径」"
                    "（分类模型 .pth），与「分割模型路径」（U-Net）区分开。"
                )
                return
            model_type = ckpt.get('model_type', 'resnet')
            num_classes = ckpt['num_classes']
            class_names = ckpt.get('class_names')
            if not class_names:
                self.error_occurred.emit("模型文件中缺少类别名称，请重新训练并保存。")
                return

            # 构建模型
            if model_type == 'resnet':
                model = models.resnet50(weights=None)
                model.fc = nn.Linear(model.fc.in_features, num_classes)
            elif model_type == 'efficientnet':
                model = models.efficientnet_b0(weights=None)
                model.classifier[1] = nn.Linear(model.classifier[1].in_features, num_classes)
            elif model_type == 'mobilenet':
                model = models.mobilenet_v3_large(weights=None)
                model.classifier[3] = nn.Linear(model.classifier[3].in_features, num_classes)
            else:
                self.error_occurred.emit(f"暂不支持的模型类型: {model_type}")
                return

            state = ckpt['model_state_dict']
            # 兼容 DataParallel 保存的模型（key 带 module. 前缀）
            if next(iter(state.keys())).startswith('module.'):
                state = {k.replace('module.', ''): v for k, v in state.items()}
            model.load_state_dict(state, strict=True)
            model.eval()

            transform = transforms.Compose([
                transforms.Resize((224, 224)),
                transforms.ToTensor(),
                transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
            ])

            self.progress_updated.emit(15)
            img_normalized = self._normalize_brightness(self.image)
            scale_factor = 0.5
            img_resized = cv2.resize(img_normalized, (int(img_normalized.shape[1] * scale_factor),
                                                      int(img_normalized.shape[0] * scale_factor)))
            self.progress_updated.emit(25)

            height, width, _ = img_resized.shape
            half_grid = self.grid_size // 2
            x_range = width - half_grid
            y_range = height - half_grid

            patches = []
            positions = []
            total = (x_range - half_grid) * (y_range - half_grid)
            idx = 0
            for i in range(half_grid, x_range):
                for j in range(half_grid, y_range):
                    grid = img_resized[j - half_grid:j + half_grid + 1, i - half_grid:i + half_grid + 1]
                    rgb = cv2.cvtColor(grid, cv2.COLOR_BGR2RGB)
                    pil_img = PILImage.fromarray(rgb)
                    patches.append(transform(pil_img))
                    positions.append((i, j))
                    idx += 1
                    if idx % 200 == 0:
                        self.progress_updated.emit(25 + int((idx / total) * 40))

            self.progress_updated.emit(65)
            # 批量推理
            batch_size = 32
            preds = []
            with torch.no_grad():
                for k in range(0, len(patches), batch_size):
                    batch = torch.stack(patches[k:k + batch_size])
                    out = model(batch)
                    _, pred = out.max(1)
                    preds.extend(pred.cpu().numpy().tolist())
                    if (k // batch_size) % 5 == 0:
                        self.progress_updated.emit(65 + int((k / len(patches)) * 30))

            self.progress_updated.emit(95)
            # 统计
            color_frequency = {}
            for pred_idx in preds:
                name = class_names[pred_idx]
                code = self._lookup_code(name)
                key = (name, code)
                color_frequency[key] = color_frequency.get(key, 0) + 1

            total_vectors = len(preds)
            structured_stats = []
            sorted_colors = sorted(color_frequency.items(), key=lambda x: x[1], reverse=True)
            for (color_name, color_code), count in sorted_colors:
                pct = (count / total_vectors) * 100
                structured_stats.append({
                    "name": color_name,
                    "code": color_code,
                    "count": count,
                    "percent": pct,
                    "std_vec": np.array([0, 0, 0]),
                    "target_vec": np.array([0, 0, 0])
                })

            report_text = "=== 颜色分析报告（模型识别）===\n\n"
            for s in structured_stats:
                report_text += f"{s['name']} ({s['code']}): {s['percent']:.2f}%\n"

            self.result_ready.emit(report_text, structured_stats, total_vectors)
            self.progress_updated.emit(100)
        except Exception as e:
            self.error_occurred.emit(f"模型识别出错: {str(e)}\n{traceback.format_exc()}")

    def _normalize_brightness(self, img):
        img_yuv = cv2.cvtColor(img, cv2.COLOR_BGR2YUV)
        img_yuv[:, :, 0] = cv2.equalizeHist(img_yuv[:, :, 0])
        return cv2.cvtColor(img_yuv, cv2.COLOR_YUV2BGR)

    def _lookup_code(self, color_name):
        if self.color_names_csv is not None and self.color_codes_csv is not None:
            for i, n in enumerate(self.color_names_csv):
                if n == color_name or (isinstance(n, str) and str(n).strip() == str(color_name).strip()):
                    return self.color_codes_csv[i] if i < len(self.color_codes_csv) else ""
        return ""


# ========== 绘图工具 ==========
class MatplotlibWindow(QMainWindow):
    """通用的 Matplotlib 独立窗口，自带工具栏（保存、缩放、移动）"""

    def __init__(self, title="Chart", parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(700, 500)
        self.parent = parent

        # 核心部件
        self.main_widget = QWidget()
        self.setCentralWidget(self.main_widget)
        layout = QVBoxLayout(self.main_widget)

        # 1. 创建 Figure
        self.figure = Figure(figsize=(5, 4), dpi=100)
        self.canvas = FigureCanvasQTAgg(self.figure)

        # 2. 创建原生工具栏
        self.toolbar = NavigationToolbar(self.canvas, self)

        # 拦截保存事件，添加我们的交互对话框
        for action in self.toolbar.actions():
            if action.text() == 'Save':
                action.triggered.disconnect()
                action.triggered.connect(self.custom_save)
                break

        # 3. 布局
        layout.addWidget(self.toolbar)
        layout.addWidget(self.canvas)

    def custom_save(self):
        """自定义保存功能，添加交互对话框"""
        # 使用主窗口记忆的最近目录
        initial_path = "chart.png"
        if hasattr(self.parent, "get_last_directory"):
            last_dir = self.parent.get_last_directory()
            if last_dir:
                initial_path = os.path.join(last_dir, "chart.png")

        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存图表", initial_path,
            IMAGE_EXPORT_FILTER
        )

        if file_path:
            if hasattr(self.parent, "set_last_directory"):
                self.parent.set_last_directory(file_path)
            fmt, _ = get_export_format_from_path(file_path)
            savefig_fmt = {'png': 'png', 'jpg': 'jpg', 'tif': 'tiff', 'pdf': 'pdf'}.get(fmt, 'png')
            dpi = getattr(self.parent, 'dpi', 300) if hasattr(self.parent, 'dpi') else 300
            try:
                self.figure.savefig(file_path, dpi=dpi, bbox_inches="tight", format=savefig_fmt)
            except Exception as e:
                QMessageBox.critical(self, "保存失败", f"保存图表时发生错误：\n{e}")

    def get_figure(self):
        return self.figure

    def draw(self):
        self.canvas.draw()


# ========== 点击查看大图 ==========
class ImageViewerWindow(QMainWindow):
    """通用大图查看窗口：支持缩放与拖拽平移，打开时自动适应窗口显示全图"""

    def __init__(self, image_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle(os.path.basename(image_path))
        self.resize(900, 650)

        central = QWidget()
        self.setCentralWidget(central)
        v = QVBoxLayout(central)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setAlignment(Qt.AlignmentFlag.AlignCenter)
        v.addWidget(self.scroll)

        self.label = QLabel()
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label.setMouseTracking(True)
        self.scroll.setWidget(self.label)

        self._pixmap = QPixmap(image_path)
        self._drag_start = None
        self._scroll_start = None
        self._initial_fit_done = False
        self.scroll.viewport().setMouseTracking(True)
        self.scroll.viewport().installEventFilter(self)
        self._zoom = 1.0
        self.apply_zoom()

    def showEvent(self, event):
        """首次显示时自动缩放以适应窗口，显示图片全貌"""
        super().showEvent(event)
        if not self._initial_fit_done and not self._pixmap.isNull():
            self._initial_fit_done = True
            vp = self.scroll.viewport()
            vw = max(100, vp.width())
            vh = max(100, vp.height())
            pw = self._pixmap.width()
            ph = self._pixmap.height()
            if pw > 0 and ph > 0:
                zoom_w = vw / pw
                zoom_h = vh / ph
                self._zoom = min(zoom_w, zoom_h, 1.0) * 0.95
                self._zoom = max(0.05, min(self._zoom, 10))
                self.apply_zoom()

    def eventFilter(self, obj, event):
        if obj == self.scroll.viewport():
            if event.type() == QEvent.Type.MouseButtonPress and event.button() == Qt.MouseButton.LeftButton:
                self._drag_start = event.pos()
                self._scroll_start = QPoint(self.scroll.horizontalScrollBar().value(),
                                            self.scroll.verticalScrollBar().value())
            elif event.type() == QEvent.Type.MouseMove and self._drag_start is not None and event.buttons() & Qt.MouseButton.LeftButton:
                delta = event.pos() - self._drag_start
                self.scroll.horizontalScrollBar().setValue(self._scroll_start.x() - delta.x())
                self.scroll.verticalScrollBar().setValue(self._scroll_start.y() - delta.y())
            elif event.type() == QEvent.Type.MouseButtonRelease and event.button() == Qt.MouseButton.LeftButton:
                self._drag_start = None
        return super().eventFilter(obj, event)

    def wheelEvent(self, event):
        if event.angleDelta().y() > 0:
            self._zoom *= 1.1
        else:
            self._zoom /= 1.1
        self._zoom = max(0.1, min(self._zoom, 10))
        self.apply_zoom()

    def apply_zoom(self):
        if not self._pixmap.isNull():
            scaled = self._pixmap.scaled(
                self._pixmap.size() * self._zoom,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            )
            self.label.setPixmap(scaled)

    def custom_save(self):
        """自定义保存功能：支持 PNG/JPEG/TIFF/PDF 多格式与高 DPI（论文用图）"""
        initial_path = os.path.basename(self.windowTitle()) or "image.png"
        base, _ = os.path.splitext(initial_path)
        if not base.lower().endswith(('.png', '.jpg', '.jpeg', '.tif', '.tiff', '.pdf')):
            initial_path = base + ".png"
        parent = self.parent()
        if hasattr(parent, "get_last_directory") and parent.get_last_directory():
            initial_path = os.path.join(parent.get_last_directory(), initial_path)

        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存图片（可选格式与高DPI）", initial_path, IMAGE_EXPORT_FILTER
        )
        if not file_path:
            return

        if hasattr(parent, "set_last_directory"):
            parent.set_last_directory(file_path)
        dpi = getattr(parent, 'dpi', 300) if hasattr(parent, 'dpi') else 300
        try:
            fmt, _ = get_export_format_from_path(file_path)
            buffer = QBuffer()
            buffer.open(QIODevice.OpenModeFlag.ReadWrite)
            self._pixmap.save(buffer, "PNG")
            pil_img = Image.open(io.BytesIO(buffer.data()))

            if fmt == 'pdf':
                fig, ax = plt.subplots(dpi=dpi)
                ax.imshow(np.array(pil_img))
                ax.axis('off')
                plt.tight_layout(pad=0)
                fig.savefig(file_path, dpi=dpi, format='pdf', bbox_inches='tight', pad_inches=0)
                plt.close(fig)
            elif fmt == 'tif':
                pil_img.save(file_path, format='TIFF', dpi=(float(dpi), float(dpi)), compression="tiff_lzw")
            elif fmt == 'jpg':
                if pil_img.mode in ('RGBA', 'P'):
                    pil_img = pil_img.convert('RGB')
                pil_img.save(file_path, format='JPEG', quality=95, dpi=(float(dpi), float(dpi)))
            else:
                pil_img.save(file_path, format='PNG', dpi=(float(dpi), float(dpi)))
            self.show_export_success_dialog(file_path)
        except Exception as e:
            QMessageBox.critical(self, "保存失败", str(e))

    def show_export_success_dialog(self, file_path):
        """显示导出成功的交互对话框"""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("保存成功")
        msg_box.setText(f"图表已保存到：\n{file_path}")
        msg_box.setIcon(QMessageBox.Icon.Information)

        open_file_btn = msg_box.addButton("打开文件", QMessageBox.ButtonRole.ActionRole)
        open_folder_btn = msg_box.addButton("打开所在文件夹", QMessageBox.ButtonRole.ActionRole)
        msg_box.addButton("确定", QMessageBox.ButtonRole.AcceptRole)

        msg_box.exec()

        clicked_btn = msg_box.clickedButton()
        if clicked_btn == open_file_btn:
            QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))
        elif clicked_btn == open_folder_btn:
            folder_path = os.path.dirname(file_path)
            if platform.system() == "Windows":
                subprocess.run(['explorer', '/select,', os.path.normpath(file_path)])
            else:
                QDesktopServices.openUrl(QUrl.fromLocalFile(folder_path))


# ========== 分割算法选择对话框 ==========
class SegmentationMethodDialog(QDialog):
    """分割算法选择对话框"""

    def __init__(self, parent=None, default_methods=None, show_reminder=True, last_used_methods=None):
        super().__init__(parent)
        self._last_used_methods = last_used_methods or []
        self.setWindowTitle("选择分割算法")
        self.setModal(True)
        self.resize(520, 460)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # 说明文字
        info_label = QLabel("请选择要使用的分割算法（可多选）：")
        info_label.setStyleSheet("font-weight: bold; font-size: 12px; padding: 5px 0;")
        layout.addWidget(info_label)

        # 使用上一次使用过的分割算法
        last_used_btn = QPushButton("使用上一次使用过的分割算法")
        last_used_btn.setToolTip("将当前勾选恢复为上次执行分割时使用的算法组合")
        last_used_btn.clicked.connect(self._apply_last_used)
        if not self._last_used_methods:
            last_used_btn.setEnabled(False)
            last_used_btn.setToolTip("暂无上一次使用记录，先执行一次分割后即可使用")
        layout.addWidget(last_used_btn)

        # 算法选择区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        scroll_layout.setSpacing(8)

        self.method_checks = {}
        methods = [
            ('GrabCut智能分割', '使用GrabCut算法，适合复杂背景的岩石区域分割。'),
            ('颜色阈值分割', '基于HSV颜色范围，适合颜色与背景差异明显的岩石。'),
            ('边缘检测分割', '基于边缘检测和形态学操作，适合边界清晰的岩石颗粒。'),
            ('自适应阈值分割', '自适应阈值，适合光照不均或亮度变化较大的图像。'),
            ('分水岭分割', '分水岭算法，适合颗粒粘连、需要分离相邻岩石区域的情况。'),
            ('K-means聚类分割', '基于颜色聚类，适合颜色分布明显但边界不规则的图像。'),
            ('深度学习分割', 'U-Net语义分割，需先训练模型。在 设置→分割模型 中配置模型路径。')
        ]

        for method_name, description in methods:
            check = QCheckBox(method_name)
            check.setToolTip(description)
            if default_methods and method_name in default_methods:
                check.setChecked(True)

            # 为每个算法创建一个卡片：名称 + 描述
            method_frame = QFrame()
            method_frame.setStyleSheet("""
                QFrame {
                    border: 1px solid #dcdfe6;
                    border-radius: 4px;
                    background-color: #ffffff;
                }
            """)
            method_layout = QVBoxLayout(method_frame)
            method_layout.setContentsMargins(8, 6, 8, 6)
            method_layout.setSpacing(3)

            top_row = QHBoxLayout()
            top_row.addWidget(check)
            top_row.addStretch()
            if method_name == "深度学习分割" and parent and hasattr(parent, "show_segmentation_model_settings_dialog"):
                settings_btn = QPushButton("设置")
                settings_btn.setMaximumWidth(60)
                settings_btn.clicked.connect(lambda: parent.show_segmentation_model_settings_dialog())
                top_row.addWidget(settings_btn)
            method_layout.addLayout(top_row)

            desc_label = QLabel(description)
            desc_label.setStyleSheet("color: #7f8c8d; font-size: 11px;")
            desc_label.setWordWrap(True)
            method_layout.addWidget(desc_label)

            scroll_layout.addWidget(method_frame)
            self.method_checks[method_name] = check

        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)

        # 不再提醒选项
        self.dont_remind_check = QCheckBox("使用默认设置且不再提醒")
        self.dont_remind_check.setChecked(not show_reminder)
        layout.addWidget(self.dont_remind_check)

        # 选择操作按钮：全选 / 全不选 / 反选
        op_layout = QHBoxLayout()
        op_layout.addStretch()
        select_all_btn = QPushButton("全选")
        select_all_btn.clicked.connect(self.select_all)
        select_none_btn = QPushButton("全不选")
        select_none_btn.clicked.connect(self.unselect_all)
        invert_btn = QPushButton("反选")
        invert_btn.clicked.connect(self.inverse_selection)
        op_layout.addWidget(select_all_btn)
        op_layout.addWidget(select_none_btn)
        op_layout.addWidget(invert_btn)
        layout.addLayout(op_layout)

        # 确认 / 取消按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        ok_btn = QPushButton("确定")
        cancel_btn = QPushButton("取消")
        ok_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

    def select_all(self):
        """全选"""
        for check in self.method_checks.values():
            check.setChecked(True)

    def unselect_all(self):
        """全不选"""
        for check in self.method_checks.values():
            check.setChecked(False)

    def inverse_selection(self):
        """反选"""
        for check in self.method_checks.values():
            check.setChecked(not check.isChecked())

    def _apply_last_used(self):
        """将勾选设为上一次使用过的分割算法"""
        if not self._last_used_methods:
            QMessageBox.information(self, "提示", "暂无上一次使用记录。请先执行一次图像分割后再使用此功能。")
            return
        for name, check in self.method_checks.items():
            check.setChecked(name in self._last_used_methods)

    def get_selected_methods(self):
        """获取选中的方法"""
        return [name for name, check in self.method_checks.items() if check.isChecked()]

    def should_remind(self):
        """是否应该提醒"""
        return not self.dont_remind_check.isChecked()


# ========== 功能说明对话框（非模态，可复制，不影响主窗口操作）==========
class HelpDialog(QDialog):
    """功能说明对话框，非模态，内容可复制"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("RoCAS 功能说明")
        self.setModal(False)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, False)
        self.resize(720, 560)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)

        self.browser = QTextBrowser()
        self.browser.setOpenExternalLinks(True)
        self.browser.setReadOnly(True)
        self.browser.setStyleSheet("""
            QTextBrowser {
                font-family: "Microsoft YaHei", "Segoe UI", sans-serif;
                font-size: 12px;
                background-color: #fafafa;
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                padding: 8px;
            }
        """)
        layout.addWidget(self.browser)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        copy_btn = QPushButton("复制全部")
        copy_btn.clicked.connect(self._copy_all)
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.close)
        btn_row.addWidget(copy_btn)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        self._load_content()

    def _load_content(self):
        """加载帮助内容"""
        help_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resources", "help", "help.html")
        tech_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs", "技术文档.md")
        if os.path.exists(help_path):
            with open(help_path, 'r', encoding='utf-8') as f:
                html = f.read()
            self.browser.setHtml(html)
        elif os.path.exists(tech_path):
            with open(tech_path, 'r', encoding='utf-8') as f:
                text = f.read()
            self.browser.setPlainText(text)
        else:
            self.browser.setPlainText("帮助文档未找到。请确保 resources/help/help.html 或 docs/技术文档.md 存在。")

    def _copy_all(self):
        """复制全部内容到剪贴板"""
        clipboard = QApplication.clipboard()
        clipboard.setText(self.browser.toPlainText())
        QMessageBox.information(self, "已复制", "内容已复制到剪贴板。")


# ========== 主窗口 ==========
class ImageEditorApp(QMainWindow):

    # ========== 初始化界面和方法 ==========
    def __init__(self):
        super().__init__()
        self.setWindowTitle("RoCAS v2.20 (By TianXiaoying, Email: dachang0220@163.com)")
        self.setWindowIcon(QIcon(r".\resources\assets\images\logo\logo.png"))
        self.resize(800, 525)

        # 初始化变量
        self.image = None
        self.image_path = None
        self.stats_data = []  # 存储最新的分析结果
        self.total_samples = 0
        self.original_image = None  # 用于裁剪恢复

        # 分割相关属性
        self.segmenter = RockSegmenter(log_callback=self.log)
        self.exporter = HighDPIExporter()
        self.segmented_image = None
        self.segmentation_mask = None
        self.original_image_for_seg = None
        self.segmentation_method = None
        
        # UNet segmentation related properties
        self.unet_model_path = ""
        self.unet_python_path = ""
        
        # Segmentation model related properties
        self.seg_model_path = ""
        self.seg_model_use_gpu = True
        self.color_model_path = ""

        # 加载分割设置
        self.load_segmentation_settings()
        # 加载日志设置
        self.load_log_settings()

        # 初始化日志自动保存
        self.init_log_auto_save()

        # 设置全局样式（扁平化、现代风）
        self.setup_styles()

        # 初始化界面布局
        self.init_ui()
        self.create_menus()
        self.add_copyright_footer()
        self.load_csv_data()  # 预加载数据
        self.load_dpi_settings()
        self.load_export_settings()

        # 通用大图查看窗口引用
        self._image_viewer = None

    def closeEvent(self, event: QCloseEvent):
        reply = QMessageBox.question(
            self,
            "退出",
            "确定要退出程序吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            event.accept()
        else:
            event.ignore()

    # ========== 界面、布局、样式 ==========
    def init_ui(self):
        # 1. 创建顶部工具栏 (替代原来的按钮堆)
        self.create_toolbar()

        # 2. 中央主区域 - 使用标签页
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # 创建标签页
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #ddd;
                background-color: #f5f6f7;
            }
            QTabBar::tab {
                background-color: #ecf0f1;
                color: #34495e;
                padding: 8px 20px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background-color: #3498db;
                color: white;
            }
            QTabBar::tab:hover {
                background-color: #bdc3c7;
            }
        """)

        # 标签页1: Single Image (现有功能)
        single_image_tab = self.create_single_image_tab()
        self.tab_widget.addTab(single_image_tab, "Single Image")

        # 标签页2: Dataset (新增功能)
        dataset_tab = self.create_dataset_tab()
        self.tab_widget.addTab(dataset_tab, "Dataset")

        main_layout.addWidget(self.tab_widget)

    def create_single_image_tab(self):
        """创建单图处理标签页"""
        tab_widget = QWidget()
        main_layout = QHBoxLayout(tab_widget)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(15)

        # --- 左侧：图像显示区 ---
        image_frame = QFrame()
        image_frame.setFrameShape(QFrame.Shape.NoFrame)
        image_layout = QVBoxLayout(image_frame)
        image_layout.setContentsMargins(5, 5, 5, 5)

        # A. 增加标题栏
        header_layout = QHBoxLayout()
        preview_title = QLabel("Image Preview")
        preview_title.setStyleSheet("font-weight: bold; color: #34495e; font-size: 14px;")
        header_layout.addWidget(preview_title)
        header_layout.addStretch()
        image_layout.addLayout(header_layout)

        # B. 图片显示 Label
        self.image_label = QLabel(
            "No image available.\n\nClick here or the button in the upper left corner to load an image.")
        self.image_label.setObjectName("ImageLabel")
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        # 设置样式：虚线边框、圆角、鼠标悬停变色
        self.image_label.setStyleSheet("""
            QLabel#ImageLabel {
                border: 2px dashed #bdc3c7;
                border-radius: 10px;
                background-color: #f8f9fa;
                color: #7f8c8d;
            }
            QLabel#ImageLabel:hover {
                background-color: #ecf0f1;
                border: 2px dashed #3498db;
                color: #3498db;
            }
        """)
        self.image_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # C. 核心：开启鼠标交互并绑定函数
        self.image_label.setCursor(Qt.CursorShape.PointingHandCursor)  # 鼠标移上去变小手
        # 自定义鼠标点击，左键放大 / 加载，右键清除
        self.image_label.mousePressEvent = self.on_preview_clicked

        image_layout.addWidget(self.image_label)

        # D. 显示当前样本名（仅在有图片时显示）
        self.image_name_label = QLabel("")
        self.image_name_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.image_name_label.setStyleSheet("color: #7f8c8d; font-size: 11px; padding-top: 4px;")
        self.image_name_label.setVisible(False)
        image_layout.addWidget(self.image_name_label)
        main_layout.addWidget(image_frame, stretch=7)  # 稍微调大一点占比

        # --- 右侧：参数与日志区 ---
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)

        # 参数设置组
        settings_group = QGroupBox("Parameter settings")
        settings_layout = QVBoxLayout()

        # 翻转模式选择
        settings_layout.addWidget(QLabel("Flip mode:"))
        self.flip_combo = QComboBox()
        self.flip_combo.addItems(["None", "Vertical flip", "Horizontal flip", "Vertical + Horizontal"])
        self.flip_combo.setFont(QFont("Microsoft YaHei", 10))
        settings_layout.addWidget(self.flip_combo)

        settings_layout.addStretch()
        settings_group.setLayout(settings_layout)

        # 日志输出组
        log_group = QGroupBox("Process Log")
        log_layout = QVBoxLayout()
        self.output_text = QTextEdit()
        self.output_text.setReadOnly(True)
        self.output_text.setFont(QFont("Consolas", 10))
        log_layout.addWidget(self.output_text)

        # 进度条放在日志区域内
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setVisible(False)
        log_layout.addWidget(self.progress_bar)

        log_group.setLayout(log_layout)

        # 日志区域自定义右键菜单
        self.output_text.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.output_text.customContextMenuRequested.connect(self.show_log_menu)

        right_layout.addWidget(settings_group, stretch=2)
        right_layout.addWidget(log_group, stretch=8)

        main_layout.addWidget(right_panel, stretch=3)  # 右侧占比 30%

        return tab_widget

    def update_dataset_tab_info(self, dataset_path):
        """更新Dataset标签页的信息显示（统计 / 结构 / 预览）"""
        if not os.path.isdir(dataset_path):
            return

        self.dataset_summary_label.setText(
            f"当前数据集: {os.path.basename(dataset_path)}  |  路径: {dataset_path}"
        )

        # 扫描图片
        supported_formats = ('.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.tif', '.webp')
        self.dataset_image_files = []
        for root, dirs, files in os.walk(dataset_path):
            for f in files:
                if f.lower().endswith(supported_formats):
                    self.dataset_image_files.append(os.path.join(root, f))

        # 统计信息
        total_images = len(self.dataset_image_files)
        formats = {}
        folders = {}
        for img_path in self.dataset_image_files:
            ext = os.path.splitext(img_path)[1].lower()
            formats[ext] = formats.get(ext, 0) + 1
            folder = os.path.dirname(img_path)
            folder_name = "Root" if folder == dataset_path else os.path.relpath(folder, dataset_path)
            folders[folder_name] = folders.get(folder_name, 0) + 1

        self.dataset_stats_table.setRowCount(6 + len(formats) + len(folders))
        row = 0
        self.dataset_stats_table.setItem(row, 0, QTableWidgetItem("Dataset Path"))
        self.dataset_stats_table.setItem(row, 1, QTableWidgetItem(dataset_path))
        row += 1
        self.dataset_stats_table.setItem(row, 0, QTableWidgetItem("Total Images"))
        self.dataset_stats_table.setItem(row, 1, QTableWidgetItem(str(total_images)))
        row += 1
        self.dataset_stats_table.setItem(row, 0, QTableWidgetItem("Image Formats"))
        self.dataset_stats_table.setItem(row, 1, QTableWidgetItem(", ".join(formats.keys())))
        row += 1
        self.dataset_stats_table.setItem(row, 0, QTableWidgetItem("Format Distribution"))
        self.dataset_stats_table.setItem(
            row, 1,
            QTableWidgetItem(", ".join([f"{k}: {v}" for k, v in formats.items()]))
        )
        row += 1
        self.dataset_stats_table.setItem(row, 0, QTableWidgetItem("Number of Folders"))
        self.dataset_stats_table.setItem(row, 1, QTableWidgetItem(str(len(folders))))
        row += 1
        self.dataset_stats_table.setItem(row, 0, QTableWidgetItem("Folder Distribution"))
        folder_str = ", ".join([f"{k}: {v}" for k, v in sorted(folders.items(), key=lambda x: x[1], reverse=True)[:10]])
        if len(folders) > 10:
            folder_str += f" ... (and {len(folders) - 10} more)"
        self.dataset_stats_table.setItem(row, 1, QTableWidgetItem(folder_str))

        # 构建树结构
        self.dataset_tree.clear()
        root_item = QTreeWidgetItem(self.dataset_tree)
        root_item.setText(0, os.path.basename(dataset_path))
        root_item.setExpanded(True)
        folder_dict = {}
        for img_path in self.dataset_image_files:
            folder = os.path.dirname(img_path)
            folder_dict.setdefault(folder, []).append(os.path.basename(img_path))

        for folder, files in sorted(folder_dict.items()):
            rel_path = os.path.relpath(folder, dataset_path)
            parent = root_item
            if rel_path != ".":
                parts = rel_path.split(os.sep)
                for part in parts:
                    found = False
                    for i in range(parent.childCount()):
                        if parent.child(i).text(0) == part:
                            parent = parent.child(i)
                            found = True
                            break
                    if not found:
                        new_item = QTreeWidgetItem(parent)
                        new_item.setText(0, part)
                        parent = new_item
            for f in files:
                file_item = QTreeWidgetItem(parent)
                file_item.setText(0, f)

        # 重置分页并刷新预览
        self.dataset_current_page = 0
        self.dataset_show_all = False
        self.update_dataset_preview()

    def update_dataset_preview(self):
        """根据分页状态刷新 Dataset 标签页中的图片预览"""
        # 清空当前格子
        for i in reversed(range(self.dataset_preview_grid.count())):
            w = self.dataset_preview_grid.itemAt(i).widget()
            if w:
                w.setParent(None)

        total = len(self.dataset_image_files)
        if total == 0:
            self.dataset_preview_label.setText("第 0 页 / 共 0 页 (共 0 张)")
            self.dataset_prev_btn.setEnabled(False)
            self.dataset_next_btn.setEnabled(False)
            return

        if self.dataset_show_all:
            start_index, end_index = 0, total
            current_page, total_pages = 1, 1
        else:
            page_size = self.dataset_preview_page_size
            total_pages = (total + page_size - 1) // page_size
            self.dataset_current_page = max(0, min(self.dataset_current_page, total_pages - 1))
            current_page = self.dataset_current_page + 1
            start_index = self.dataset_current_page * page_size
            end_index = min(start_index + page_size, total)

        self.dataset_preview_label.setText(f"第 {current_page} 页 / 共 {total_pages} 页 (共 {total} 张)")

        # 按钮可用性
        if self.dataset_show_all or total_pages <= 1:
            self.dataset_prev_btn.setEnabled(False)
            self.dataset_next_btn.setEnabled(False)
        else:
            self.dataset_prev_btn.setEnabled(current_page > 1)
            self.dataset_next_btn.setEnabled(current_page < total_pages)

        cols = 4
        for idx, img_path in enumerate(self.dataset_image_files[start_index:end_index]):
            try:
                pil_img = Image.open(img_path)
                if pil_img.mode != "RGB":
                    pil_img = pil_img.convert("RGB")
                thumb_size = 150
                pil_img.thumbnail((thumb_size, thumb_size))
                img_array = np.array(pil_img)
                h, w, ch = img_array.shape
                bytes_per_line = ch * w
                q_image = QImage(img_array.data, w, h, bytes_per_line, QImage.Format.Format_RGB888)
                pixmap = QPixmap.fromImage(q_image)

                # 容器：图片 + 文件名
                container = QWidget()
                v = QVBoxLayout(container)
                v.setContentsMargins(0, 0, 0, 0)
                v.setSpacing(3)

                img_label = QLabel()
                img_label.setPixmap(pixmap)
                img_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                img_label.setStyleSheet("""
                    QLabel {
                        border: 2px solid #ddd;
                        border-radius: 5px;
                        padding: 5px;
                        background-color: white;
                    }
                    QLabel:hover {
                        border: 2px solid #3498db;
                    }
                """)
                img_label.setCursor(Qt.CursorShape.PointingHandCursor)
                img_label.mousePressEvent = lambda e, p=img_path: self.open_image_viewer(p)

                name_label = QLabel(os.path.basename(img_path))
                name_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                name_label.setStyleSheet("color: #555; font-size: 10px;")

                v.addWidget(img_label)
                v.addWidget(name_label)

                row = idx // cols
                col = idx % cols
                self.dataset_preview_grid.addWidget(container, row, col)
            except Exception as e:
                self.log(f"Error loading dataset preview: {img_path} - {e}")

    def dataset_prev_page(self):
        """Dataset 预览上一页"""
        if self.dataset_show_all:
            self.dataset_show_all = False
        if self.dataset_current_page > 0:
            self.dataset_current_page -= 1
        self.update_dataset_preview()

    def dataset_next_page(self):
        """Dataset 预览下一页"""
        if self.dataset_show_all:
            self.dataset_show_all = False
            self.dataset_current_page = 0
        total = len(self.dataset_image_files)
        if total > 0:
            total_pages = (total + self.dataset_preview_page_size - 1) // self.dataset_preview_page_size
            if self.dataset_current_page < total_pages - 1:
                self.dataset_current_page += 1
        self.update_dataset_preview()

    def dataset_load_all(self):
        """Dataset 预览加载全部"""
        self.dataset_show_all = True
        self.dataset_current_page = 0
        self.update_dataset_preview()

    def create_dataset_tab(self):
        """创建数据集管理标签页（集成 DatasetInfo 功能，简洁风格）"""
        tab_widget = QWidget()
        main_layout = QVBoxLayout(tab_widget)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(10)

        # 顶部：数据集概要 + 加载按钮
        header_layout = QHBoxLayout()
        self.dataset_summary_label = QLabel("尚未加载数据集")
        self.dataset_summary_label.setStyleSheet("font-weight: bold; color: #34495e; font-size: 13px;")
        header_layout.addWidget(self.dataset_summary_label)
        header_layout.addStretch()

        load_dataset_btn = QPushButton("加载数据集文件夹")
        load_dataset_btn.setIcon(QIcon(r".\resources\assets\images\button\addFolder.png"))
        load_dataset_btn.clicked.connect(self.addFolder)
        header_layout.addWidget(load_dataset_btn)
        main_layout.addLayout(header_layout)

        # 中部：标签页（统计 / 结构 / 预览）
        self.dataset_tabs = QTabWidget()

        # 统计信息
        stats_widget = QWidget()
        stats_layout = QVBoxLayout(stats_widget)
        stats_layout.setContentsMargins(10, 10, 10, 10)
        self.dataset_stats_table = QTableWidget()
        self.dataset_stats_table.setColumnCount(2)
        self.dataset_stats_table.setHorizontalHeaderLabels(["属性", "数值"])
        header = self.dataset_stats_table.horizontalHeader()
        # 默认给第一列一个稍宽的初始宽度，但保持“可拖动调整”
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.dataset_stats_table.setColumnWidth(0, 180)
        header.setStretchLastSection(True)
        self.dataset_stats_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        stats_layout.addWidget(self.dataset_stats_table)
        self.dataset_tabs.addTab(stats_widget, "统计信息")

        # 文件夹结构
        structure_widget = QWidget()
        structure_layout = QVBoxLayout(structure_widget)
        structure_layout.setContentsMargins(10, 10, 10, 10)
        self.dataset_tree = QTreeWidget()
        self.dataset_tree.setHeaderLabel("数据集结构")
        structure_layout.addWidget(self.dataset_tree)
        self.dataset_tabs.addTab(structure_widget, "文件结构")

        # 图片预览
        preview_widget = QWidget()
        preview_layout = QVBoxLayout(preview_widget)
        preview_layout.setContentsMargins(10, 10, 10, 10)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        inner = QWidget()
        self.dataset_preview_grid = QGridLayout(inner)
        self.dataset_preview_grid.setSpacing(10)
        scroll_area.setWidget(inner)
        preview_layout.addWidget(scroll_area)

        # 分页控制
        ctrl_layout = QHBoxLayout()
        self.dataset_preview_label = QLabel("第 0 页 / 共 0 页 (共 0 张)")
        ctrl_layout.addWidget(self.dataset_preview_label)
        ctrl_layout.addStretch()
        self.dataset_prev_btn = QPushButton("上一页")
        self.dataset_next_btn = QPushButton("下一页")
        self.dataset_all_btn = QPushButton("加载全部")
        ctrl_layout.addWidget(self.dataset_prev_btn)
        ctrl_layout.addWidget(self.dataset_next_btn)
        ctrl_layout.addWidget(self.dataset_all_btn)
        preview_layout.addLayout(ctrl_layout)

        self.dataset_tabs.addTab(preview_widget, "图片预览")

        main_layout.addWidget(self.dataset_tabs)

        # 初始化分页相关变量
        self.dataset_image_files = []
        self.dataset_preview_page_size = 20
        self.dataset_current_page = 0
        self.dataset_show_all = False

        # 绑定按钮到槽函数
        self.dataset_prev_btn.clicked.connect(self.dataset_prev_page)
        self.dataset_next_btn.clicked.connect(self.dataset_next_page)
        self.dataset_all_btn.clicked.connect(self.dataset_load_all)

        return tab_widget

    def setup_styles(self):
        """设置现代化的扁平风格 QSS"""
        style_sheet = """
            QMainWindow {
                background-color: #f5f6f7;
            }
            /* 工具栏样式 */
            QToolBar {
                background-color: #ffffff;
                border-bottom: 1px solid #e0e0e0;
                padding: 5px;
                spacing: 10px;
            }
            QToolButton {
                background-color: transparent;
                border: none;
                border-radius: 4px;
                min-width: 35px;      /* 设置统一的最小宽度，整齐排列 */
                padding: 1px;
                margin: 1px;
                color: #333;
                font-family: "Microsoft YaHei";
            }
            QToolButton:hover {
                background-color: #e6f7ff; /* 悬浮时的浅蓝色背景 */
                border: 1px solid #1890ff; /* 悬浮时的细边框 */
            }
            QToolButton:pressed {
                background-color: #bae7ff;
            }

            /* 右侧面板样式 */
            QGroupBox {
                border: 1px solid #dcdfe6;
                border-radius: 6px;
                margin-top: 10px;
                background-color: #ffffff;
                font-family: "Microsoft YaHei";
                font-weight: bold;
                color: #2c3e50;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 5px;
                left: 10px;
            }

            /* 状态显示区 */
            QLabel#ImageLabel {
                background-color: #eef2f5;
                border: 2px dashed #cbd5e0;
                border-radius: 8px;
                color: #a0aec0;
            }

            /* Tooltip 样式 */
            QToolTip {
                border: 1px solid #333;
                background-color: #333;
                color: white;
                padding: 4px;
                opacity: 200;
            }

            QTextEdit {
                border: 1px solid #dcdfe6;
                border-radius: 4px;
                background-color: #fafafa;
                selection-background-color: #1890ff;
            }

            /* 日志区域右键菜单样式*/
            QMenu {
                background-color: #ffffff;
                border: 1px solid #d1d4d8;
            }

            QMenu::item {
                padding: 6px 30px 6px 20px;
                background-color: transparent;
            }

            QMenu::item:selected {
                background-color: #3b8ee0;
                color: #ffffff;
            }

            /* 菜单栏样式：保持白色背景，精致的悬浮感 */
            QMenuBar {
                background-color: #ffffff;
                border-bottom: 1px solid #dcdfe6;
            }

            QMenuBar::item {
                background: transparent;
                padding: 6px 12px;
            }

            /* 强制隐藏菜单中的小图标 */
            QMenu::icon {
                width: 0px;
            }

            QMenuBar::item:selected {
                background-color: #f0f7ff; /* 浅蓝色选中效果 */
                color: #1890ff;
            }

            QMenu {
                background-color: #ffffff;
                border: 1px solid #dcdfe6;
            }

            QMenu::item:selected {
                background-color: #1890ff;
                color: white;
            }
             """
        self.setStyleSheet(style_sheet)

    def create_menus(self):
        menubar = self.menuBar()
        # 统一菜单栏样式：纯白背景，无边框感
        menubar.setStyleSheet("QMenuBar { background-color: #ffffff; border: none; }")

        # --- File 菜单 ---
        file_menu = menubar.addMenu('&File')

        # 统一菜单栏样式：纯白背景，无边框感
        menubar.setStyleSheet("QMenuBar { background-color: #ffffff; border: none; }")

        # 创建纯文字 Action
        open_menu_act = QAction('Open Image...', self)
        open_menu_act.setShortcut('Ctrl+O')
        open_menu_act.triggered.connect(self.load_image)
        file_menu.addAction(open_menu_act)

        save_menu_act = QAction('Save Result', self)
        save_menu_act.setShortcut('Ctrl+S')
        save_menu_act.triggered.connect(self.save_image)
        file_menu.addAction(save_menu_act)

        # --- Settings 菜单 ---
        settings_menu = menubar.addMenu('&Settings')

        # Path Settings
        path_settings_action = QAction('Path Settings', self)
        path_settings_action.setShortcut('Ctrl+P')
        path_settings_action.setToolTip('Configure log, segmentation model, and color recognition model paths')
        path_settings_action.triggered.connect(self.show_path_settings_dialog)
        settings_menu.addAction(path_settings_action)

        # Segmentation Model Settings
        seg_model_action = QAction('Segmentation Model Settings', self)
        seg_model_action.setShortcut('Ctrl+G')
        seg_model_action.setToolTip('Configure U-Net segmentation model path and parameters')
        seg_model_action.triggered.connect(self.show_segmentation_model_settings_dialog)
        settings_menu.addAction(seg_model_action)

        # 添加DPI设置菜单项
        settings_menu.addSeparator()
        dpi_settings_action = QAction('Export DPI Settings', self)
        dpi_settings_action.triggered.connect(self.show_dpi_settings_dialog)
        dpi_settings_action.setShortcut('Ctrl+D')
        settings_menu.addAction(dpi_settings_action)

        # 导出设置菜单项
        export_settings_action = QAction('Export Settings', self)
        export_settings_action.triggered.connect(self.show_export_settings_dialog)
        export_settings_action.setShortcut('Ctrl+E')
        settings_menu.addAction(export_settings_action)

        # 添加分割算法设置
        settings_menu.addSeparator()
        seg_settings_action = QAction('Segmentation Algorithm Settings', self)
        seg_settings_action.triggered.connect(self.show_segmentation_settings_dialog)
        seg_settings_action.setShortcut('Ctrl+M')
        settings_menu.addAction(seg_settings_action)

        # 添加日志设置
        log_settings_action = QAction('Log Settings', self)
        log_settings_action.triggered.connect(self.show_log_settings_dialog)
        log_settings_action.setShortcut('Ctrl+L')
        settings_menu.addAction(log_settings_action)

        # --- Analysis 菜单 ---
        analysis_menu = menubar.addMenu('&Analysis')
        run_menu_act = QAction('Start Recognition', self)
        run_menu_act.setShortcut('F5')
        run_menu_act.triggered.connect(self.start_processing)
        analysis_menu.addAction(run_menu_act)

        # --- Help 菜单 ---
        help_menu = menubar.addMenu('&Help')
        # 修复：使用正确的方法调用方式
        help_menu_action = QAction('About RoCAS', self)
        help_menu_action.setShortcut('F1')
        help_menu_action.triggered.connect(self.show_about_dialog)
        help_menu.addAction(help_menu_action)

    # ========== 添加DPI设置对话框方法 ==========
    def show_dpi_settings_dialog(self):
        """显示DPI设置对话框"""
        dialog = DPISettingsDialog(self)

        # 加载当前设置
        dialog.dpi_group.button(self.dpi).setChecked(True)
        dialog.width_spin.setValue(self.fig_width)
        dialog.height_spin.setValue(self.fig_height)
        dialog.save_individual.setChecked(self.save_individual)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            settings = dialog.get_settings()
            self.save_dpi_settings(
                settings['dpi'],
                settings['width'],
                settings['height'],
                settings['save_individual']
            )
            # 更新当前设置
            self.dpi = settings['dpi']
            self.fig_width = settings['width']
            self.fig_height = settings['height']
            self.save_individual = settings['save_individual']

            QMessageBox.information(
                self,
                "设置已保存",
                f"DPI设置已保存：\nDPI: {self.dpi}\n尺寸: {self.fig_width}×{self.fig_height}英寸\n单独保存: {'是' if self.save_individual else '否'}"
            )

    def create_toolbar(self):
        """创建顶部图标工具栏"""
        toolbar = QToolBar("功能工具栏")
        toolbar.setIconSize(QSize(22, 22))
        toolbar.setMovable(False)  # 禁止拖动
        toolbar.layout().setSpacing(4)  # 图标贴得近
        toolbar.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)

        self.addToolBar(toolbar)

        # 定义统一的字体大小
        btn_font = QFont("Microsoft YaHei", 9)

        # 1. 加载图片
        self.load_action = QAction(QIcon(r".\resources\assets\images\button\load.png"), "Open", self)
        self.load_action.setStatusTip("Tips：从本地文件夹打开一张图片")
        self.load_action.setToolTip("加载图片\n点击选择一张 png 或 jpg 图片进行分析")
        self.load_action.triggered.connect(self.load_image)
        toolbar.addAction(self.load_action)
        toolbar.widgetForAction(self.load_action).setFont(btn_font)

        # 2. 翻转图片
        self.flip_action = QAction(QIcon(r".\resources\assets\images\button\flip.png"), "Flip", self)
        self.flip_action.setStatusTip("Tips:从右侧选中一种翻转方式,之后点击翻转按钮")
        self.flip_action.setToolTip("执行翻转\n根据右侧设置的模式对图片进行翻转")
        self.flip_action.triggered.connect(self.flip_image)
        toolbar.addAction(self.flip_action)

        # 3. 裁剪图片
        self.crop_action = QAction(QIcon(r".\resources\assets\images\button\crop.png"), "Crop", self)
        self.crop_action.setStatusTip("Tips：手动对图片进行裁剪，选择想要识别的目标区域")
        self.crop_action.setToolTip("裁剪图片\n手动框选目标区域")
        self.crop_action.triggered.connect(self.crop_image)
        toolbar.addAction(self.crop_action)
        toolbar.widgetForAction(self.crop_action).setFont(btn_font)

        # 5. 保存结果
        self.save_action = QAction(QIcon(r".\resources\assets\images\button\save.png"), "Save", self)
        self.save_action.setStatusTip("Tips：保存当前显示的图片")
        self.save_action.setToolTip("保存图片\n将处理后的结果保存到本地")
        self.save_action.triggered.connect(self.save_image)
        toolbar.addAction(self.save_action)
        toolbar.widgetForAction(self.save_action).setFont(btn_font)

        toolbar.addSeparator()

        # 6. 自适应分割-添加文件夹
        self.addFolder_action = QAction(QIcon(r".\resources\assets\images\button\addFolder.png"), "Dataset",
                                        self)
        self.addFolder_action.setStatusTip("Tips：选择图像数据集所在文件夹")
        self.addFolder_action.setToolTip("加载所有需要图像分割的图片")
        self.addFolder_action.triggered.connect(self.addFolder)
        toolbar.addAction(self.addFolder_action)
        toolbar.widgetForAction(self.addFolder_action).setFont(btn_font)

        # 7. 一键开启自适应分割
        self.seg_action = QAction(QIcon(r".\resources\assets\images\button\segmentation.png"), "seg", self)
        self.seg_action.setStatusTip("Tips：一键开启自适应分割算法；最后生成原图、目标区域图、分割图以及分割掩码")
        self.seg_action.setToolTip("自适应算法包括：\n\n1.边缘检测算法\n2.阈值分割算法\n3.GrabCut分割算法")
        self.seg_action.triggered.connect(self.seg)
        toolbar.addAction(self.seg_action)
        toolbar.widgetForAction(self.seg_action).setFont(btn_font)

        # 4. 核心功能：识别颜色
        self.process_action = QAction(QIcon(r".\resources\assets\images\button\recognize.png"), "Analyze",
                                      self)
        self.process_action.setStatusTip("Tips：开启颜色智能识别")
        self.process_action.setToolTip("分析图片中的岩石颜色分布")
        self.process_action.triggered.connect(self.start_processing)
        toolbar.addAction(self.process_action)
        toolbar.widgetForAction(self.process_action).setFont(btn_font)

        # 6. 高清截图
        self.screenshot_action = QAction(QIcon(r".\resources\assets\images\button\screenshot-fill.png"),
                                         "Snapshot",
                                         self)
        self.screenshot_action.triggered.connect(self.capture_high_res_screenshot)

        # 将截图功能放在右侧
        empty = QWidget()
        empty.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        toolbar.addWidget(empty)
        toolbar.addAction(self.screenshot_action)
        toolbar.widgetForAction(self.screenshot_action).setFont(btn_font)

    # 设置-导出对话框
    def show_export_settings_dialog(self):
        """显示导出设置对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("导出设置")
        dialog.setModal(True)
        dialog.resize(500, 400)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # 标题区域
        title_label = QLabel("导出设置")
        title_label.setStyleSheet("""
            font-weight: bold;
            font-size: 16px;
            color: #2c3e50;
            padding: 10px;
        """)
        layout.addWidget(title_label)

        # 标题显示选项
        title_group = QGroupBox("图片标题设置")
        title_layout = QVBoxLayout()

        self.title_check = QCheckBox("在导出的图片中显示标题（样本名称）")
        self.title_check.setChecked(self.show_title_in_export)
        self.title_check.setToolTip("勾选后，导出的对比图会显示标题；不勾选则只显示图片")
        title_layout.addWidget(self.title_check)

        title_hint = QLabel("提示：单独保存的原图、分割图、掩码默认不显示标题")
        title_hint.setStyleSheet("color: #7f8c8d; font-size: 11px; padding-left: 20px;")
        title_layout.addWidget(title_hint)

        title_group.setLayout(title_layout)
        layout.addWidget(title_group)

        # 文件名格式选项
        filename_group = QGroupBox("文件名格式设置")
        filename_layout = QVBoxLayout()

        self.filename_time_radio = QRadioButton("时间+样本名 (例如: 20250101_120000_样本名_方法_comparison_style1.png)")
        self.filename_custom_radio = QRadioButton("自定义前缀 (例如: Rock_样本名_方法_comparison_style1.png)")

        if self.filename_format == 'time_name':
            self.filename_time_radio.setChecked(True)
        else:
            self.filename_custom_radio.setChecked(True)

        filename_layout.addWidget(self.filename_time_radio)
        filename_layout.addWidget(self.filename_custom_radio)

        # 自定义前缀输入
        prefix_layout = QHBoxLayout()
        prefix_layout.addWidget(QLabel("自定义前缀:"))
        self.custom_prefix_edit = QLineEdit(self.custom_filename_prefix)
        self.custom_prefix_edit.setPlaceholderText("例如: Rock, Sample, Test")
        self.custom_prefix_edit.setEnabled(self.filename_custom_radio.isChecked())
        prefix_layout.addWidget(self.custom_prefix_edit)

        # 当选择自定义前缀时，启用输入框
        self.filename_custom_radio.toggled.connect(self.custom_prefix_edit.setEnabled)

        filename_layout.addLayout(prefix_layout)
        filename_group.setLayout(filename_layout)
        layout.addWidget(filename_group)

        # 示例显示
        example_label = QLabel("文件名示例:")
        example_label.setStyleSheet("font-weight: bold; color: #34495e;")
        layout.addWidget(example_label)

        self.example_label = QLabel()
        self.example_label.setStyleSheet("""
            background-color: #f8f9fa;
            border: 1px solid #dee2e6;
            border-radius: 4px;
            padding: 8px;
            color: #495057;
            font-family: 'Consolas', monospace;
            font-size: 11px;
        """)
        self.update_filename_example()
        layout.addWidget(self.example_label)

        # 连接信号更新示例
        self.filename_time_radio.toggled.connect(self.update_filename_example)
        self.filename_custom_radio.toggled.connect(self.update_filename_example)
        self.custom_prefix_edit.textChanged.connect(self.update_filename_example)

        layout.addStretch()

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        ok_btn = QPushButton("确定")
        ok_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 8px 20px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            show_title = self.title_check.isChecked()
            filename_format = 'time_name' if self.filename_time_radio.isChecked() else 'custom'
            custom_prefix = self.custom_prefix_edit.text().strip() or "Rock"

            self.save_export_settings(show_title, filename_format, custom_prefix)
            QMessageBox.information(
                self,
                "设置已保存",
                f"导出设置已保存：\n"
                f"显示标题: {'是' if show_title else '否'}\n"
                f"文件名格式: {'时间+样本名' if filename_format == 'time_name' else '自定义前缀'}\n"
                f"自定义前缀: {custom_prefix}"
            )

        # 清理临时控件
        self.title_check = None
        self.filename_time_radio = None
        self.filename_custom_radio = None
        self.custom_prefix_edit = None
        self.example_label = None

    def update_filename_example(self):
        """更新文件名示例"""
        if not hasattr(self, 'example_label') or not self.example_label:
            return

        if self.filename_time_radio.isChecked():
            example = "20250101_120000_样本名_方法名_comparison_style1.png"
        else:
            prefix = self.custom_prefix_edit.text().strip() or "Rock"
            example = f"{prefix}_样本名_方法名_comparison_style1.png"

        self.example_label.setText(example)

    # ========== 自适应分割 ==========
    def addFolder(self):
        """选择图像数据集文件夹"""
        last_dir = self.get_last_directory()
        folder_path = QFileDialog.getExistingDirectory(
            self,
            "选择图像数据集文件夹",
            last_dir
        )

        if not folder_path:
            return

        self.segmentation_folder = folder_path
        self.log(f"已选择数据集文件夹: {folder_path}")
        # 记忆数据集所在目录
        self.set_last_directory(folder_path)

        # 更新Dataset标签页的信息显示（不再弹出独立窗口）
        self.update_dataset_tab_info(folder_path)

    def seg(self):
        """单张图像分割功能（带进度条）"""
        if self.image is None:
            QMessageBox.warning(self, "警告", "请先加载图片！")
            return

        # 选择分割算法
        selected_methods = self.default_segmentation_methods.copy()
        show_reminder = self.show_segmentation_reminder

        if show_reminder:
            dialog = SegmentationMethodDialog(
                self,
                default_methods=self.default_segmentation_methods,
                show_reminder=show_reminder
            )
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return

            selected_methods = dialog.get_selected_methods()
            if not selected_methods:
                QMessageBox.warning(self, "警告", "请至少选择一个分割算法！")
                return

            # 保存"不再提醒"设置
            if not dialog.should_remind():
                self.save_segmentation_settings(selected_methods, False)

        # 创建进度对话框
        progress_dialog = QProgressDialog("正在分割图像，请稍候...", "取消", 0, 100, self)
        progress_dialog.setWindowTitle("图像分割")
        progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
        progress_dialog.setMinimumDuration(0)  # 立即显示
        progress_dialog.setAutoClose(True)
        progress_dialog.setAutoReset(True)
        progress_dialog.setValue(0)
        
        # 为进度对话框添加固定大小（400x120）
        progress_dialog.setFixedSize(400, 120)
        
        # 实现居中显示功能
        def center_dialog():
            screen = QApplication.primaryScreen()
            screen_geometry = screen.geometry()
            dialog_geometry = progress_dialog.geometry()
            x = (screen_geometry.width() - dialog_geometry.width()) // 2
            y = (screen_geometry.height() - dialog_geometry.height()) // 2
            progress_dialog.move(x, y)
        
        # 延迟移动，确保进度框已经显示
        QTimer.singleShot(100, center_dialog)

        try:
            self.log("=" * 60)
            self.log("开始图像分割处理")
            self.log("=" * 60)

            # 保存原始图像
            self.original_image_for_seg = self.image.copy()
            image_name = os.path.basename(self.image_path) if self.image_path else "unknown.jpg"

            # 更新进度
            progress_dialog.setLabelText("正在初始化分割算法...")
            progress_dialog.setValue(10)
            QApplication.processEvents()  # 更新UI

            # 创建进度回调函数
            total_methods = len(selected_methods)
            current_method = 0

            def progress_callback(message):
                """进度回调，更新进度条"""
                nonlocal current_method
                if "完成" in message or "成功" in message:
                    current_method += 1
                    progress = 10 + int((current_method / total_methods) * 80)
                    progress_dialog.setValue(progress)
                    progress_dialog.setLabelText(f"正在处理: {message}")
                else:
                    progress_dialog.setLabelText(f"正在处理: {message}")
                QApplication.processEvents()

            # 执行分割（使用选定的方法）
            results = self.segmenter.segment_by_methods(
                self.image.copy(),
                image_name,
                selected_methods,
                log_callback=progress_callback
            )

            progress_dialog.setValue(90)
            progress_dialog.setLabelText("正在准备预览窗口...")
            QApplication.processEvents()

            if not results:
                progress_dialog.close()
                QMessageBox.warning(self, "分割失败", "所有分割方法都失败了，请检查图像质量。")
                return

            progress_dialog.setValue(100)
            progress_dialog.close()

            self.log(f"\n分割完成，成功方法数: {len(results)}")

            # 显示分割结果预览窗口（修复：只传2个参数）
            self.show_segmentation_results_preview(results, image_name)

        except Exception as e:
            progress_dialog.close()
            self.log(f"分割失败: {str(e)}")
            error_msg = f"分割过程中发生错误:\n{str(e)}\n\n{traceback.format_exc()}"
            QMessageBox.critical(self, "分割错误", error_msg)

    def show_segmentation_result(self, original, segmented, mask, method, image_name):
        """显示分割结果并询问是否导出"""
        # 更新显示为分割结果
        self.display_image_on_label(segmented)
        self.log(f"分割结果已显示在预览区，使用的分割方法: {method}")

        # 创建分割结果预览窗口
        preview_window = SegmentationPreviewWindow(
            original, segmented, mask, method, image_name, self
        )
        preview_window.show()

    def show_segmentation_results_preview(self, results, image_name):
        """显示多算法分割结果预览窗口"""
        preview_window = MultiMethodSegmentationPreviewWindow(
            results, image_name, self
        )
        preview_window.show()

    def export_segmentation_high_dpi(self, original, segmented, mask, method, image_name):
        """导出高DPI分割对比图（使用保存的设置，带进度条）"""
        # 从设置中读取DPI和尺寸
        dpi = self.dpi
        width = self.fig_width
        height = self.fig_height
        save_individual = self.save_individual
        show_title = self.show_title_in_export  # 从设置读取

        # 选择保存目录
        last_dir = self.get_last_directory()
        output_dir = QFileDialog.getExistingDirectory(
            self,
            "选择保存目录",
            last_dir
        )

        if not output_dir:
            return

        # 记忆导出目录
        self.set_last_directory(output_dir)

        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)

        # 创建进度对话框
        progress_dialog = QProgressDialog("正在导出高DPI图片，请稍候...", "取消", 0, 100, self)
        progress_dialog.setWindowTitle("导出图片")
        progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
        progress_dialog.setMinimumDuration(0)
        progress_dialog.setAutoClose(True)
        progress_dialog.setAutoReset(True)
        progress_dialog.setValue(0)

        exported_files = []

        try:
            self.log(f"开始导出高DPI图片 (DPI={dpi})...")
            progress_dialog.setLabelText("正在初始化导出器...")
            progress_dialog.setValue(5)
            QApplication.processEvents()

            # 创建导出器（传入标题显示设置）
            exporter = HighDPIExporter(show_title=show_title)

            # 生成文件名
            base_name = os.path.splitext(os.path.basename(image_name))[0]
            method_clean = method.replace(' ', '_').replace('/', '_')

            # 导出样式1（三列）
            self.log("正在生成样式1对比图...")
            progress_dialog.setLabelText("正在生成样式1对比图...")
            progress_dialog.setValue(20)
            QApplication.processEvents()

            fig1 = exporter.create_comparison_style1(
                original, segmented, mask, method, dpi, width, height, show_title=show_title
            )
            style1_filename = self.generate_export_filename(base_name, method_clean, "comparison_style1")
            style1_path = os.path.join(output_dir, style1_filename)

            # 确保目录存在
            os.makedirs(os.path.dirname(style1_path), exist_ok=True)
            fig1.savefig(style1_path, dpi=dpi, bbox_inches='tight', format='PNG')
            plt.close(fig1)
            exported_files.append(style1_path)
            self.log(f"样式1已保存: {style1_path}")

            # 导出样式2（两列）
            self.log("正在生成样式2对比图...")
            progress_dialog.setLabelText("正在生成样式2对比图...")
            progress_dialog.setValue(50)
            QApplication.processEvents()

            fig2 = exporter.create_comparison_style2(
                original, segmented, method, image_name, dpi, width, height, show_title=show_title
            )
            style2_filename = self.generate_export_filename(base_name, method_clean, "comparison_style2")
            style2_path = os.path.join(output_dir, style2_filename)

            # 确保目录存在
            os.makedirs(os.path.dirname(style2_path), exist_ok=True)
            fig2.savefig(style2_path, dpi=dpi, bbox_inches='tight', format='PNG')
            plt.close(fig2)
            exported_files.append(style2_path)
            self.log(f"样式2已保存: {style2_path}")

            # 单独保存原图、分割图、掩码（不显示标题）
            if save_individual:
                self.log("正在单独保存原图、分割图、掩码...")
                progress_dialog.setLabelText("正在单独保存原图、分割图、掩码...")
                progress_dialog.setValue(70)
                QApplication.processEvents()

                # 生成文件夹名
                if self.filename_format == 'time_name':
                    
                    time_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    individual_dir_name = f"{time_str}_{base_name}_{method_clean}_individual"
                else:
                    individual_dir_name = f"{self.custom_filename_prefix}_{base_name}_{method_clean}_individual"

                individual_dir = os.path.join(output_dir, individual_dir_name)
                os.makedirs(individual_dir, exist_ok=True)

                paths = exporter.save_individual_images(
                    original, segmented, mask, image_name, individual_dir, dpi, show_title=False
                )
                exported_files.extend([paths['original'], paths['segmented'], paths['mask']])
                self.log(f"原图已保存: {paths['original']}")
                self.log(f"分割图已保存: {paths['segmented']}")
                self.log(f"掩码已保存: {paths['mask']}")

            progress_dialog.setValue(100)
            progress_dialog.close()

            # 使用统一的导出成功对话框
            self.show_export_success_dialog(
                exported_files,
                title="导出成功",
                message_prefix="图片已成功导出"
            )

        except Exception as e:
            progress_dialog.close()
            self.log(f"导出失败: {str(e)}")
            
            QMessageBox.critical(
                self,
                "导出错误",
                f"导出过程中发生错误:\n{str(e)}\n\n{traceback.format_exc()}"
            )

    # ========== DPI设置相关方法 ==========
    def load_dpi_settings(self):
        """从QSettings加载DPI设置"""
        settings = QSettings("RockAnalysisTool", "DPI")
        self.dpi = settings.value("dpi", 300, type=int)
        self.fig_width = settings.value("fig_width", 15.0, type=float)
        self.fig_height = settings.value("fig_height", 5.0, type=float)
        self.save_individual = settings.value("save_individual", True, type=bool)

    def save_dpi_settings(self, dpi, width, height, save_individual):
        """保存DPI设置到QSettings"""
        settings = QSettings("RockAnalysisTool", "DPI")
        settings.setValue("dpi", dpi)
        settings.setValue("fig_width", width)
        settings.setValue("fig_height", height)
        settings.setValue("save_individual", save_individual)

    def show_export_success_dialog(self, file_paths, title="导出成功", message_prefix="文件已成功导出"):
        """
        统一的导出成功对话框
        file_paths: 可以是单个文件路径（字符串）或文件路径列表
        """
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle(title)

        # 判断是单个文件还是多个文件
        if isinstance(file_paths, str):
            file_paths = [file_paths]
            is_single = True
        else:
            is_single = len(file_paths) == 1

        # 构建消息文本
        if is_single:
            msg = f"{message_prefix}：\n{os.path.basename(file_paths[0])}"
            msg += f"\n\n保存位置:\n{os.path.dirname(file_paths[0])}"
        else:
            msg = f"{message_prefix} {len(file_paths)} 个文件：\n\n"
            for i, path in enumerate(file_paths[:5], 1):  # 最多显示5个
                msg += f"{i}. {os.path.basename(path)}\n"
            if len(file_paths) > 5:
                msg += f"... 还有 {len(file_paths) - 5} 个文件\n"
            msg += f"\n保存位置:\n{os.path.dirname(file_paths[0])}"

        msg_box.setText(msg)
        msg_box.setIcon(QMessageBox.Icon.Information)

        # 添加按钮
        if is_single:
            # 单个文件：显示"打开图片"和"打开文件夹"
            open_file_btn = msg_box.addButton("打开图片", QMessageBox.ButtonRole.ActionRole)
            open_folder_btn = msg_box.addButton("打开文件夹", QMessageBox.ButtonRole.ActionRole)
        else:
            # 多个文件：只显示"打开文件夹"
            open_folder_btn = msg_box.addButton("打开文件夹", QMessageBox.ButtonRole.ActionRole)
            open_file_btn = None

        msg_box.addButton("确定", QMessageBox.ButtonRole.AcceptRole)

        msg_box.exec()

        # 处理按钮点击
        clicked_btn = msg_box.clickedButton()
        if clicked_btn == open_file_btn and is_single:
            QDesktopServices.openUrl(QUrl.fromLocalFile(file_paths[0]))
        elif clicked_btn == open_folder_btn:
            folder_path = os.path.dirname(file_paths[0])
            if platform.system() == "Windows":
                if is_single:
                    # 打开文件夹并选中文件
                    subprocess.run(['explorer', '/select,', os.path.normpath(file_paths[0])])
                else:
                    # 只打开文件夹
                    subprocess.run(['explorer', os.path.normpath(folder_path)])
            else:
                QDesktopServices.openUrl(QUrl.fromLocalFile(folder_path))

    # ========== 按钮、信息、日志显示等相关操作 ==========
    def load_csv_data(self):
        """程序启动时加载一次 CSV，避免重复读取"""
        csv_path = r".\resources\files\color.csv"
        try:
            if not os.path.exists(csv_path):
                self.log(f"Error: CSV file not found at {csv_path}")
                # 禁用分析按钮防止崩溃
                self.process_action.setEnabled(False)
                return

            df = pd.read_csv(csv_path, encoding='GBK', sep=',')

            # 检查缺失值
            if df.iloc[:, 6:9].isnull().any().any():
                self.log("Warning: CSV contains missing values in vector columns.")

            # 提取数据存为类属性
            self.standard_vectors = df.iloc[:, 6:9].values
            self.color_names = df['岩石颜色'].values
            self.color_codes = df['Munsell颜色代码'].values
            self.log("Color database loaded successfully.")

        except Exception as e:
            self.log(f"Failed to load CSV: {e}")
            QMessageBox.critical(self, "Data Error", f"Cannot load color database:\n{e}")

    def show_log_menu(self, position):
        """创建并显示日志框的右键菜单"""
        menu = QMenu(self)

        clear_action = QAction("Clear Log", self)
        clear_action.triggered.connect(self.output_text.clear)

        copy_action = QAction("Copy Selected", self)
        copy_action.triggered.connect(self.output_text.copy)

        menu.addAction(copy_action)
        menu.addSeparator()
        menu.addAction(clear_action)

        menu.exec(self.output_text.mapToGlobal(position))

    def log(self, message):
        """同时输出详细日志到文本框和底部状态栏，并自动保存"""
        time_str = datetime.datetime.now().strftime("%H:%M:%S")

        # 1. 更新底部状态小字
        if hasattr(self, 'status_label'):
            self.status_label.setText(message[:50])

        # 2. 更新详细历史日志
        if hasattr(self, 'output_text'):
            log_entry = f"[{time_str}] {message}"
            self.output_text.append(log_entry)
            self.output_text.ensureCursorVisible()

            # 添加到缓冲区
            self.log_buffer.append(log_entry)

            # 自动保存（每10条或每5秒）
            if self.log_auto_save and len(self.log_buffer) >= 10:
                self.save_log_file()

        # 定期保存（每5秒）
        if self.log_auto_save and hasattr(self, 'log_timer'):
            if not hasattr(self, 'log_timer'):
                self.log_timer = QTimer()
                self.log_timer.timeout.connect(self.save_log_file)
                self.log_timer.start(5000)  # 5秒

    def save_log_file(self):
        """保存日志文件"""
        if not self.log_auto_save or not self.log_buffer:
            return

        try:
            # 获取计算机名
            computer_name = socket.gethostname()
            date_str = datetime.datetime.now().strftime("%Y%m%d")

            # 生成文件名
            log_filename = f"{computer_name}_{date_str}.log"
            log_filepath = os.path.join(self.log_save_path, log_filename)

            # 追加写入
            with open(log_filepath, 'a', encoding='utf-8') as f:
                for entry in self.log_buffer:
                    f.write(entry + '\n')

            # 清空缓冲区
            self.log_buffer.clear()

        except Exception as e:
            print(f"保存日志失败: {e}")

    def load_export_settings(self):
        """加载导出设置"""
        settings = QSettings("RockAnalysisTool", "Export")
        # 是否显示标题
        self.show_title_in_export = settings.value("show_title", False, type=bool)
        # 文件名格式：'time_name' 或 'custom'
        self.filename_format = settings.value("filename_format", "time_name", type=str)
        # 自定义文件名前缀
        self.custom_filename_prefix = settings.value("custom_prefix", "Rock", type=str)

    def save_export_settings(self, show_title, filename_format, custom_prefix):
        """保存导出设置"""
        settings = QSettings("RockAnalysisTool", "Export")
        settings.setValue("show_title", show_title)
        settings.setValue("filename_format", filename_format)
        settings.setValue("custom_prefix", custom_prefix)
        self.show_title_in_export = show_title
        self.filename_format = filename_format
        self.custom_filename_prefix = custom_prefix

    # 3. 添加分割设置和日志设置相关方法
    def load_segmentation_settings(self):
        """Load segmentation settings"""
        settings = QSettings("RockAnalysisTool", "Segmentation")
        default_methods_str = settings.value("default_methods", "GrabCut Intelligent Segmentation,Color Threshold Segmentation", type=str)
        self.default_segmentation_methods = default_methods_str.split(',') if default_methods_str else ['GrabCut Intelligent Segmentation']
        self.show_segmentation_reminder = settings.value("show_reminder", True, type=bool)
        
        # Load UNet settings
        self.unet_model_path = (settings.value("unet_model_path", "", type=str) or "").strip()
        self.unet_python_path = (settings.value("unet_python_path", "", type=str) or "").strip()
        
        # Load segmentation model settings
        self.seg_model_path = (settings.value("seg_model_path", "", type=str) or "").strip()
        self.seg_model_use_gpu = settings.value("seg_model_use_gpu", True, type=bool)
        
        # Load color model path
        color_settings = QSettings("RockAnalysisTool", "ColorModel")
        self.color_model_path = (color_settings.value("color_model_path", "", type=str) or "").strip()
        
        # Configure segmenter with loaded settings
        if self.seg_model_path:
            self.segmenter.set_dl_config(self.seg_model_path, self.seg_model_use_gpu)

    def save_segmentation_settings(self, methods, show_reminder):
        """Save segmentation settings"""
        settings = QSettings("RockAnalysisTool", "Segmentation")
        settings.setValue("default_methods", ','.join(methods))
        settings.setValue("show_reminder", show_reminder)
        self.default_segmentation_methods = methods
        self.show_segmentation_reminder = show_reminder

    def load_log_settings(self):
        """Load log settings"""
        settings = QSettings("RockAnalysisTool", "Log")
        self.log_auto_save = settings.value("auto_save", False, type=bool)
        self.log_save_path = settings.value("save_path", os.path.join(os.getcwd(), "logs"), type=str)
        # Ensure log directory exists
        os.makedirs(self.log_save_path, exist_ok=True)

    def save_log_settings(self, auto_save, save_path):
        """Save log settings"""
        settings = QSettings("RockAnalysisTool", "Log")
        settings.setValue("auto_save", auto_save)
        settings.setValue("save_path", save_path)
        self.log_auto_save = auto_save
        self.log_save_path = save_path
        # 确保日志目录存在
        os.makedirs(self.log_save_path, exist_ok=True)

    def init_log_auto_save(self):
        """Initialize log auto-save"""
        self.log_buffer = []
        if self.log_auto_save:
            self.log("Log auto-save has been enabled")

    def show_log_settings_dialog(self):
        """Show log settings dialog"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Log Settings")
        dialog.setModal(True)
        dialog.resize(500, 300)

        layout = QVBoxLayout(dialog)

        # Auto-save option
        auto_save_check = QCheckBox("Enable log auto-save")
        auto_save_check.setChecked(self.log_auto_save)
        layout.addWidget(auto_save_check)

        # Save path
        path_layout = QHBoxLayout()
        path_layout.addWidget(QLabel("Log save path:"))
        path_edit = QLineEdit(self.log_save_path)
        path_edit.setReadOnly(True)
        path_layout.addWidget(path_edit)
        browse_btn = QPushButton("Browse...")
        browse_btn.clicked.connect(lambda: self.browse_log_path(path_edit))
        path_layout.addWidget(browse_btn)
        layout.addLayout(path_layout)

        # Description
        info_label = QLabel("Log file naming rule: ComputerName_Date.log\nExample: DESKTOP-ABC123_20250101.log")
        info_label.setStyleSheet("color: #7f8c8d; font-size: 11px;")
        layout.addWidget(info_label)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        ok_btn = QPushButton("OK")
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.save_log_settings(auto_save_check.isChecked(), path_edit.text())
            QMessageBox.information(self, "Settings Saved", "Log settings have been saved")

    def browse_log_path(self, path_edit):
        """Browse log save path"""
        last_dir = path_edit.text() or self.get_last_directory()
        path = QFileDialog.getExistingDirectory(self, "Select log save directory", last_dir)
        if path:
            path_edit.setText(path)
            # 记忆目录
            self.set_last_directory(path)

    def show_segmentation_settings_dialog(self):
        """Show segmentation algorithm settings dialog"""
        dialog = SegmentationMethodDialog(
            self,
            default_methods=self.default_segmentation_methods,
            show_reminder=self.show_segmentation_reminder
        )
        dialog.setWindowTitle("Segmentation Algorithm Default Settings")
        dialog.dont_remind_check.setText("Don't remind to select algorithms (use current selection as default)")

        if dialog.exec() == QDialog.DialogCode.Accepted:
            methods = dialog.get_selected_methods()
            if methods:
                self.save_segmentation_settings(methods, dialog.should_remind())
                QMessageBox.information(self, "Settings Saved", f"Default segmentation algorithms have been set to: {', '.join(methods)}")

    def show_path_settings_dialog(self):
        """Unified path settings: log, segmentation model, color recognition model paths"""
        d = QDialog(self)
        d.setWindowTitle("Path Settings")
        d.setModal(True)
        d.resize(560, 320)
        layout = QVBoxLayout(d)
        layout.setContentsMargins(15, 15, 15, 15)

        info = QLabel("All paths related to models and logs are configured here. Please distinguish:\n"  
                      "• Segmentation model path: for U-Net image segmentation;\n• Color recognition model path: for model recognition (color). Do not mix them up.")
        info.setStyleSheet("color: #7f8c8d; font-size: 11px;")
        info.setWordWrap(True)
        layout.addWidget(info)

        def add_path_row(label_text, line_edit, browse_callback):
            row = QHBoxLayout()
            row.addWidget(QLabel(label_text))
            row.addWidget(line_edit)
            btn = QPushButton("Browse...")
            btn.clicked.connect(browse_callback)
            row.addWidget(btn)
            layout.addLayout(row)

        log_edit = QLineEdit(self.log_save_path)
        log_edit.setPlaceholderText("Log file save directory")

        def browse_log():
            p = QFileDialog.getExistingDirectory(d, "Select Log Save Directory", log_edit.text() or self.get_last_directory())
            if p:
                log_edit.setText(p)
                self.set_last_directory(p)

        add_path_row("Log Save Path:", log_edit, browse_log)

        seg_edit = QLineEdit(getattr(self, 'seg_model_path', '') or '')
        seg_edit.setPlaceholderText("U-Net image segmentation model .pth/.pt (not color recognition model)")

        def browse_seg():
            p, _ = QFileDialog.getOpenFileName(d, "Select Image Segmentation Model (U-Net)", self.get_last_directory(),
                                               "PyTorch (*.pt *.pth);;All (*)")
            if p:
                seg_edit.setText(p)
                self.set_last_directory(p)

        add_path_row("Segmentation Model Path (U-Net):", seg_edit, browse_seg)

        color_edit = QLineEdit(getattr(self, 'color_model_path', '') or '')
        color_edit.setPlaceholderText("Color classification model .pth/.pt (not U-Net segmentation model)")

        def browse_color():
            p, _ = QFileDialog.getOpenFileName(d, "Select Color Recognition Model (Classification)", self.get_last_directory(),
                                               "PyTorch (*.pth *.pt);;All (*)")
            if p:
                color_edit.setText(p)
                self.set_last_directory(p)

        add_path_row("Color Recognition Model Path (Classification):", color_edit, browse_color)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        ok_btn = QPushButton("OK")

        def on_ok():
            log_path = (os.path.abspath(log_edit.text().strip()) if log_edit.text().strip() else os.path.abspath(
                os.path.join(os.getcwd(), "logs")))
            seg_path = seg_edit.text().strip()
            color_path = color_edit.text().strip()
            QSettings("RockAnalysisTool", "Log").setValue("save_path", log_path)
            QSettings("RockAnalysisTool", "Segmentation").setValue("seg_model_path", seg_path)
            QSettings("RockAnalysisTool", "ColorModel").setValue("color_model_path", color_path)
            self.log_save_path = log_path
            os.makedirs(self.log_save_path, exist_ok=True)
            self.seg_model_path = seg_path
            self.color_model_path = color_path
            if hasattr(self, 'segmenter') and self.segmenter:
                self.segmenter.set_dl_config(seg_path, getattr(self, 'seg_model_use_gpu', True))
            if self.log_auto_save:
                self.init_log_auto_save()
            d.accept()
            QMessageBox.information(self, "Settings Saved", "Path settings have been saved.")

        ok_btn.clicked.connect(on_ok)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(d.reject)
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

        d.exec()

    def show_segmentation_model_settings_dialog(self):
        """Show segmentation model settings dialog"""
        d = QDialog(self)
        d.setWindowTitle("Segmentation Model Settings")
        d.setModal(True)
        d.resize(560, 380)
        layout = QVBoxLayout(d)
        layout.setContentsMargins(15, 15, 15, 15)

        info = QLabel("Configure U-Net deep learning segmentation model path and parameters")
        info.setStyleSheet("color: #7f8c8d; font-size: 11px;")
        info.setWordWrap(True)
        layout.addWidget(info)

        # Model path
        path_grp = QGroupBox("Model Path")
        path_layout = QHBoxLayout(path_grp)
        self._seg_model_path_edit = QLineEdit(self.seg_model_path or '')
        self._seg_model_path_edit.setPlaceholderText("U-Net segmentation model .pth/.pt file")
        path_layout.addWidget(self._seg_model_path_edit)
        browse_btn = QPushButton("Browse...")

        def browse_model():
            p, _ = QFileDialog.getOpenFileName(d, "Select U-Net Model File",
                                               self.get_last_directory(),
                                               "PyTorch Model (*.pth *.pt *.bin);;All Files (*)")
            if p:
                self._seg_model_path_edit.setText(p)
                self.set_last_directory(p)

        browse_btn.clicked.connect(browse_model)
        path_layout.addWidget(browse_btn)
        layout.addWidget(path_grp)

        # GPU settings
        gpu_layout = QHBoxLayout()
        gpu_layout.addWidget(QLabel("Use GPU acceleration:"))
        self._seg_use_gpu_check = QCheckBox()
        self._seg_use_gpu_check.setChecked(self.seg_model_use_gpu)
        gpu_layout.addWidget(self._seg_use_gpu_check)
        gpu_layout.addStretch()
        layout.addLayout(gpu_layout)

        # Python path (for external training scripts)
        py_grp = QGroupBox("Python Interpreter Path (for external training)")
        py_layout = QHBoxLayout(py_grp)
        self._unet_python_edit = QLineEdit(self.unet_python_path or '')
        self._unet_python_edit.setPlaceholderText("Python interpreter path, e.g., python or full path")
        py_layout.addWidget(self._unet_python_edit)
        py_browse = QPushButton("Browse...")

        def browse_py():
            p, _ = QFileDialog.getOpenFileName(d, "Select Python Interpreter",
                                               self.get_last_directory(),
                                               "Executable (*.exe);;All Files (*)")
            if p:
                self._unet_python_edit.setText(p)
                self.set_last_directory(p)

        py_browse.clicked.connect(browse_py)
        py_layout.addWidget(py_browse)
        layout.addWidget(py_grp)

        # Training button
        train_btn = QPushButton("Launch U-Net Model Training...")
        train_btn.clicked.connect(lambda: self.show_seg_model_training_dialog())
        layout.addWidget(train_btn)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        ok_btn = QPushButton("OK")

        def on_ok():
            path = self._seg_model_path_edit.text().strip()
            use_gpu = self._seg_use_gpu_check.isChecked()
            py_path = self._unet_python_edit.text().strip()
            settings = QSettings("RockAnalysisTool", "Segmentation")
            settings.setValue("seg_model_path", path)
            settings.setValue("seg_model_use_gpu", use_gpu)
            settings.setValue("unet_python_path", py_path)
            self.seg_model_path = path
            self.seg_model_use_gpu = use_gpu
            self.unet_python_path = py_path
            if self.segmenter:
                self.segmenter.set_dl_config(path, use_gpu)
            d.accept()
            QMessageBox.information(self, "Settings Saved", f"Segmentation model settings have been saved.\nModel path: {path if path else 'Not set'}")

        ok_btn.clicked.connect(on_ok)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(d.reject)
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

        d.exec()

    def show_seg_model_training_dialog(self):
        """Show segmentation model training dialog"""
        # This method would open the training dialog
        # For now, we'll just show a message
        QMessageBox.information(self, "Training Dialog", "U-Net model training dialog would open here.")

    def on_preview_clicked(self, event):
        """Triggered when user clicks on preview area"""
        # Right click: Clear current image (with one-time prompt)
        if event.button() == Qt.MouseButton.RightButton:
            self.on_clear_image_requested()
            return

        # Left click: If there's an image, open image viewer; otherwise trigger image loading
        if event.button() == Qt.MouseButton.LeftButton:
            if self.image is not None and self.image_path:
                self.open_image_viewer(self.image_path)
            else:
                self.load_image()

    def on_clear_image_requested(self):
        """Clear currently loaded image (right click)"""
        if self.image is None:
            return

        settings = QSettings("RockAnalysisTool", "MainWindow")
        ask = settings.value("ask_clear_image", True, type=bool)

        if ask:
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Clear Image")
            msg_box.setIcon(QMessageBox.Icon.Warning)
            msg_box.setText("Right-clicking the preview area will clear the currently loaded image.\n\nDo you want to clear the current image?")
            msg_box.setStandardButtons(
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel
            )
            checkbox = QCheckBox("Don't ask again")
            msg_box.setCheckBox(checkbox)
            result = msg_box.exec()

            if checkbox.isChecked():
                settings.setValue("ask_clear_image", False)

            if result != QMessageBox.StandardButton.Yes:
                return

        # Execute clear operation
        self.image = None
        self.image_path = None
        self.display_pixmap = None
        self.image_label.clear()
        self.image_label.setText(
            "No image available.\n\nClick here or the button in the upper left corner to load an image."
        )

        if hasattr(self, "image_name_label"):
            self.image_name_label.setText("")
            self.image_name_label.setVisible(False)

        # 禁用与当前图片相关的操作
        if hasattr(self, "process_action"):
            self.process_action.setEnabled(False)
        if hasattr(self, "save_action"):
            self.save_action.setEnabled(False)

    def show_about_dialog(self):
        """About dialog"""
        about_text = """
            <h3>RoCAS v2.20</h3>
            <p><b>Rock Color Analysis System</b></p>
            <hr>
            <p>Developed by: <b>Tian Xiaoying</b></p>
            <p>Affiliation: Your Lab / University Name</p>
            <p>Email: <a href='mailto:dachang0220@163.com'>dachang0220@163.com</a></p>
            <br>
            <p><i>A professional tool for rock surface chromaticity recognition 
            and petrological data analysis.</i></p>
        """
        QMessageBox.about(self, "About RoCAS", about_text)

    def add_copyright_footer(self):
        """Add status bar: left corner shows status, right corner shows copyright info"""
        status_bar = self.statusBar()

        # Left side: status information label
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("font-size: 12px; margin-right: 5px;")
        status_bar.addWidget(self.status_label)  # Add status label to left side

        # Right side: copyright information (use addPermanentWidget to keep it on the right)
        copyright_label = QLabel("Copyright © 2025-2026 TianXiaoying.All Rights Reserved.")
        copyright_label.setStyleSheet("font-size: 12px; color: #909399;")
        status_bar.addPermanentWidget(copyright_label)  # Add copyright label to right side

    def show_image_info(self):
        """Show current image information"""
        if self.image is not None:
            height, width = self.image.shape[:2]
            self.log(f"Current image size: {width}x{height}")

            # Calculate recommended grid size
            recommended = min(25, min(width, height) // 10)
            if recommended % 2 == 0:
                recommended += 1
            self.log(f"Recommended grid size: {recommended} (odd num)")
        else:
            self.log("No image loaded")

    def display_image_on_label(self, img_array):
        """Convert OpenCV image to QPixmap and display"""
        if img_array is None:
            return

        height, width, channel = img_array.shape
        bytes_per_line = 3 * width
        # OpenCV uses BGR, Qt uses RGB, need to convert
        img_rgb = cv2.cvtColor(img_array, cv2.COLOR_BGR2RGB)
        q_img = QImage(img_rgb.data, width, height, bytes_per_line, QImage.Format.Format_RGB888)
        pixmap = QPixmap.fromImage(q_img)

        # Scale to fit label while maintaining aspect ratio
        scaled_pixmap = pixmap.scaled(self.image_label.size(), Qt.AspectRatioMode.KeepAspectRatio,
                                      Qt.TransformationMode.SmoothTransformation)
        self.image_label.setPixmap(scaled_pixmap)

    # ========== Color recognition single operations: image loading, cropping, flipping, and recognition ==========
    def load_image(self):
        last_dir = self.get_last_directory()
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Choose rock image",
                last_dir,
                "Image files (*.jpg *.png *.jpeg *.bmp *.tif);;All(*)")

            if not file_path:
                return

            # 保存图片路径
            self.image_path = file_path
            self._last_image_path = file_path

            if not os.path.exists(file_path):
                QMessageBox.critical(self, "Error", "The file does not exist or is missing.")
                return
            file_size = os.path.getsize(file_path)

            if file_size > 10 * 1024 * 1024:
                QMessageBox.warning(self, "warn", "Large files may cause program lag.")

            self.log(f"Try loading images: {file_path}")

            try:
                pil_image = Image.open(file_path)
                self.image = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)
                self.log("Image successfully loaded using PIL.")
                self.set_last_directory(file_path)
                self._display_image_centered()
                self.log(f"Image loaded successfully: {file_path}")
                self.show_image_info()
                # 更新样本名显示
                if hasattr(self, "image_name_label"):
                    self.image_name_label.setText(f"{os.path.basename(file_path)}")
                    self.image_name_label.setVisible(True)

            except Exception as e:
                self.log(f"PIL failed to load; try OpenCV instead.: {str(e)}")
                self.image = cv2_imread(file_path)
                if self.image is None:
                    raise Exception("Unable to load images using OpenCV")

            if self.image is None or self.image.size == 0:
                QMessageBox.critical(self, "错误", "无法解码图片文件!")
                return

            self._display_image_centered()
            self.log(f"Image loaded successfully: {file_path}")
            self.crop_action.setEnabled(True)
            self.process_action.setEnabled(True)
            self.save_action.setEnabled(True)
        except Exception as e:
            self.log(f"An error occurred while loading the image.: {str(e)}")
            QMessageBox.critical(self, "Error", f"An error occurred while loading the image.: {str(e)}")

    def save_image(self):
        """Save processed rock analysis result image"""
        if self.image is None:
            QMessageBox.warning(self, "Warning", "No image to save!")
            return

        try:
            # Handle memory layout issue after cropping
            temp_img = np.ascontiguousarray(self.image)

            h, w, c = temp_img.shape
            bytes_per_line = c * w

            # Convert to QImage
            qimg = QImage(temp_img.data, w, h, bytes_per_line, QImage.Format.Format_BGR888).copy()

            # Call core export function
            self._execute_export(QPixmap.fromImage(qimg), "Rock_Analysis")

        except Exception as e:
            self.log(f"Save failed: {str(e)}")
            QMessageBox.critical(self, "Error", f"Failed to convert image data: {e}")

    def flip_image(self):
        """Flip image function - using English matching"""
        if self.image is None:
            QMessageBox.warning(self, "Warning", "Please load an image first")
            return

        mode = self.flip_combo.currentText()
        if mode == "None":
            QMessageBox.warning(self, "Warning", "Please select a flip mode in the parameters settings on the right")
            self.log("No flip mode selected")
            return
        elif mode == "Vertical flip":
            self.image = cv2.flip(self.image, 0)
        elif mode == "Horizontal flip":
            self.image = cv2.flip(self.image, 1)
        elif mode == "Vertical + Horizontal":
            self.image = cv2.flip(self.image, -1)

        self.display_image_on_label(self.image)
        self.log(f"Executed: {mode}")

    def crop_image(self):
        """
        Use enhanced crop window to manually select crop area and update display image
        """
        if self.image is None:
            QMessageBox.critical(self, "Error", "Please load an image first!")
            return

        # Save original image for cancel operation
        self.original_image = self.image.copy()

        try:
            # Create enhanced crop window
            self.crop_window = EnhancedCropWindow(self)

            # Check image size
            height, width = self.image.shape[:2]
            if height <= 0 or width <= 0:
                raise Exception("Invalid image size")

            # Note: Pass a copy of the original image to avoid modifying original data
            self.crop_window.set_image(self.image.copy())
            self.crop_window.cropConfirmed.connect(self.on_crop_confirmed)
            self.crop_window.cropCancelled.connect(self.on_crop_cancelled)

            # Center display
            screen_geometry = QApplication.primaryScreen().geometry()
            window_geometry = self.crop_window.frameGeometry()
            window_geometry.moveCenter(screen_geometry.center())
            self.crop_window.move(window_geometry.topLeft())

            self.crop_window.show()

        except Exception as e:
            self.log(f"Crop window creation failed: {str(e)}")
            QMessageBox.critical(self, "Error", f"Crop window creation failed: {str(e)}")

    def on_crop_confirmed(self, rect):
        """Crop confirmation callback"""
        try:
            # Execute crop operation
            x, y, w, h = rect.x(), rect.y(), rect.width(), rect.height()

            # Validate crop area
            if w <= 0 or h <= 0:
                raise Exception("Invalid crop area")

            # Crop image - Note: Ensure using original image copy
            cropped_image = self.original_image[y:y + h, x:x + w]

            # Validate cropped image
            if cropped_image.size == 0:
                raise Exception("Cropped image is empty")

            # Update main image
            self.image = cropped_image

            # Update display
            self._display_image_centered()

            # Update log
            self.log(f"Image cropped: Position({x},{y}), Size{w}x{h}")
            self.log(f"Current image size: {self.image.shape[1]}x{self.image.shape[0]}")

            # Enable related functions
            self.process_action.setEnabled(True)
            self.save_action.setEnabled(True)

            # Show success message
            QMessageBox.information(self, "Success", f"Image cropping completed! \nNew size: {w}x{h}")

        except Exception as e:
            # Restore original image if crop failed
            self.image = self.original_image.copy() if hasattr(self, 'original_image') else None
            self.log(f"Crop failed: {str(e)}")
            QMessageBox.critical(self, "Error", f"Crop failed: {str(e)}")
            if self.image is not None:
                self._display_image_centered()

    def on_crop_cancelled(self):
        """Crop cancellation callback"""
        self.log("User cancelled crop operation")
        # Restore original image display
        if hasattr(self, 'original_image'):
            self.image = self.original_image.copy()
            self._display_image_centered()

    # Add exception handling in EnhancedCropWindow class:
    def set_image(self, cv_image):
        """Set OpenCV image"""
        if cv_image is None:
            return

        try:
            # Check image size
            height, width, channel = cv_image.shape
            if height <= 0 or width <= 0:
                raise Exception("Invalid image size")

            # Convert OpenCV image to QImage
            bytes_per_line = 3 * width
            img_rgb = cv2.cvtColor(cv_image, cv2.COLOR_BGR2RGB)
            self.qimage = QImage(img_rgb.data, width, height,
                                 bytes_per_line, QImage.Format.Format_RGB888).copy()

            if self.qimage.isNull():
                raise Exception("QImage creation failed")

            # Set to crop widget
            self.crop_widget.set_image(self.qimage)
            self.crop_widget.original_size = QPoint(width, height)

            # Update zoom slider
            current_scale = int(self.crop_widget.scale_factor * 100)
            self.zoom_slider.setValue(current_scale)
            self.zoom_value_label.setText(f"{current_scale}%")

        except Exception as e:
            print(f"Error setting image: {e}")
            QMessageBox.critical(self, "Error", f"Failed to set image: {str(e)}")

    def on_crop_confirmed(self, rect):
        """裁剪确认回调"""
        try:
            # 执行裁剪操作
            x, y, w, h = rect.x(), rect.y(), rect.width(), rect.height()

            # 验证裁剪区域
            if w <= 0 or h <= 0:
                raise Exception("无效的裁剪区域")

            # 裁剪图像
            self.image = self.image[y:y + h, x:x + w]

            # 验证裁剪后图像
            if self.image.size == 0:
                raise Exception("裁剪后图像为空")

            # 更新显示
            self._display_image_centered()
            self.log(f"图片已裁剪: 位置({x},{y}), 大小{w}x{h}")
            self.log(f"当前图片尺寸: {self.image.shape[1]}x{self.image.shape[0]}")

            self.process_action.setEnabled(True)
            self.save_action.setEnabled(True)

        except Exception as e:
            # 裁剪失败恢复原始图像
            if hasattr(self, 'original_image'):
                self.image = self.original_image
            self.log(f"裁剪失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"裁剪失败: {str(e)}")

    def on_crop_cancelled(self):
        """裁剪取消回调"""
        self.log("用户取消了裁剪操作")

    def start_processing(self):
        """开始处理图像逻辑"""
        if self.image is None:
            QMessageBox.critical(self, "错误", "请先加载图片！")
            return

        # 获取当前图片尺寸
        height, width = self.image.shape[:2]

        # 计算合适的网格大小范围
        min_grid = 5
        max_grid = min(width, height) // 4  # 网格大小不超过图片尺寸的1/4
        if max_grid < min_grid:
            max_grid = min_grid + 10

        # 根据图片尺寸动态设置默认值
        default_grid = min(25, max_grid)

        grid_size, ok = QInputDialog.getInt(
            self, "输入网格大小",
            f"请输入网格大小, 如 3、7、9、15 (当前图片大小：{width}x{height}={width * height})",
            value=default_grid, min=min_grid, max=max_grid, step=1
        )
        if not ok:  # 用户点了取消
            return

        # 确保网格大小为奇数
        if grid_size % 2 == 0:
            grid_size += 1
            self.log(f"网格大小已调整为奇数: {grid_size}")
            QMessageBox.information(self, "提示", f"网格大小已调整为奇数: {grid_size}")

        self.log(f"开始分析，网格大小: {grid_size}px, 图片尺寸: {width}x{height}")
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)

        # 启动后台线程
        self.worker = ColorAnalysisWorker(
            self.image,
            grid_size,
            self.standard_vectors,
            self.color_names,
            self.color_codes
        )

        # 连接信号
        self.worker.progress_updated.connect(self.progress_bar.setValue)
        self.worker.error_occurred.connect(lambda err: (
            QMessageBox.critical(self, "错误", err),
            self.progress_bar.setVisible(False)
        ))
        self.worker.result_ready.connect(self.on_analysis_finished)

        self.worker.start()

    def capture_high_res_screenshot(self):
        """捕捉软件全貌截图（含边框与标题栏）"""
        # 捕捉包含标题栏的完整窗口几何体
        rect = self.frameGeometry()
        screenshot = self.screen().grabWindow(0, rect.x(), rect.y(), rect.width(), rect.height())

        # 传入截图，前缀设为 RoCAS_UI
        self._execute_export(screenshot, "RoCAS_Full_View")

    # ========== 导出图片、导出表格等、识别结果、分析报告等其他功能 ==========
    def _execute_export(self, pixmap_to_save, default_prefix="RoCAS"):
        """
        内部核心导出逻辑：
        整合了 TIFF-300DPI、交互对话框、自动命名、中文兼容
        """
        # 1. 自动生成带时间戳的文件名
        now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        join_path = os.path.join(self.get_last_directory(), f"{now_str}_{default_prefix}.png")
        # 2. 弹出保存对话框
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Image",
            join_path,
            "PNG Files (*.png);;TIFF Files (*.tif *.tiff)",
        )
        if not save_path:
            return

        try:

            # 3. 将 QPixmap 转换为 PIL 对象进行处理 (兼容所有格式)
            buffer = QBuffer()
            buffer.open(QIODevice.OpenModeFlag.ReadWrite)
            pixmap_to_save.save(buffer, "PNG")
            pil_img = Image.open(io.BytesIO(buffer.data()))

            # 4. 执行保存逻辑
            if save_path.lower().endswith(('.tif', '.tiff')):
                # TIFF 模式：无损压缩 + 300 DPI
                pil_img.save(save_path, format='TIFF', dpi=(300.0, 300.0), compression="tiff_lzw")
                log_msg = f"TIFF (300DPI) Exported: {save_path}"
            else:
                # PNG 模式：原生保存
                pil_img.save(save_path, format='PNG')
                log_msg = f"PNG Exported: {save_path}"

            # 保存图片保存目录
            self.set_last_directory(os.path.dirname(save_path))

            # 5. 更新日志
            if hasattr(self, 'output_text'):
                self.log(log_msg)

            # 6. 弹出交互式对话框
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Export Success")
            msg_box.setText(f"Successfully saved to:\n{os.path.basename(save_path)}")
            msg_box.setIcon(QMessageBox.Icon.Information)

            open_file_btn = msg_box.addButton("Open Image", QMessageBox.ButtonRole.ActionRole)
            open_folder_btn = msg_box.addButton("Open Folder", QMessageBox.ButtonRole.ActionRole)
            msg_box.addButton("OK", QMessageBox.ButtonRole.AcceptRole)

            msg_box.exec()

            # 7. 交互逻辑处理
            clicked_btn = msg_box.clickedButton()
            if clicked_btn == open_file_btn:
                QDesktopServices.openUrl(QUrl.fromLocalFile(save_path))
            elif clicked_btn == open_folder_btn:
                if platform.system() == "Windows":
                    # 打开文件夹并选中该文件
                    subprocess.run(['explorer', '/select,', os.path.normpath(save_path)])
                else:
                    QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(save_path)))

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"An error occurred:\n{str(e)}")

    def on_analysis_finished(self, report_text, stats_list, total_count):
        """分析完成，隐藏进度条并弹出独立的结果分析窗口"""
        self.progress_bar.setVisible(False)
        self.log(f"分析完成。总采样点: {total_count}")

        # 保存数据到实例变量中
        self.stats_data = stats_list
        self.total_samples = total_count

        # 获取文件名
        file_name = "未知样本"
        if hasattr(self, 'image_path') and self.image_path:
            file_name = os.path.basename(self.image_path)
        elif hasattr(self, '_last_image_path') and self._last_image_path:
            file_name = os.path.basename(self._last_image_path)

        try:
            # 如果已存在结果窗口，先关闭它
            if hasattr(self, 'result_window') and self.result_window:
                self.result_window.close()
                self.result_window = None

            self.result_window = AnalysisResultWindow(stats_list, total_count, file_name, self)
            self.result_window.show()
        except Exception as e:
            self.log(f"弹窗失败: {str(e)}")
            
            print(f"弹窗失败详情: {traceback.format_exc()}")
            QMessageBox.warning(self, "提示", "分析已完成，但结果窗口显示失败，请检查日志。")

    def get_last_directory(self):
        """获取上次使用的目录"""
        settings = QSettings("RockAnalysisTool", "MainWindow")
        return settings.value("last_directory", "", type=str)

    def set_last_directory(self, file_path):
        """设置上次使用的目录"""
        # 既支持文件路径也支持文件夹路径
        if file_path and os.path.isdir(file_path):
            directory = file_path
        else:
            directory = os.path.dirname(file_path)
        settings = QSettings("RockAnalysisTool", "MainWindow")
        settings.setValue("last_directory", directory)

    def _display_image_centered(self):
        """图片居中显示"""
        if self.image is not None:
            canvas_width = self.image_label.width()
            canvas_height = self.image_label.height()
            self.display_pixmap = self.convert_to_display(self.image, canvas_width, canvas_height)
            self.image_label.setPixmap(self.display_pixmap)
            self.image_label.setText("")

    def open_image_viewer(self, image_path):
        """打开可缩放的大图查看窗口"""
        if not os.path.exists(image_path):
            return
        self._image_viewer = ImageViewerWindow(image_path, self)
        self._image_viewer.show()

    def convert_to_display(self, img, canvas_width, canvas_height):
        try:
            if canvas_width <= 0 or canvas_height <= 0:
                canvas_width = 600
                canvas_height = 500

            img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            img_height, img_width, channel = img_rgb.shape

            if img_width <= 0 or img_height <= 0:
                raise Exception("无效的图像尺寸")

            original_ratio = img_width / img_height
            canvas_ratio = canvas_width / canvas_height

            if original_ratio > canvas_ratio:
                new_width = canvas_width
                new_height = max(1, int(canvas_width / original_ratio))
            else:
                new_height = canvas_height
                new_width = max(1, int(canvas_height * original_ratio))

            resized_img = cv2.resize(img_rgb, (new_width, new_height), interpolation=cv2.INTER_AREA)

            q_image = QImage(
                resized_img.data,
                new_width,
                new_height,
                new_width * channel,
                QImage.Format.Format_RGB888
            )

            return QPixmap.fromImage(q_image)
        except Exception as e:
            self.log(f"图像转换错误: {str(e)}")
            return QPixmap()

    def export_single_segmentation_result(self, original, segmented, mask, method, image_name, output_dir):
        """导出单个分割结果（支持指定目录，返回文件列表）"""
        dpi = self.dpi
        width = self.fig_width
        height = self.fig_height
        save_individual = self.save_individual
        show_title = self.show_title_in_export

        exported_files = []

        try:
            # 确保输出目录存在
            os.makedirs(output_dir, exist_ok=True)

            exporter = HighDPIExporter(show_title=show_title)
            base_name = os.path.splitext(os.path.basename(image_name))[0]
            method_clean = method.replace(' ', '_').replace('/', '_')

            # 导出样式1
            fig1 = exporter.create_comparison_style1(
                original, segmented, mask, method, dpi, width, height, show_title=show_title
            )
            style1_filename = self.generate_export_filename(base_name, method_clean, "comparison_style1")
            style1_path = os.path.join(output_dir, style1_filename)
            os.makedirs(os.path.dirname(style1_path), exist_ok=True)
            fig1.savefig(style1_path, dpi=dpi, bbox_inches='tight', format='PNG')
            plt.close(fig1)
            exported_files.append(style1_path)

            # 导出样式2
            fig2 = exporter.create_comparison_style2(
                original, segmented, method, image_name, dpi, width, height, show_title=show_title
            )
            style2_filename = self.generate_export_filename(base_name, method_clean, "comparison_style2")
            style2_path = os.path.join(output_dir, style2_filename)
            os.makedirs(os.path.dirname(style2_path), exist_ok=True)
            fig2.savefig(style2_path, dpi=dpi, bbox_inches='tight', format='PNG')
            plt.close(fig2)
            exported_files.append(style2_path)

            # 单独保存
            if save_individual:
                if self.filename_format == 'time_name':
                    
                    time_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    individual_dir_name = f"{time_str}_{base_name}_{method_clean}_individual"
                else:
                    individual_dir_name = f"{self.custom_filename_prefix}_{base_name}_{method_clean}_individual"

                individual_dir = os.path.join(output_dir, individual_dir_name)
                os.makedirs(individual_dir, exist_ok=True)
                paths = exporter.save_individual_images(
                    original, segmented, mask, image_name, individual_dir, dpi, show_title=False
                )
                exported_files.extend([paths['original'], paths['segmented'], paths['mask']])

            return exported_files

        except Exception as e:
            self.log(f"导出失败: {str(e)}")
            
            self.log(traceback.format_exc())
            return []

    def generate_export_filename(self, base_name, method_clean, suffix, extension='.png'):
        """根据设置生成导出文件名"""
        if self.filename_format == 'time_name':
            # 时间+样本名格式
            
            time_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{time_str}_{base_name}_{method_clean}_{suffix}{extension}"
        else:
            # 自定义前缀格式
            filename = f"{self.custom_filename_prefix}_{base_name}_{method_clean}_{suffix}{extension}"
        return filename



# ========== DPI调整并导出类 ==========
class HighDPIExporter:
    """高DPI图片导出工具类 - 包含所有图片导出功能"""

    def __init__(self, show_title=False, sci_colors=None):
        self.supported_dpi = [300, 600, 1200]
        self.supported_formats = ['PNG', 'TIFF', 'PDF', 'JPEG']
        self.show_title = show_title
        self.sci_colors = sci_colors  # SCI 配色，用于图表等

    def _apply_sci_theme(self):
        """应用 SCI 配色到 matplotlib"""
        if self.sci_colors:
            try:
                from matplotlib import cycler
                plt.rcParams['axes.prop_cycle'] = cycler(color=self.sci_colors)
            except Exception:
                pass

    def create_comparison_style1(self, original, segmented, mask, method_name,
                                 dpi=300, fig_width=15, fig_height=5, show_title=True):
        """创建样式1对比图：原图+分割图+掩码（三列）"""
        self._apply_sci_theme()
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial']
        plt.rcParams['axes.unicode_minus'] = False

        fig, axes = plt.subplots(1, 3, figsize=(fig_width, fig_height), dpi=dpi)

        # 原图
        axes[0].imshow(cv2.cvtColor(original, cv2.COLOR_BGR2RGB))
        if show_title:
            axes[0].set_title('原始图像', fontsize=14, fontweight='bold')
        axes[0].axis('off')

        # 分割图
        axes[1].imshow(cv2.cvtColor(segmented, cv2.COLOR_BGR2RGB))
        if show_title:
            axes[1].set_title(f'分割结果 ({method_name})', fontsize=14, fontweight='bold')
        axes[1].axis('off')

        # 掩码
        if len(mask.shape) == 2:
            axes[2].imshow(mask, cmap='gray')
        else:
            axes[2].imshow(mask)
        if show_title:
            axes[2].set_title('分割掩码', fontsize=14, fontweight='bold')
        axes[2].axis('off')

        plt.tight_layout(pad=2.0)
        return fig

    def create_comparison_style2(self, original, segmented, method_name, image_name,
                                 dpi=300, fig_width=12, fig_height=6, show_title=True):
        """创建样式2对比图：原图+分割图（两列）"""
        self._apply_sci_theme()
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial']
        plt.rcParams['axes.unicode_minus'] = False

        fig, axes = plt.subplots(1, 2, figsize=(fig_width, fig_height), dpi=dpi)

        # 原图
        axes[0].imshow(cv2.cvtColor(original, cv2.COLOR_BGR2RGB))
        if show_title:
            axes[0].set_title(f'原始图像: {os.path.basename(image_name)}', fontsize=14, fontweight='bold')
        axes[0].axis('off')

        # 分割图
        axes[1].imshow(cv2.cvtColor(segmented, cv2.COLOR_BGR2RGB))
        if show_title:
            axes[1].set_title(f'分割结果 - {method_name}', fontsize=14, fontweight='bold')
        axes[1].axis('off')

        plt.tight_layout(pad=2.0)
        return fig

    def create_comparison_style3(self, original, segmented, mask, method_name, image_name,
                                 dpi=300, fig_width=18, fig_height=6, show_title=True):
        """创建样式3对比图：原图+分割图+掩码+透明主体图（四列）"""
        self._apply_sci_theme()
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial']
        plt.rcParams['axes.unicode_minus'] = False

        fig, axes = plt.subplots(1, 4, figsize=(fig_width, fig_height), dpi=dpi)

        # 原图
        axes[0].imshow(cv2.cvtColor(original, cv2.COLOR_BGR2RGB))
        if show_title:
            axes[0].set_title('原始图像', fontsize=14, fontweight='bold')
        axes[0].axis('off')

        # 分割图
        axes[1].imshow(cv2.cvtColor(segmented, cv2.COLOR_BGR2RGB))
        if show_title:
            axes[1].set_title(f'分割结果 ({method_name})', fontsize=14, fontweight='bold')
        axes[1].axis('off')

        # 掩码
        if len(mask.shape) == 2:
            axes[2].imshow(mask, cmap='gray')
        else:
            axes[2].imshow(mask)
        if show_title:
            axes[2].set_title('分割掩码', fontsize=14, fontweight='bold')
        axes[2].axis('off')

        # 透明主体图（使用掩码作为alpha通道）
        h, w = segmented.shape[:2]
        if len(mask.shape) == 3:
            alpha = cv2.cvtColor(mask, cv2.COLOR_BGR2GRAY)
        else:
            alpha = mask
        if alpha.max() <= 1:
            alpha = (alpha * 255).astype(np.uint8)
        if alpha.shape[0] != h or alpha.shape[1] != w:
            alpha = cv2.resize(alpha, (w, h), interpolation=cv2.INTER_NEAREST)

        rgb = cv2.cvtColor(segmented, cv2.COLOR_BGR2RGB)
        alpha_channel = alpha.astype(np.float32) / 255.0
        rgba = np.dstack([rgb, alpha_channel])

        axes[3].imshow(rgba)
        if show_title:
            axes[3].set_title('透明主体', fontsize=14, fontweight='bold')
        axes[3].axis('off')

        plt.tight_layout(pad=2.0)
        return fig

    def save_individual_images(self, original, segmented, mask, image_name,
                               output_dir, dpi=300, show_title=False, export_fmt='png'):
        """单独保存原图、分割图、掩码到指定文件夹（不显示标题），支持 PNG/JPEG/TIFF/PDF"""
        ext = '.pdf' if export_fmt == 'pdf' else (
            '.tif' if export_fmt == 'tif' else ('.jpg' if export_fmt == 'jpg' else '.png'))
        savefig_fmt = {'png': 'png', 'jpg': 'jpg', 'tif': 'tiff', 'pdf': 'pdf'}.get(export_fmt, 'png')

        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)

        base_name = os.path.splitext(os.path.basename(image_name))[0]

        # 保存原图（无标题）
        original_path = os.path.join(output_dir, f"{base_name}_original{ext}")
        fig_orig = plt.figure(figsize=(10, 10), dpi=dpi)
        plt.imshow(cv2.cvtColor(original, cv2.COLOR_BGR2RGB))
        plt.axis('off')
        if show_title:
            plt.title(f'原始图像: {base_name}', fontsize=16, fontweight='bold', pad=20)
        plt.tight_layout(pad=0)
        fig_orig.savefig(original_path, dpi=dpi, bbox_inches='tight', format=savefig_fmt, pad_inches=0)
        plt.close(fig_orig)

        # 保存分割图（无标题）- 带黑色背景的岩石主体图
        segmented_path = os.path.join(output_dir, f"{base_name}_segmented{ext}")
        fig_seg = plt.figure(figsize=(10, 10), dpi=dpi)
        plt.imshow(cv2.cvtColor(segmented, cv2.COLOR_BGR2RGB))
        plt.axis('off')
        if show_title:
            plt.title(f'分割结果: {base_name}', fontsize=16, fontweight='bold', pad=20)
        plt.tight_layout(pad=0)
        fig_seg.savefig(segmented_path, dpi=dpi, bbox_inches='tight', format=savefig_fmt, pad_inches=0)
        plt.close(fig_seg)

        # 保存掩码（无标题）
        mask_path = os.path.join(output_dir, f"{base_name}_mask{ext}")
        fig_mask = plt.figure(figsize=(10, 10), dpi=dpi)
        if len(mask.shape) == 2:
            plt.imshow(mask, cmap='gray')
        else:
            plt.imshow(mask)
        plt.axis('off')
        if show_title:
            plt.title(f'分割掩码: {base_name}', fontsize=16, fontweight='bold', pad=20)
        plt.tight_layout(pad=0)
        fig_mask.savefig(mask_path, dpi=dpi, bbox_inches='tight', format=savefig_fmt, pad_inches=0)
        plt.close(fig_mask)

        # 保存透明背景的岩石主体图（掩码区域保留，其他区域透明；仅PNG支持透明）
        subject_transparent_path = None
        try:
            h, w = original.shape[:2]
            # 确保掩码是单通道的
            if len(mask.shape) == 3:
                alpha = cv2.cvtColor(mask, cv2.COLOR_BGR2GRAY)
            else:
                alpha = mask.copy()

            # 确保alpha值在0-255范围内
            if alpha.max() <= 1:
                alpha = (alpha * 255).astype(np.uint8)
            if alpha.shape[0] != h or alpha.shape[1] != w:
                alpha = cv2.resize(alpha, (w, h), interpolation=cv2.INTER_NEAREST)

            # 关键修复：反转mask，让岩石区域变成255（不透明）
            alpha = cv2.bitwise_not(alpha)

            # 将原图转为RGB（关键修改：使用original而不是segmented）
            rgb = cv2.cvtColor(original, cv2.COLOR_BGR2RGB)

            # 创建透明度通道，掩码值为255的地方为不透明，其他地方为透明
            alpha_channel = alpha.astype(np.float32) / 255.0  # 归一化到0-1

            # 合并RGB和Alpha通道
            rgba = np.dstack([rgb, (alpha_channel * 255).astype(np.uint8)])
            subject_transparent_path = os.path.join(output_dir, f"{base_name}_subject_transparent.png")
            Image.fromarray(rgba, 'RGBA').save(subject_transparent_path)
        except Exception as e:
            print(f"保存透明背景图像时出错: {e}")
            pass

        # 保存黑色背景的岩石主体图（掩码区域保留，其他区域为黑色）
        subject_black_bg_path = None
        try:
            h, w = original.shape[:2]
            # 确保掩码是单通道的
            if len(mask.shape) == 3:
                alpha = cv2.cvtColor(mask, cv2.COLOR_BGR2GRAY)
            else:
                alpha = mask.copy()

            # 确保alpha值在0-255范围内
            if alpha.max() <= 1:
                alpha = (alpha * 255).astype(np.uint8)
            if alpha.shape[0] != h or alpha.shape[1] != w:
                alpha = cv2.resize(alpha, (w, h), interpolation=cv2.INTER_NEAREST)

            # 关键修复：反转mask，让岩石区域变成255
            alpha = cv2.bitwise_not(alpha)

            # 将原图转为RGB（关键修改：使用original而不是segmented）
            rgb = cv2.cvtColor(original, cv2.COLOR_BGR2RGB)

            # 创建全黑背景
            black_bg = np.zeros_like(rgb, dtype=np.uint8)

            # 创建alpha遮罩（归一化到0-1）
            alpha_mask = alpha.astype(np.float32) / 255.0

            # 将掩码区域的内容复制到黑色背景上
            subject_with_black_bg = np.where(alpha_mask[..., np.newaxis] > 0.5, rgb, black_bg)

            subject_black_bg_path = os.path.join(output_dir, f"{base_name}_subject_black_bg.png")
            Image.fromarray(subject_with_black_bg, 'RGB').save(subject_black_bg_path)
        except Exception as e:
            print(f"保存黑色背景图像时出错: {e}")
            pass

        return {
            'original': original_path,
            'segmented': segmented_path,
            'mask': mask_path,
            'subject_transparent': subject_transparent_path,
            'subject_black_bg': subject_black_bg_path
        }

    def create_subject_only_image(self, original, segmented, mask, output_dir, base_name):
        """创建仅包含岩石主体的图像（去除背景）- 背景透明的岩石主体图"""
        try:
            # 确保掩码是单通道的
            if len(mask.shape) == 3:
                alpha = cv2.cvtColor(mask, cv2.COLOR_BGR2GRAY)
            else:
                alpha = mask.copy()

            # 确保alpha值在0-255范围内
            if alpha.max() <= 1:
                alpha = (alpha * 255).astype(np.uint8)

            # 将原图转为RGB（关键修改：使用original而不是segmented）
            # rgb = cv2.cvtColor(original, cv2.COLOR_BGR2RGB)
            # 反转mask，让岩石区域变成255
            alpha = cv2.bitwise_not(alpha)
            # 将原图转为RGB（关键修改：使用original而不是segmented）
            rgb = cv2.cvtColor(original, cv2.COLOR_BGR2RGB)

            # 使用掩码作为alpha通道，岩石区域为不透明，背景区域为透明
            alpha_channel = alpha.astype(np.float32) / 255.0
            rgba = np.dstack([rgb, (alpha_channel * 255).astype(np.uint8)])
            subject_only_path = os.path.join(output_dir, f"{base_name}_subject_only.png")
            Image.fromarray(rgba, 'RGBA').save(subject_only_path)

            return subject_only_path
        except Exception as e:
            print(f"创建透明主体图像时出错: {e}")
            return None

    def create_subject_with_black_background(self, original, segmented, mask, output_dir, base_name):
        """创建黑色背景的岩石主体图像（去除背景）- 背景黑色的岩石主体图"""
        try:
            # 确保掩码是单通道的
            if len(mask.shape) == 3:
                alpha = cv2.cvtColor(mask, cv2.COLOR_BGR2GRAY)
            else:
                alpha = mask.copy()

            # 确保alpha值在0-255范围内
            if alpha.max() <= 1:
                alpha = (alpha * 255).astype(np.uint8)

            # 将原图转为RGB（关键修改：使用original而不是segmented）
            # rgb = cv2.cvtColor(original, cv2.COLOR_BGR2RGB)
            # 反转mask，让岩石区域变成255
            alpha = cv2.bitwise_not(alpha)
            # 将原图转为RGB（关键修改：使用original而不是segmented）
            rgb = cv2.cvtColor(original, cv2.COLOR_BGR2RGB)
            # 创建全黑背景
            black_bg = np.zeros_like(rgb, dtype=np.uint8)

            # 创建alpha遮罩（归一化到0-1）
            alpha_mask = alpha.astype(np.float32) / 255.0

            # 将掩码区域的内容复制到黑色背景上
            subject_with_black_bg = np.where(alpha_mask[..., np.newaxis] > 0.5, rgb, black_bg)

            subject_black_bg_path = os.path.join(output_dir, f"{base_name}_subject_black_bg.png")
            Image.fromarray(subject_with_black_bg, 'RGB').save(subject_black_bg_path)

            return subject_black_bg_path
        except Exception as e:
            print(f"创建黑色背景主体图像时出错: {e}")
            return None

    def create_mask_overlay(self, original, mask, output_dir, base_name, alpha=0.5):
        """创建掩码覆盖在原图上的图像"""
        try:
            # 转换为RGB格式
            rgb = cv2.cvtColor(original, cv2.COLOR_BGR2RGB)

            # 确保掩码是单通道的
            if len(mask.shape) == 3:
                gray_mask = cv2.cvtColor(mask, cv2.COLOR_BGR2GRAY)
            else:
                gray_mask = mask.copy()

            # 创建彩色掩码（红色）
            colored_mask = np.zeros_like(rgb)
            colored_mask[:, :, 0] = gray_mask  # R通道为红色

            # 创建覆盖图像
            overlay = rgb.copy()
            overlay[gray_mask > 0] = (overlay[gray_mask > 0] * (1 - alpha) +
                                      colored_mask[gray_mask > 0] * alpha).astype(np.uint8)

            overlay_path = os.path.join(output_dir, f"{base_name}_mask_overlay.png")
            cv2.imwrite(overlay_path, cv2.cvtColor(overlay, cv2.COLOR_RGB2BGR))

            return overlay_path
        except Exception as e:
            print(f"创建掩码覆盖图像时出错: {e}")
            return None

    def create_edge_highlight_image(self, original, mask, output_dir, base_name):
        """创建边缘高亮图像"""
        try:
            # 转换为RGB格式
            rgb = cv2.cvtColor(original, cv2.COLOR_BGR2RGB)

            # 确保掩码是单通道的
            if len(mask.shape) == 3:
                gray_mask = cv2.cvtColor(mask, cv2.COLOR_BGR2GRAY)
            else:
                gray_mask = mask.copy()

            # 找到轮廓
            contours, _ = cv2.findContours(gray_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

            # 在图像上绘制轮廓
            result = rgb.copy()
            cv2.drawContours(result, contours, -1, (0, 255, 0), thickness=2)  # 绿色轮廓

            edge_highlight_path = os.path.join(output_dir, f"{base_name}_edge_highlight.png")
            cv2.imwrite(edge_highlight_path, cv2.cvtColor(result, cv2.COLOR_RGB2BGR))

            return edge_highlight_path
        except Exception as e:
            print(f"创建边缘高亮图像时出错: {e}")
            return None

    def create_binary_mask_image(self, mask, output_dir, base_name):
        """创建二值掩码图像"""
        try:
            binary_mask_path = os.path.join(output_dir, f"{base_name}_binary_mask.png")
            cv2.imwrite(binary_mask_path, mask)
            return binary_mask_path
        except Exception as e:
            print(f"创建二值掩码图像时出错: {e}")
            return None

    def create_heatmap_overlay(self, original, mask, output_dir, base_name):
        """创建热力图叠加图像"""
        try:
            # 转换为RGB格式
            rgb = cv2.cvtColor(original, cv2.COLOR_BGR2RGB)

            # 确保掩码是单通道的
            if len(mask.shape) == 3:
                gray_mask = cv2.cvtColor(mask, cv2.COLOR_BGR2GRAY)
            else:
                gray_mask = mask.copy()

            # 归一化掩码到0-1范围
            norm_mask = gray_mask.astype(np.float32) / 255.0

            # 创建热力图
            heatmap = cv2.applyColorMap((norm_mask * 255).astype(np.uint8), cv2.COLORMAP_JET)
            heatmap_rgb = cv2.cvtColor(heatmap, cv2.COLOR_BGR2RGB)

            # 混合原图和热力图
            alpha = 0.3
            blended = cv2.addWeighted(rgb, 1 - alpha, heatmap_rgb, alpha, 0)

            heatmap_path = os.path.join(output_dir, f"{base_name}_heatmap_overlay.png")
            cv2.imwrite(heatmap_path, cv2.cvtColor(blended, cv2.COLOR_RGB2BGR))

            return heatmap_path
        except Exception as e:
            print(f"创建热力图叠加图像时出错: {e}")
            return None

    def save_comprehensive_export(self, original, segmented, mask, image_name, output_dir,
                                  export_formats=['png'], dpi=300, include_individual=True):
        """保存所有类型的导出图像，包括各种格式和样式"""
        base_name = os.path.splitext(os.path.basename(image_name))[0]
        exported_files = []

        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)

        # 为每种格式保存不同的图像组合
        for fmt in export_formats:
            # 保存不同样式的对比图
            try:
                # 样式1对比图
                fig1 = self.create_comparison_style1(original, segmented, mask, "深度学习分割", dpi=dpi)
                style1_path = os.path.join(output_dir, f"{base_name}_comparison_style1.{fmt}")
                fig1.savefig(style1_path, dpi=dpi, bbox_inches='tight',
                             format={'png': 'png', 'jpg': 'jpg', 'tif': 'tiff', 'pdf': 'pdf'}.get(fmt, 'png'))
                plt.close(fig1)
                exported_files.append(style1_path)

                # 样式2对比图
                fig2 = self.create_comparison_style2(original, segmented, "深度学习分割", image_name, dpi=dpi)
                style2_path = os.path.join(output_dir, f"{base_name}_comparison_style2.{fmt}")
                fig2.savefig(style2_path, dpi=dpi, bbox_inches='tight',
                             format={'png': 'png', 'jpg': 'jpg', 'tif': 'tiff', 'pdf': 'pdf'}.get(fmt, 'png'))
                plt.close(fig2)
                exported_files.append(style2_path)

                # 样式3对比图（包含透明主体图）
                fig3 = self.create_comparison_style3(original, segmented, mask, "深度学习分割", image_name, dpi=dpi)
                style3_path = os.path.join(output_dir, f"{base_name}_comparison_style3.{fmt}")
                fig3.savefig(style3_path, dpi=dpi, bbox_inches='tight',
                             format={'png': 'png', 'jpg': 'jpg', 'tif': 'tiff', 'pdf': 'pdf'}.get(fmt, 'png'))
                plt.close(fig3)
                exported_files.append(style3_path)
            except Exception as e:
                print(f"保存对比图时出错: {e}")

        # 保存单独的图像（如果需要）
        if include_individual:
            individual_result = self.save_individual_images(original, segmented, mask, image_name,
                                                            output_dir, dpi, show_title=False, export_fmt='png')
            for path_key, path_value in individual_result.items():
                if path_value:
                    exported_files.append(path_value)

        # 保存额外的特殊图像
        try:
            # 透明背景主体图像
            subject_transparent_path = self.create_subject_only_image(original, segmented, mask, output_dir, base_name)
            if subject_transparent_path:
                exported_files.append(subject_transparent_path)

            # 黑色背景主体图像
            subject_black_bg_path = self.create_subject_with_black_background(original, segmented, mask, output_dir,
                                                                              base_name)
            if subject_black_bg_path:
                exported_files.append(subject_black_bg_path)

            # 掩码覆盖图像
            overlay_path = self.create_mask_overlay(original, mask, output_dir, base_name)
            if overlay_path:
                exported_files.append(overlay_path)

            # 边缘高亮图像
            edge_path = self.create_edge_highlight_image(original, mask, output_dir, base_name)
            if edge_path:
                exported_files.append(edge_path)

            # 二值掩码图像
            binary_path = self.create_binary_mask_image(mask, output_dir, base_name)
            if binary_path:
                exported_files.append(binary_path)

            # 热力图叠加图像
            heatmap_path = self.create_heatmap_overlay(original, mask, output_dir, base_name)
            if heatmap_path:
                exported_files.append(heatmap_path)
        except Exception as e:
            print(f"保存特殊图像时出错: {e}")

        return exported_files

    def generate_export_filename(self, base_name, method_clean, suffix, extension='.png'):
        """根据设置生成导出文件名"""
        if self.filename_format == 'time_name':
            # 时间+样本名格式
            time_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{time_str}_{base_name}_{method_clean}_{suffix}{extension}"
        else:
            # 自定义前缀格式
            filename = f"{self.custom_filename_prefix}_{base_name}_{method_clean}_{suffix}{extension}"
        return filename


# ========== DPI设置对话框 ==========
class DPISettingsDialog(QDialog):
    """DPI和尺寸设置对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("高DPI导出设置")
        self.setModal(True)
        self.resize(400, 300)

        layout = QVBoxLayout(self)

        # DPI设置
        dpi_group = QGroupBox("DPI设置")
        dpi_layout = QVBoxLayout()

        self.dpi_300 = QRadioButton("300 DPI (标准论文)")
        self.dpi_600 = QRadioButton("600 DPI (高清)")
        self.dpi_1200 = QRadioButton("1200 DPI (超高清)")
        self.dpi_300.setChecked(True)

        self.dpi_group = QButtonGroup()
        self.dpi_group.addButton(self.dpi_300, 300)
        self.dpi_group.addButton(self.dpi_600, 600)
        self.dpi_group.addButton(self.dpi_1200, 1200)

        dpi_layout.addWidget(self.dpi_300)
        dpi_layout.addWidget(self.dpi_600)
        dpi_layout.addWidget(self.dpi_1200)
        dpi_group.setLayout(dpi_layout)

        # 图像尺寸设置
        size_group = QGroupBox("图像尺寸 (英寸)")
        size_layout = QVBoxLayout()

        width_layout = QHBoxLayout()
        width_layout.addWidget(QLabel("宽度:"))
        self.width_spin = QDoubleSpinBox()
        self.width_spin.setRange(1.0, 50.0)
        self.width_spin.setValue(15.0)
        self.width_spin.setSingleStep(0.5)
        width_layout.addWidget(self.width_spin)
        size_layout.addLayout(width_layout)

        height_layout = QHBoxLayout()
        height_layout.addWidget(QLabel("高度:"))
        self.height_spin = QDoubleSpinBox()
        self.height_spin.setRange(1.0, 50.0)
        self.height_spin.setValue(5.0)
        self.height_spin.setSingleStep(0.5)
        height_layout.addWidget(self.height_spin)
        size_layout.addLayout(height_layout)

        size_group.setLayout(size_layout)

        # 保存选项
        save_group = QGroupBox("保存选项")
        save_layout = QVBoxLayout()

        self.save_individual = QCheckBox("单独保存原图、分割图、掩码")
        self.save_individual.setChecked(True)
        save_layout.addWidget(self.save_individual)

        save_group.setLayout(save_layout)

        layout.addWidget(dpi_group)
        layout.addWidget(size_group)
        layout.addWidget(save_group)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        ok_btn = QPushButton("确定")
        cancel_btn = QPushButton("取消")
        ok_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

    def get_settings(self):
        """获取设置值"""
        dpi = self.dpi_group.checkedId()
        return {
            'dpi': dpi if dpi > 0 else 300,
            'width': self.width_spin.value(),
            'height': self.height_spin.value(),
            'save_individual': self.save_individual.isChecked()
        }


# ========== 识别结果独立窗口 ==========
class AnalysisResultWindow(QMainWindow):
    def __init__(self, stats_data, total_samples, file_name, parent=None):
        super().__init__(parent)
        self.stats_data = stats_data
        self.total_samples = total_samples
        # 对filename进行分割，将文件后缀去掉
        # Path(file_name).stem
        self.file_name = file_name.split('.')[0]

        self.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint |
            Qt.WindowType.WindowCloseButtonHint
        )

        self.setWindowTitle(f"识别结果分析 - {file_name}")
        self.resize(850, 650)
        self.init_ui()

    def init_ui(self):
        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # 设置图标
        # self.setWindowIcon(QIcon(r"resources\assets\images\button\results.png"))
        # 换成相对路径
        self.setWindowIcon(QIcon(r"resources\assets\images\button\results.png"))

        # 顶部：简洁标题 + 按钮行（风格参考 Dataset）
        header_layout = QHBoxLayout()
        title_lbl = QLabel(f"样本: {self.file_name} | 总采样点: {self.total_samples}")
        title_lbl.setStyleSheet("font-weight: bold; color: #34495e; font-size: 13px;")
        header_layout.addWidget(title_lbl)
        header_layout.addStretch()

        btn_layout = QHBoxLayout()
        self.btn_3d = QPushButton("3D颜色空间图")
        self.btn_hist = QPushButton("颜色分布直方图")
        self.btn_export = QPushButton("导出数据")

        for btn in [self.btn_3d, self.btn_hist, self.btn_export]:
            btn.setFixedHeight(28)
            btn.setMinimumWidth(110)
            btn_layout.addWidget(btn)
        header_layout.addLayout(btn_layout)

        layout.addLayout(header_layout)

        # [中间内容区 - 简洁表格样式]
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        self.card_layout = QVBoxLayout(container)
        self.card_layout.setContentsMargins(10, 10, 10, 10)
        self.card_layout.setSpacing(5)
        self.card_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        # 生成简单的数据行
        for i, item in enumerate(self.stats_data):
            if isinstance(item, dict):
                row = self.create_simple_row(i + 1, item)
            else:
                continue
            self.card_layout.addWidget(row)

        scroll.setWidget(container)
        layout.addWidget(scroll)

        # 绑定按钮事件
        self.btn_3d.clicked.connect(self.show_3d_view)
        self.btn_hist.clicked.connect(self.show_hist_view)
        self.btn_export.clicked.connect(self.export_data)

    def create_simple_row(self, idx, data):
        row_frame = QFrame()
        row_frame.setStyleSheet("""
            QFrame { 
                background: white; border: none;
                border-bottom: 1px solid #f2f2f2; padding: 12px 5px;
            }
            QFrame:hover { background: #f9f9f9; }
        """)

        main_layout = QHBoxLayout(row_frame)
        main_layout.setSpacing(25)

        # 1. 序号
        idx_label = QLabel(f"{idx:02d}")
        idx_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #bdc3c7;")
        main_layout.addWidget(idx_label)

        # 2. 颜色名称和代码区 (增加固定文字标签)
        name_code_container = QWidget()
        nc_layout = QVBoxLayout(name_code_container)
        nc_layout.setContentsMargins(0, 0, 0, 0)
        nc_layout.setSpacing(4)

        # 使用辅助函数创建带标签的行
        nc_layout.addWidget(self._make_labeled_text("ColorName:", data.get('name', '未知'), "#2c3e50", True))
        nc_layout.addWidget(self._make_labeled_text("ColorCode:", data.get('code', '未知'), "#2c3e50", True))

        main_layout.addWidget(name_code_container, stretch=2)

        # 3. 向量对比区 (上下排列 + 颜色预览)
        vector_container = QWidget()
        v_layout = QVBoxLayout(vector_container)
        v_layout.setContentsMargins(0, 0, 0, 0)
        v_layout.setSpacing(4)

        std_v = [int(x) for x in data.get('std_vec', [0, 0, 0])]
        act_v = [int(x) for x in data.get('target_vec', [0, 0, 0])]

        # v_layout.addWidget(self._make_vector_line("标准向量:", std_v, "#3498db"))
        # v_layout.addWidget(self._make_vector_line("实测向量:", act_v, "#e74c3c"))
        v_layout.addWidget(self._make_vector_line("STD", std_v, "#2c3e50"))
        v_layout.addWidget(self._make_vector_line("ACT", act_v, "#2c3e50"))

        main_layout.addWidget(vector_container, stretch=2)

        # --- 4. 统计信息 (占比和频次) ---
        stat_frame = QWidget()
        stat_layout = QVBoxLayout(stat_frame)
        stat_layout.setContentsMargins(0, 0, 0, 0)
        stat_layout.setSpacing(4)
        stat_layout.setAlignment(Qt.AlignmentFlag.AlignRight)  # 整体靠右对齐

        # 准备数据
        percent_val = f"{data.get('percent', 0.0):.2f}%"
        count_val = f"{data.get('count', 0)} px"

        # 使用辅助函数创建带标签的行，颜色统一为 #2c3e50
        # 提示：如果希望标签也靠右，可以在 _make_labeled_text 里调整，或者直接在这里添加
        stat_layout.addWidget(self._make_labeled_text("占比:", percent_val, "#2c3e50", True))
        stat_layout.addWidget(self._make_labeled_text("频次:", count_val, "#2c3e50", True))

        main_layout.addWidget(stat_frame, stretch=2)

        return row_frame

    def _make_labeled_text(self, label_text, value_text, color, is_bold):
        """辅助函数：创建 [标签: 变量] 结构的横向部件"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)

        # 固定文字标签
        lbl = QLabel(label_text)
        lbl.setStyleSheet("color: #95a5a6; font-size: 11px; font-weight: normal; min-width: 70px;")

        # 变量内容 (使用 QLineEdit 方便复制)
        entry = QLineEdit(value_text)
        entry.setReadOnly(True)
        bold_style = "font-weight: bold;" if is_bold else ""
        entry.setStyleSheet(f"border: none; background: transparent; color: {color}; {bold_style} font-size: 12px;")

        layout.addWidget(lbl)
        layout.addWidget(entry)
        layout.addStretch()
        return widget

    def _make_vector_line(self, label_text, vec, text_color):
        """辅助函数：创建 [向量标签: 坐标值 [色块]]"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        lbl = QLabel(label_text)
        lbl.setStyleSheet("color: #7f8c8d; font-size: 11px; min-width: 60px;")

        vec_str = f"{vec[0]}, {vec[1]}, {vec[2]}"
        entry = QLineEdit(vec_str)
        entry.setReadOnly(True)
        entry.setFixedWidth(110)
        entry.setStyleSheet(f"""
            border: 1px solid #f0f0f0; background: #fafafa; 
            color: {text_color}; font-family: 'Consolas'; font-size: 11px;
            padding: 1px 4px; border-radius: 2px;
        """)

        # 颜色预览色块
        color_swatch = QFrame()
        color_swatch.setFixedSize(35, 16)
        color_swatch.setStyleSheet(f"""
            background-color: rgb({vec[0]}, {vec[1]}, {vec[2]});
            border: 1px solid #ddd; border-radius: 3px;
        """)

        layout.addWidget(lbl)
        layout.addWidget(entry)
        layout.addWidget(color_swatch)
        layout.addStretch()
        return widget

    def _make_vector_line(self, label_text, vec, text_color):
        """辅助函数：创建 单行标签+坐标+色块 的部件"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)

        # 引导文字
        lbl = QLabel(f"{label_text}:")
        lbl.setStyleSheet(f"color: #7f8c8d; font-size: 11px; min-width: 55px;")

        # RGB数值文本 (可复制)
        vec_str = f"{vec[0]}, {vec[1]}, {vec[2]}"
        entry = QLineEdit(vec_str)
        entry.setReadOnly(True)
        entry.setStyleSheet(f"""
            border: none; 
            background: #f8f9fa; 
            color: {text_color}; 
            font-family: 'Consolas'; 
            font-size: 11px;
            padding: 2px 5px;
            border-radius: 2px;
        """)
        entry.setFixedWidth(100)

        # 颜色预览色块
        color_swatch = QFrame()
        color_swatch.setFixedSize(30, 14)
        # 绘制对应的颜色
        color_swatch.setStyleSheet(f"""
            background-color: rgb({vec[0]}, {vec[1]}, {vec[2]});
            border: 1px solid #ddd;
            border-radius: 2px;
        """)

        layout.addWidget(lbl)
        layout.addWidget(entry)
        layout.addWidget(color_swatch)
        layout.addStretch()

        return widget

    def show_3d_view(self):
        """弹出带工具栏的独立 3D 颜色空间窗口"""
        # 从父窗口获取DPI设置，如果没有则使用默认值300
        dpi = 300
        if self.parent() and hasattr(self.parent(), 'dpi'):
            dpi = self.parent().dpi

        self.plot_3d_win = MatplotlibWindow(f"3D 颜色空间 - {self.file_name}", self)

        fig = self.plot_3d_win.get_figure()
        fig.set_dpi(dpi)  # 设置DPI
        ax = fig.add_subplot(111, projection='3d')

        # 2. 准备数据（取前 50 个主要颜色）
        top_stats = self.stats_data[:50]

        # 收集数据
        xs, ys, zs, colors, sizes = [], [], [], [], []

        for item in top_stats:
            if isinstance(item, dict):
                target_vec = item.get('target_vec', [0, 0, 0])
                percent = item.get('percent', 0.0)
            else:
                # 兼容旧格式
                target_vec = item[5] if len(item) > 5 else [0, 0, 0]
                percent = item[3] if len(item) > 3 else 0.0

            if isinstance(target_vec, np.ndarray):
                target_vec = target_vec.tolist()

            xs.append(target_vec[0])
            ys.append(target_vec[1])
            zs.append(target_vec[2])
            colors.append(np.array(target_vec) / 255.0)
            sizes.append(percent * 20 + 50)

        if not xs:  # 没有数据
            QMessageBox.warning(self, "警告", "没有可显示的颜色数据")
            return

        # 3. 绘图
        ax.scatter(xs, ys, zs, c=colors, s=sizes,
                   alpha=0.8, edgecolors='black', linewidth=0.5)

        # 4. 设置轴标签和样式
        ax.set_xlabel('Red (R)', fontsize=9, fontweight='bold')
        ax.set_ylabel('Green (G)', fontsize=9, fontweight='bold')
        ax.set_zlabel('Blue (B)', fontsize=9, fontweight='bold')
        ax.set_title(f'RGB 坐标系下的 3D 颜色空间分布\n{self.file_name}',
                     fontsize=11, fontweight='bold', pad=20)

        ax.set_xlim([0, 255])
        ax.set_ylim([0, 255])
        ax.set_zlim([0, 255])
        ax.grid(True, alpha=0.3)

        # 5. 刷新画布并显示窗口
        self.plot_3d_win.draw()
        self.plot_3d_win.show()

    def export_data(self):
        """导出结果数据 (整合 TXT 报告、PNG 表格和 Excel 表格)"""
        default_name = f"{os.path.splitext(self.file_name)[0]}_识别结果"

        save_path, selected_filter = QFileDialog.getSaveFileName(
            self, "导出识别结果", default_name,
            "Text Files (*.txt);;"
            "PNG (*.png);;JPEG (*.jpg *.jpeg);;TIFF (*.tif *.tiff);;PDF (*.pdf);;"
            "Excel (*.xlsx);;CSV (*.csv)"
        )

        if not save_path:
            return

        try:
            ext = os.path.splitext(save_path)[1].lower()

            if ext == '.txt':
                self.export_txt(save_path)
            elif ext in ('.png', '.jpg', '.jpeg', '.tif', '.tiff', '.pdf'):
                self.export_png(save_path)
            elif ext == '.xlsx':
                self.export_excel(save_path)
            elif ext == '.csv':
                self.export_csv(save_path)
            else:
                save_path = save_path + '.txt'
                self.export_txt(save_path)

            # 导出成功，弹出交互对话框
            self.show_export_success_dialog(save_path)

        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"错误详情: {str(e)}")

    def export_png(self, save_path):
        # --- 导出识别结果图（支持 PNG/JPEG/TIFF/PDF，高DPI）---
        fmt, _ = get_export_format_from_path(save_path)
        savefig_fmt = {'png': 'png', 'jpg': 'jpg', 'tif': 'tiff', 'pdf': 'pdf'}.get(fmt, 'png')
        dpi = getattr(self.parent_app, 'dpi', 300) if hasattr(self, 'parent_app') and self.parent_app else 300

        # 动态计算高度
        fig_h = max(6, len(self.stats_data) * 0.5 + 2)
        fig = Figure(figsize=(12, fig_h), dpi=dpi)
        ax = fig.add_subplot(111)
        ax.axis('off')

        # 准备表格数据
        col_labels = ["序号", "颜色名称", "代码", "占比(%)", "频数", "标准RGB", "实测RGB"]
        table_vals = []

        for i, item in enumerate(self.stats_data):
            if isinstance(item, dict):
                name = item.get('name', '未知')
                code = item.get('code', '未知')
                percent = item.get('percent', 0.0)
                count = item.get('count', 0)
                std_vec = item.get('std_vec', [0, 0, 0])
                target_vec = item.get('target_vec', [0, 0, 0])
            else:
                name = item[0] if len(item) > 0 else '未知'
                code = item[1] if len(item) > 1 else '未知'
                count = item[2] if len(item) > 2 else 0
                percent = item[3] if len(item) > 3 else 0.0
                std_vec = item[4] if len(item) > 4 else [0, 0, 0]
                target_vec = item[5] if len(item) > 5 else [0, 0, 0]

            if isinstance(std_vec, np.ndarray):
                std_vec = std_vec.tolist()
            if isinstance(target_vec, np.ndarray):
                target_vec = target_vec.tolist()

            std_s = f"{int(std_vec[0])},{int(std_vec[1])},{int(std_vec[2])}"
            act_s = f"{int(target_vec[0])},{int(target_vec[1])},{int(target_vec[2])}"
            table_vals.append([
                str(i + 1), name, code, f"{percent:.2f}%",
                str(count), std_s, act_s
            ])

        # 绘制表格
        the_table = ax.table(cellText=table_vals, colLabels=col_labels,
                             loc='center', cellLoc='center', colColours=["#e6e6e6"] * 7)
        the_table.auto_set_font_size(False)
        the_table.set_fontsize(10)
        the_table.scale(1, 1.8)

        ax.set_title(f"岩石颜色识别结果表 - {self.file_name}\n(总采样点: {self.total_samples})",
                     fontsize=14, pad=20)

        fig.savefig(save_path, dpi=dpi, bbox_inches='tight', pad_inches=0.5, format=savefig_fmt)
        plt.close(fig)

    def export_excel(self, save_path):
        """导出为Excel：增加两列专门用于颜色填充预览"""
        # 1. 准备数据：增加 '标准颜色预览' 和 '实测颜色预览' 两列空占位
        rows = []
        for i, item in enumerate(self.stats_data):
            # 处理字典或列表格式
            if isinstance(item, dict):
                std_v = [int(x) for x in item.get('std_vec', [0, 0, 0])]
                act_v = [int(x) for x in item.get('target_vec', [0, 0, 0])]
                name, code, perc, count = item.get('name', ''), item.get('code', ''), item.get('percent', 0), item.get(
                    'count', 0)
            else:
                name, code, count, perc, std_v, act_v = item[0], item[1], item[2], item[3], item[4], item[5]

            rows.append([
                i + 1, name, code, round(perc, 2), count,
                std_v[0], std_v[1], std_v[2],
                act_v[0], act_v[1], act_v[2],
                "", ""  # 最后两列留空，准备填色
            ])

        cols = ['序号', '颜色名称', '代码', '占比(%)', '频数',
                '标准R', '标准G', '标准B', '实测R', '实测G', '实测B',
                '标准RGB预览', '实测RGB预览']
        df = pd.DataFrame(rows, columns=cols)

        # 2. 写入 Excel 并进行单元格染色
        try:
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='识别结果')
                worksheet = writer.sheets['识别结果']

                # --- 调整样式 ---
                # 设置色块预览列（L列和M列）宽一点，方便观察
                worksheet.column_dimensions['L'].width = 18
                worksheet.column_dimensions['M'].width = 18

                for row_idx in range(2, len(df) + 2):
                    # 获取该行的 RGB
                    sr, sg, sb = rows[row_idx - 2][5:8]  # 标准
                    ar, ag, ab = rows[row_idx - 2][8:11]  # 实测

                    # 转换为 HEX (注意：openpyxl 的 PatternFill 需要 RRGGBB)
                    hex_std = f"{int(sr):02x}{int(sg):02x}{int(sb):02x}"
                    hex_act = f"{int(ar):02x}{int(ag):02x}{int(ab):02x}"

                    # 填充最后两列 (Column 12 和 13)
                    cell_std = worksheet.cell(row=row_idx, column=12)
                    cell_act = worksheet.cell(row=row_idx, column=13)

                    cell_std.fill = PatternFill(start_color=hex_std, end_color=hex_std, fill_type="solid")
                    cell_act.fill = PatternFill(start_color=hex_act, end_color=hex_act, fill_type="solid")

                # 设置标题行颜色（灰色）
                header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                for cell in worksheet[1]:
                    cell.fill = header_fill

            QMessageBox.information(self, "成功", f"Excel已生成！\n请查看最后两列的颜色对比。")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存Excel失败: {str(e)}")

    def export_csv(self, save_path):
        """导出为CSV文件"""
        # 准备数据
        data = []
        for i, item in enumerate(self.stats_data):
            if isinstance(item, dict):
                name = item.get('name', '未知')
                code = item.get('code', '未知')
                percent = item.get('percent', 0.0)
                count = item.get('count', 0)
                std_vec = item.get('std_vec', [0, 0, 0])
                target_vec = item.get('target_vec', [0, 0, 0])
            else:
                # 兼容旧格式
                name = item[0] if len(item) > 0 else '未知'
                code = item[1] if len(item) > 1 else '未知'
                count = item[2] if len(item) > 2 else 0
                percent = item[3] if len(item) > 3 else 0.0
                std_vec = item[4] if len(item) > 4 else [0, 0, 0]
                target_vec = item[5] if len(item) > 5 else [0, 0, 0]

            if isinstance(std_vec, np.ndarray):
                std_vec = std_vec.tolist()
            if isinstance(target_vec, np.ndarray):
                target_vec = target_vec.tolist()

            data.append({
                '序号': i + 1,
                '颜色名称': name,
                '代码': code,
                '占比(%)': f"{percent:.2f}",
                '频数': count,
                '标准R': std_vec[0],
                '标准G': std_vec[1],
                '标准B': std_vec[2],
                '实测R': target_vec[0],
                '实测G': target_vec[1],
                '实测B': target_vec[2]
            })

        # 创建DataFrame并保存
        df = pd.DataFrame(data)
        df.to_csv(save_path, index=False, encoding='utf-8-sig')

    def show_export_success_dialog(self, file_path):
        """显示导出成功的交互对话框"""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("导出成功")
        msg_box.setText(f"文件已成功导出到：\n{file_path}")
        msg_box.setIcon(QMessageBox.Icon.Information)

        # 添加按钮
        open_file_btn = msg_box.addButton("打开文件", QMessageBox.ButtonRole.ActionRole)
        open_folder_btn = msg_box.addButton("打开所在文件夹", QMessageBox.ButtonRole.ActionRole)
        ok_btn = msg_box.addButton("确定", QMessageBox.ButtonRole.AcceptRole)

        msg_box.exec()

        # 处理按钮点击
        clicked_btn = msg_box.clickedButton()
        if clicked_btn == open_file_btn:
            QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))
        elif clicked_btn == open_folder_btn:
            folder_path = os.path.dirname(file_path)
            if platform.system() == "Windows":
                # 打开文件夹并选中该文件
                subprocess.run(['explorer', '/select,', os.path.normpath(file_path)])
            else:
                QDesktopServices.openUrl(QUrl.fromLocalFile(folder_path))

    def export_txt(self, save_path):
        # --- 导出 TXT ---
        if save_path.endswith('.txt'):
            with open(save_path, "w", encoding="utf-8") as f:
                f.write("=" * 60 + "\n")
                f.write("岩石颜色识别报告\n")
                f.write(f"样本名称: {self.file_name}\n")
                f.write(f"导出时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"总采样点: {self.total_samples}\n")
                f.write("=" * 60 + "\n\n")

                header = f"{'序号':<5}{'颜色名称':<12}{'代码':<12}{'占比(%)':<10}{'频数':<10}{'实测RGB':<20}\n"
                f.write(header)
                f.write("-" * 80 + "\n")

                for i, item in enumerate(self.stats_data):
                    if isinstance(item, dict):
                        name = item.get('name', '未知')
                        code = item.get('code', '未知')
                        percent = item.get('percent', 0.0)
                        count = item.get('count', 0)
                        target_vec = item.get('target_vec', [0, 0, 0])
                    else:
                        # 兼容旧格式
                        name = item[0] if len(item) > 0 else '未知'
                        code = item[1] if len(item) > 1 else '未知'
                        count = item[2] if len(item) > 2 else 0
                        percent = item[3] if len(item) > 3 else 0.0
                        target_vec = item[5] if len(item) > 5 else [0, 0, 0]

                    if isinstance(target_vec, np.ndarray):
                        target_vec = target_vec.tolist()

                    rgb_str = f"({int(target_vec[0])},{int(target_vec[1])},{int(target_vec[2])})"
                    f.write(f"{i + 1:<5}{name:<12}{code:<12}{percent:<10.2f}{count:<10}{rgb_str:<20}\n")

    def show_hist_view(self):
        """显示统计直方图窗口"""
        if not self.stats_data:
            return

        # 从父窗口获取DPI设置，如果没有则使用默认值300
        dpi = 300
        if self.parent() and hasattr(self.parent(), 'dpi'):
            dpi = self.parent().dpi

        self.hist_win = MatplotlibWindow(f"颜色分布直方图 - {self.file_name}", self)
        fig = self.hist_win.get_figure()
        fig.set_dpi(dpi)  # 设置DPI
        ax = fig.add_subplot(111)

        # 2. 准备数据
        plot_stats = self.stats_data[:12]
        names = []
        values = []
        colors = []

        for item in plot_stats:
            if isinstance(item, dict):
                name = item.get('name', '未知')
                code = item.get('code', '未知')
                percent = item.get('percent', 0.0)
                std_vec = item.get('std_vec', [0, 0, 0])
            else:
                # 兼容旧格式
                name = item[0] if len(item) > 0 else '未知'
                code = item[1] if len(item) > 1 else '未知'
                percent = item[3] if len(item) > 3 else 0.0
                std_vec = item[4] if len(item) > 4 else [0, 0, 0]

            # 直方图x轴颜色代码，不显示颜色名称
            # names.append(f"{name}\n({code})")
            names.append(f"{code}")
            values.append(percent)

            # 使用标准RGB值作为颜色
            if isinstance(std_vec, np.ndarray):
                std_vec = std_vec.tolist()
            rgb_normalized = [c / 255.0 for c in std_vec[:3]]
            colors.append(rgb_normalized)

        if not names:
            QMessageBox.warning(self, "警告", "没有可显示的数据")
            return

        # 3. 绘制柱状图
        bars = ax.bar(range(len(names)), values, color=colors, edgecolor='black', linewidth=1.2)

        # 4. 设置轴标签和样式
        ax.set_xlabel('颜色类别', fontsize=12, fontweight='bold')
        ax.set_ylabel('占比 (%)', fontsize=12, fontweight='bold')
        # ax.set_title(f'颜色分布直方图\n{self.file_name}',
        #             fontsize=14, fontweight='bold', pad=20)

        # 设置x轴刻度
        ax.set_xticks(range(len(names)))
        ax.set_xticklabels(names, rotation=45, ha='right', fontsize=9)

        # 设置y轴范围
        ax.set_ylim([0, max(values) * 1.2 if values else 100])

        # 在柱状图上添加数值标签
        for i, (bar, value) in enumerate(zip(bars, values)):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2., height,
                    f'{value:.2f}%',
                    ha='center', va='bottom', fontsize=8, fontweight='bold')

        # 设置网格
        ax.grid(True, alpha=0.3, axis='y', linestyle='--')

        # 设置字体和样式以提高清晰度
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial']
        plt.rcParams['axes.unicode_minus'] = False

        # 设置刻度标签
        ax.tick_params(axis='both', which='major', labelsize=10)

        plt.tight_layout()

        # 5. 刷新画布并显示窗口
        self.hist_win.draw()
        self.hist_win.show()


# ========== 识别核心算法 ==========
class ColorRecognizeMethods():
    """颜色识别核心算法"""

    def __init__(self):
        self.progress_var = None
        self.df = self._create_color_database()
        self.standard_vectors = None
        self.color_names = None
        self.color_codes = None
        self._prepare_color_data()

    def _create_color_database(self):
        """创建颜色数据库"""
        df = pd.read_csv(r"resources\files\color.csv", encoding='GBK', sep=',')
        if df.iloc[:, 6:9].isnull().any().any():
            print("数据中存在缺失值，请检查数据。")

        # 提前提取标准颜色向量、颜色名称和颜色代码，避免在循环中重复提取
        standard_vectors = df.iloc[:, 6:9].values
        color_names = df['岩石颜色'].values
        color_codes = df['Munsell颜色代码'].values

        return df

    def _prepare_color_data(self):
        """准备颜色数据"""
        if self.df is not None:
            self.standard_vectors = self.df.iloc[:, 6:9].values  # 修复：使用正确的列索引
            self.color_names = self.df['岩石颜色'].values
            self.color_codes = self.df['Munsell颜色代码'].values

    # 计算标准颜色向量
    @staticmethod
    def calculate_color_vector(row):
        return row['R'], row['G'], row['B']

    # 计算欧氏距离
    @staticmethod
    def euclidean_distance(vector1, vector2):
        return np.linalg.norm(np.array(vector1) - np.array(vector2))

    # 计算余弦相似度
    @staticmethod
    def cosine_similarity(vector1, vector2):
        vector1 = np.array(vector1)
        vector2 = np.array(vector2)
        dot_product = np.dot(vector1, vector2)
        magnitude1 = np.linalg.norm(vector1)
        magnitude2 = np.linalg.norm(vector2)
        return dot_product / (magnitude1 * magnitude2) if magnitude1 and magnitude2 else 0


# ========== 裁剪区域选择控件 ==========
class CropRectWidget(QWidget):
    """自定义裁剪区域选择控件"""
    rectChanged = pyqtSignal(QRect)
    rectConfirmed = pyqtSignal(QRect)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMouseTracking(True)

        # 状态变量
        self.image = None
        self.scale_factor = 1.0
        self.offset = QPoint(0, 0)
        self.original_size = QPoint(0, 0)

        # 裁剪矩形
        self.crop_rect = QRect()
        self.start_point = QPoint()
        self.end_point = QPoint()
        self.dragging = False
        self.resizing = False
        self.resize_edge = None
        self.dragging_image = False
        self.last_mouse_pos = QPoint()
        self.drawing_new_rect = False

        # 显示选项
        self.show_grid = True
        self.grid_size = 50
        self.show_coords = True
        self.show_pixel_info = True

        # 颜色配置
        self.rect_color = QColor(255, 255, 255)  # 白色边框
        self.rect_fill_color = QColor(255, 255, 255, 30)  # 轻微半透明
        self.grid_color = QColor(200, 200, 200, 100)
        self.text_color = QColor(255, 255, 255)

        # 边缘调整大小区域宽度
        self.edge_margin = 8

        # 调试标志
        self.debug_mode = False

        # 初始化UI
        self.init_ui()

    def init_ui(self):
        self.setMinimumSize(100, 100)
        self.setStyleSheet("background-color: #2c2c2c;")

    def resizeEvent(self, event):
        """窗口大小改变事件"""
        super().resizeEvent(event)
        # 如果已经有图片，重新适应窗口
        if self.image is not None:
            self.fit_to_view()

    def fit_to_view(self):
        """自适应窗口大小，确保图片居中显示"""
        if self.image is None or self.width() == 0 or self.height() == 0:
            return

        img_w, img_h = self.image.width(), self.image.height()
        view_w, view_h = self.width(), self.height()

        # 计算保持长宽比的最大缩放比例
        scale_w = view_w / img_w if img_w > 0 else 1.0
        scale_h = view_h / img_h if img_h > 0 else 1.0
        self.scale_factor = min(scale_w, scale_h, 1.0)  # 最大1:1显示

        # 计算缩放后的图像尺寸
        scaled_w = img_w * self.scale_factor
        scaled_h = img_h * self.scale_factor

        # 计算居中偏移
        self.offset = QPoint(
            int((view_w - scaled_w) // 2),
            int((view_h - scaled_h) // 2)
        )

        self.update()

    def get_image_point(self, widget_point):
        """将控件坐标转换为图像坐标"""
        if self.scale_factor <= 0:
            return QPoint(-1, -1)

        x = int(round((widget_point.x() - self.offset.x()) / self.scale_factor))
        y = int(round((widget_point.y() - self.offset.y()) / self.scale_factor))

        return QPoint(x, y)

    def get_widget_point(self, image_point):
        """将图像坐标转换为控件坐标"""
        x = int(round(image_point.x() * self.scale_factor + self.offset.x()))
        y = int(round(image_point.y() * self.scale_factor + self.offset.y()))
        return QPoint(x, y)

    def get_scaled_rect(self, rect):
        """将图像矩形转换为控件矩形"""
        if rect.isNull():
            return QRect()

        top_left = self.get_widget_point(rect.topLeft())
        bottom_right = self.get_widget_point(rect.bottomRight())

        return QRect(top_left, bottom_right)

    def paintEvent(self, event):
        """绘制事件"""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # 绘制背景
        painter.fillRect(self.rect(), QColor(45, 45, 45))  # 深灰色背景

        if self.image is None:
            painter.setPen(QColor(200, 200, 200))
            painter.setFont(QFont("Microsoft YaHei", 12))
            painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter,
                             "请加载图像...")
            return

        # 计算缩放后的图像尺寸
        scaled_w = int(self.original_size.x() * self.scale_factor)
        scaled_h = int(self.original_size.y() * self.scale_factor)

        if scaled_w <= 0 or scaled_h <= 0:
            return

        # 缩放图像并保持长宽比
        scaled_image = self.image.scaled(
            scaled_w, scaled_h,
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation
        )

        # 居中绘制图像
        painter.drawImage(self.offset.x(), self.offset.y(), scaled_image)

        # 如果图片较小，绘制边界框
        if scaled_w < self.width() and scaled_h < self.height():
            painter.setPen(QPen(QColor(100, 100, 100), 1))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRect(self.offset.x() - 1, self.offset.y() - 1,
                             scaled_w + 2, scaled_h + 2)

        if self.show_grid and self.scale_factor > 0.1:
            self.draw_grid(painter)

        if not self.crop_rect.isNull():
            scaled_rect = self.get_scaled_rect(self.crop_rect)
            self.draw_crop_rect(painter, scaled_rect)

        if self.show_coords:
            self.draw_coordinate_info(painter)

    def debug_print(self, message):
        """调试打印"""
        if self.debug_mode:
            print(f"[CropRectWidget] {message}")

    def set_image(self, image):
        """设置要裁剪的图像"""
        self.image = image
        if self.image is not None:
            self.original_size = QPoint(self.image.width(), self.image.height())
            self.scale_factor = 1.0
            self.offset = QPoint(0, 0)
            self.fit_to_view()
        self.update()

    def draw_grid(self, painter):
        """绘制网格"""
        if self.scale_factor <= 0:
            return

        painter.setPen(QPen(self.grid_color, 1, Qt.PenStyle.DotLine))

        grid_spacing = int(self.grid_size * self.scale_factor)
        if grid_spacing < 10:
            return

        start_x = self.offset.x()
        end_x = start_x + int(self.original_size.x() * self.scale_factor)
        x = start_x
        while x <= end_x:
            painter.drawLine(x, self.offset.y(), x,
                             self.offset.y() + int(self.original_size.y() * self.scale_factor))
            x += grid_spacing

        start_y = self.offset.y()
        end_y = start_y + int(self.original_size.y() * self.scale_factor)
        y = start_y
        while y <= end_y:
            painter.drawLine(self.offset.x(), y,
                             self.offset.x() + int(self.original_size.x() * self.scale_factor), y)
            y += grid_spacing

    def draw_crop_rect(self, painter, rect):
        """绘制裁剪矩形 - 简化大方版"""
        if rect.isNull():
            return

        # 轻微半透明填充（几乎看不见）
        painter.setBrush(QBrush(self.rect_fill_color))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRect(rect)

        # 白色实线边框（比虚线更清晰）
        pen = QPen(self.rect_color, 2, Qt.PenStyle.SolidLine)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.setPen(pen)
        painter.drawRect(rect)

        # 显示尺寸信息（白色文字，简洁大方）
        painter.setFont(QFont("Arial", 10, QFont.Weight.Bold))

        # 使用实际的裁剪矩形尺寸，不是缩放后的rect
        width = self.crop_rect.width()
        height = self.crop_rect.height()
        text = f"{width} × {height} px"

        # 计算文本位置（在矩形上方居中显示）
        text_width = painter.fontMetrics().horizontalAdvance(text)
        text_x = rect.center().x() - text_width // 2
        text_y = rect.top() - 8

        # 绘制白色文字
        painter.setPen(QPen(QColor(255, 255, 255), 1))
        painter.drawText(text_x, text_y, text)

    def draw_coordinate_info(self, painter):
        """绘制坐标信息"""
        painter.setFont(QFont("Consolas", 10))

        mouse_pos = self.mapFromGlobal(self.cursor().pos())
        if self.rect().contains(mouse_pos):
            img_point = self.get_image_point(mouse_pos)

            if (0 <= img_point.x() < self.original_size.x() and
                    0 <= img_point.y() < self.original_size.y()):

                info_text = f"X: {img_point.x():4d}  Y: {img_point.y():4d}"

                if not self.crop_rect.isNull():
                    if self.crop_rect.contains(img_point):
                        rel_x = img_point.x() - self.crop_rect.left()
                        rel_y = img_point.y() - self.crop_rect.top()
                        info_text += f"  (ΔX: {rel_x:+4d}, ΔY: {rel_y:+4d})"

                # 在右下角显示坐标信息
                text_width = painter.fontMetrics().horizontalAdvance(info_text)
                text_x = self.width() - text_width - 10
                text_y = self.height() - 10

                # 绘制文本
                painter.setPen(QPen(QColor(255, 255, 255), 1))
                painter.drawText(text_x, text_y - 5, info_text)

    def mousePressEvent(self, event):
        """鼠标按下事件"""
        if self.image is None:
            return

        if event.button() == Qt.MouseButton.LeftButton:
            widget_pos = event.pos()
            img_pos = self.get_image_point(widget_pos)

            # 确保在图像范围内
            if not (0 <= img_pos.x() < self.original_size.x() and
                    0 <= img_pos.y() < self.original_size.y()):
                return

            # 检查是否在调整大小的边缘
            if not self.crop_rect.isNull():
                self.resize_edge = self.get_resize_edge(widget_pos)
                if self.resize_edge:
                    self.resizing = True
                    self.start_point = img_pos
                    return

                # 检查是否在裁剪矩形内
                scaled_rect = self.get_scaled_rect(self.crop_rect)
                if scaled_rect.contains(widget_pos):
                    self.dragging = True
                    self.start_point = img_pos
                    return

            # 检查是否按住Ctrl/Shift键拖动图像
            modifiers = event.modifiers()
            if modifiers & (Qt.KeyboardModifier.ControlModifier | Qt.KeyboardModifier.ShiftModifier):
                self.dragging_image = True
                self.last_mouse_pos = widget_pos
                self.setCursor(Qt.CursorShape.ClosedHandCursor)
                return

            # 开始绘制新矩形
            self.drawing_new_rect = True
            self.dragging = True
            self.crop_rect = QRect(img_pos, img_pos)
            self.start_point = img_pos
            self.end_point = img_pos

        elif event.button() == Qt.MouseButton.RightButton:
            # 右键清除选区
            self.crop_rect = QRect()
            self.update()
            self.rectChanged.emit(QRect())

        self.update()

    def mouseMoveEvent(self, event):
        """鼠标移动事件"""
        if self.image is None:
            return

        widget_pos = event.pos()
        img_pos = self.get_image_point(widget_pos)

        # 处理图像拖动
        if self.dragging_image:
            delta = widget_pos - self.last_mouse_pos
            self.offset += delta
            self.last_mouse_pos = widget_pos
            self.update()
            return

        # 确保在图像范围内
        img_pos.setX(max(0, min(img_pos.x(), self.original_size.x() - 1)))
        img_pos.setY(max(0, min(img_pos.y(), self.original_size.y() - 1)))

        # 更新鼠标光标
        cursor_set = False

        if not self.crop_rect.isNull():
            resize_edge = self.get_resize_edge(widget_pos)
            if resize_edge:
                self.set_resize_cursor(resize_edge)
                cursor_set = True
            else:
                scaled_rect = self.get_scaled_rect(self.crop_rect)
                if scaled_rect.contains(widget_pos):
                    self.setCursor(Qt.CursorShape.SizeAllCursor)
                    cursor_set = True

        # 检查是否按住Ctrl/Shift键显示手型光标
        modifiers = event.modifiers()
        if not cursor_set and modifiers & (Qt.KeyboardModifier.ControlModifier | Qt.KeyboardModifier.ShiftModifier):
            self.setCursor(Qt.CursorShape.OpenHandCursor)
            cursor_set = True

        if not cursor_set:
            self.setCursor(Qt.CursorShape.CrossCursor)

        # 处理调整大小
        if self.resizing and self.resize_edge:
            self.resize_crop_rect(img_pos)
            self.update()
            return

        # 处理拖拽
        if self.dragging:
            self.end_point = img_pos

            if self.drawing_new_rect:
                # 绘制新矩形
                self.crop_rect = QRect(self.start_point, self.end_point).normalized()
            else:
                # 移动现有矩形
                delta = img_pos - self.start_point
                new_rect = self.crop_rect.translated(delta)

                # 确保不超出图像边界
                if (new_rect.left() >= 0 and new_rect.top() >= 0 and
                        new_rect.right() < self.original_size.x() and
                        new_rect.bottom() < self.original_size.y()):
                    self.crop_rect = new_rect
                    self.start_point = img_pos

            self.update()
            self.rectChanged.emit(self.crop_rect)
            return

        self.update()

    def mouseReleaseEvent(self, event):
        """鼠标释放事件"""
        if event.button() == Qt.MouseButton.LeftButton:
            self.dragging = False
            self.resizing = False
            self.drawing_new_rect = False
            self.resize_edge = None
            self.dragging_image = False
            self.setCursor(Qt.CursorShape.ArrowCursor)

            # 如果矩形太小，清除它
            if self.crop_rect.width() < 5 and self.crop_rect.height() < 5:
                self.crop_rect = QRect()
                self.rectChanged.emit(QRect())

        self.update()

    def wheelEvent(self, event):
        """滚轮缩放事件"""
        if self.image is None or self.scale_factor <= 0:
            return

        try:
            # 获取鼠标位置
            mouse_pos = event.position().toPoint()

            # 获取缩放前的图像坐标
            old_img_pos = self.get_image_point(mouse_pos)

            # 计算缩放因子
            delta = event.angleDelta().y()
            zoom_factor = 1.1 if delta > 0 else 0.9

            # 计算新的缩放比例
            new_scale = self.scale_factor * zoom_factor

            # 限制缩放范围 (5% 到 500%)
            if new_scale < 0.05:
                new_scale = 0.05
            elif new_scale > 5.0:
                new_scale = 5.0

            # 如果缩放比例没有变化，直接返回
            if abs(new_scale - self.scale_factor) < 0.001:
                return

            # 更新缩放比例
            self.scale_factor = new_scale

            # 调整偏移量，使鼠标位置对应的图像点保持不变
            new_offset_x = mouse_pos.x() - old_img_pos.x() * self.scale_factor
            new_offset_y = mouse_pos.y() - old_img_pos.y() * self.scale_factor

            # 确保偏移量是有效的
            img_width = self.original_size.x()
            img_height = self.original_size.y()

            # 计算缩放后的图像尺寸
            scaled_width = img_width * self.scale_factor
            scaled_height = img_height * self.scale_factor

            # 限制偏移量，防止图像被拖出视图
            max_offset_x = max(0, int(scaled_width - self.width()))
            max_offset_y = max(0, int(scaled_height - self.height()))

            # 确保偏移量在合理范围内
            new_offset_x = max(-max_offset_x, min(0, int(new_offset_x)))
            new_offset_y = max(-max_offset_y, min(0, int(new_offset_y)))

            # 更新偏移量
            self.offset = QPoint(new_offset_x, new_offset_y)

            # 刷新显示
            self.update()

        except Exception as e:
            self.debug_print(f"滚轮事件错误: {e}")

    def get_resize_edge(self, widget_point):
        """获取调整大小的边缘"""
        if self.crop_rect.isNull():
            return None

        scaled_rect = self.get_scaled_rect(self.crop_rect)
        margin = self.edge_margin

        left_edge = abs(widget_point.x() - scaled_rect.left()) <= margin
        right_edge = abs(widget_point.x() - scaled_rect.right()) <= margin
        top_edge = abs(widget_point.y() - scaled_rect.top()) <= margin
        bottom_edge = abs(widget_point.y() - scaled_rect.bottom()) <= margin

        if left_edge and top_edge:
            return 'nw'
        elif left_edge and bottom_edge:
            return 'sw'
        elif right_edge and top_edge:
            return 'ne'
        elif right_edge and bottom_edge:
            return 'se'
        elif left_edge:
            return 'w'
        elif right_edge:
            return 'e'
        elif top_edge:
            return 'n'
        elif bottom_edge:
            return 's'

        return None

    def set_resize_cursor(self, edge):
        """设置调整大小的光标"""
        cursors = {
            'n': Qt.CursorShape.SizeVerCursor,
            's': Qt.CursorShape.SizeVerCursor,
            'w': Qt.CursorShape.SizeHorCursor,
            'e': Qt.CursorShape.SizeHorCursor,
            'nw': Qt.CursorShape.SizeFDiagCursor,
            'se': Qt.CursorShape.SizeFDiagCursor,
            'ne': Qt.CursorShape.SizeBDiagCursor,
            'sw': Qt.CursorShape.SizeBDiagCursor
        }
        self.setCursor(cursors.get(edge, Qt.CursorShape.ArrowCursor))

    def resize_crop_rect(self, img_pos):
        """调整裁剪矩形大小"""
        rect = self.crop_rect

        if 'n' in self.resize_edge:
            rect.setTop(int(max(0, min(img_pos.y(), rect.bottom() - 1))))
        if 's' in self.resize_edge:
            rect.setBottom(int(min(self.original_size.y() - 1, max(img_pos.y(), rect.top() + 1))))
        if 'w' in self.resize_edge:
            rect.setLeft(int(max(0, min(img_pos.x(), rect.right() - 1))))
        if 'e' in self.resize_edge:
            rect.setRight(int(min(self.original_size.x() - 1, max(img_pos.x(), rect.left() + 1))))

        self.crop_rect = rect.normalized()
        self.rectChanged.emit(self.crop_rect)

    def reset_view(self):
        """重置视图"""
        self.fit_to_view()
        self.update()

    def clear_selection(self):
        """清除选区"""
        self.crop_rect = QRect()
        self.update()
        self.rectChanged.emit(QRect())


# ========== 多个算法预览窗口 ==========
class MultiMethodSegmentationPreviewWindow(QMainWindow):
    """多算法分割结果预览窗口"""

    def __init__(self, results, image_name, parent=None):
        super().__init__(parent)
        self.results = results
        self.image_name = image_name
        self.parent_app = parent

        self.setWindowTitle(f"分割结果预览 - {os.path.basename(image_name)}")
        self.setWindowIcon(QIcon(r"resources\assets\images\button\segmentation.png"))
        self.resize(900, 600)

        self.init_ui()

    def init_ui(self):
        """初始化UI"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # 顶部：简洁标题行（参考 Dataset）
        header_layout = QHBoxLayout()
        title_label = QLabel(f"图像: {os.path.basename(self.image_name)} | 成功方法数: {len(self.results)}")
        title_label.setStyleSheet("font-weight: bold; color: #34495e; font-size: 13px;")
        header_layout.addWidget(title_label)
        header_layout.addStretch()
        layout.addLayout(header_layout)

        # 滚动区域显示所有算法的结果
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("border: 1px solid #dcdfe6; border-radius: 5px;")

        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setContentsMargins(10, 10, 10, 10)
        scroll_layout.setSpacing(15)

        # 为每个算法创建显示区域
        for method_name, result_data in self.results.items():
            method_group = self.create_method_result_group(method_name, result_data)
            scroll_layout.addWidget(method_group)

        scroll_layout.addStretch()
        scroll_area.setWidget(scroll_widget)
        layout.addWidget(scroll_area)

        # 按钮区域（简单扁平风格）
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        if self.parent_app:
            screenshot_btn = QPushButton("界面捕获")
            screenshot_btn.setIcon(QIcon(r"resources\assets\images\button\screenshot-fill.png"))
            screenshot_btn.clicked.connect(lambda: self.parent_app.capture_widget_screenshot(self, "SegPreview"))
            btn_layout.addWidget(screenshot_btn)
        export_all_btn = QPushButton("导出所有结果")
        export_all_btn.setIcon(QIcon(r"resources\assets\images\button\save.png"))
        export_all_btn.clicked.connect(self.export_all_results)
        btn_layout.addWidget(export_all_btn)

        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.close)
        btn_layout.addWidget(close_btn)

        layout.addLayout(btn_layout)

    def create_method_result_group(self, method_name, result_data):
        """为单个算法创建结果显示组"""
        group = QGroupBox(f"{method_name} (得分: {result_data['score']:.3f})")
        layout = QHBoxLayout(group)
        layout.setSpacing(10)

        # 生成 subject_transparent 和 subject_black_bg 图像
        original = result_data['original']
        mask = result_data['mask']
        h, w = original.shape[:2]
        subject_transparent = None
        subject_black_bg = None

        try:
            # 确保掩码是单通道的
            if len(mask.shape) == 3:
                alpha = cv2.cvtColor(mask, cv2.COLOR_BGR2GRAY)
            else:
                alpha = mask.copy()

            # 确保alpha值在0-255范围内
            if alpha.max() <= 1:
                alpha = (alpha * 255).astype(np.uint8)
            if alpha.shape[0] != h or alpha.shape[1] != w:
                alpha = cv2.resize(alpha, (w, h), interpolation=cv2.INTER_NEAREST)

            # 反转mask，让岩石区域变成255（不透明）
            alpha = cv2.bitwise_not(alpha)

            # 生成透明背景图像
            rgb = cv2.cvtColor(original, cv2.COLOR_BGR2RGB)
            alpha_channel = alpha.astype(np.float32) / 255.0
            rgba = np.dstack([rgb, (alpha_channel * 255).astype(np.uint8)])
            # 保持 RGBA 格式，不转换为 BGR，以便正确显示透明效果
            subject_transparent = rgba

            # 生成黑色背景图像
            black_bg = np.zeros_like(rgb, dtype=np.uint8)
            alpha_mask = alpha.astype(np.float32) / 255.0
            subject_with_black_bg = np.where(alpha_mask[..., np.newaxis] > 0.5, rgb, black_bg)
            subject_black_bg = cv2.cvtColor(subject_with_black_bg, cv2.COLOR_RGB2BGR)
        except Exception as e:
            print(f"生成透明和黑色背景图像时出错: {e}")

        # 原图
        orig_label = QLabel()
        orig_pixmap = self.cv2_to_pixmap(result_data['original'])
        orig_label.setPixmap(orig_pixmap.scaled(200, 200, Qt.AspectRatioMode.KeepAspectRatio,
                                                Qt.TransformationMode.SmoothTransformation))
        orig_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        orig_label.setStyleSheet("border: 2px solid #3498db; border-radius: 5px; padding: 5px;")

        orig_frame = QFrame()
        orig_layout = QVBoxLayout(orig_frame)
        orig_layout.addWidget(QLabel("原图"))
        orig_layout.addWidget(orig_label)
        layout.addWidget(orig_frame)
        # 分割图（使用透明背景效果）
        seg_label = QLabel()
        if subject_transparent is not None:
            seg_pixmap = self.cv2_to_pixmap(subject_transparent)
        else:
            seg_pixmap = self.cv2_to_pixmap(result_data['segmented'])
        seg_label.setPixmap(
            seg_pixmap.scaled(200, 200, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
        seg_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        seg_label.setStyleSheet("border: 2px solid #2ecc71; border-radius: 5px; padding: 5px;")

        seg_frame = QFrame()
        seg_layout = QVBoxLayout(seg_frame)
        seg_layout.addWidget(QLabel("分割图"))
        seg_layout.addWidget(seg_label)
        layout.addWidget(seg_frame)

        # 掩码
        mask_label = QLabel()
        if len(result_data['mask'].shape) == 2:
            mask_rgb = cv2.cvtColor(result_data['mask'], cv2.COLOR_GRAY2RGB)
        else:
            mask_rgb = result_data['mask']
        mask_pixmap = self.cv2_to_pixmap(mask_rgb)
        mask_label.setPixmap(mask_pixmap.scaled(200, 200, Qt.AspectRatioMode.KeepAspectRatio,
                                                Qt.TransformationMode.SmoothTransformation))
        mask_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        mask_label.setStyleSheet("border: 2px solid #e74c3c; border-radius: 5px; padding: 5px;")

        mask_frame = QFrame()
        mask_layout = QVBoxLayout(mask_frame)
        mask_layout.addWidget(QLabel("分割掩码"))
        mask_layout.addWidget(mask_label)
        layout.addWidget(mask_frame)

        # 移除透明背景预览框，分割结果已经显示了 subject_transparent 图像

        # 主体黑色背景图
        if subject_black_bg is not None:
            black_bg_label = QLabel()
            black_bg_pixmap = self.cv2_to_pixmap(subject_black_bg)
            black_bg_label.setPixmap(
                black_bg_pixmap.scaled(200, 200, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
            black_bg_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            black_bg_label.setStyleSheet("border: 2px solid #f39c12; border-radius: 5px; padding: 5px;")

            black_bg_frame = QFrame()
            black_bg_layout = QVBoxLayout(black_bg_frame)
            black_bg_layout.addWidget(QLabel("黑色背景"))
            black_bg_layout.addWidget(black_bg_label)
            layout.addWidget(black_bg_frame)

        # 导出单个按钮
        export_btn = QPushButton(f"导出\n{method_name}")
        export_btn.clicked.connect(lambda checked, m=method_name, d=result_data: self.export_single_result(m, d))
        layout.addWidget(export_btn)

        return group

    def cv2_to_pixmap(self, cv_img):
        """将OpenCV图像转换为QPixmap"""
        if len(cv_img.shape) == 2:
            height, width = cv_img.shape
            q_img = QImage(cv_img.data, width, height, width, QImage.Format.Format_Grayscale8)
        else:
            # 彩色图或RGBA图
            height, width, channel = cv_img.shape
            if channel == 4:
                # RGBA图像
                bytes_per_line = 4 * width
                # 已经是RGBA格式，直接使用
                q_img = QImage(cv_img.data, width, height, bytes_per_line, QImage.Format.Format_RGBA8888)
            else:
                # 普通彩色图
                bytes_per_line = 3 * width
                img_rgb = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
                q_img = QImage(img_rgb.data, width, height, bytes_per_line, QImage.Format.Format_RGB888)
        return QPixmap.fromImage(q_img)

    def export_single_result(self, method_name, result_data):
        """导出单个算法的结果"""
        if self.parent_app:
            self.parent_app.export_segmentation_high_dpi(
                result_data['original'],
                result_data['segmented'],
                result_data['mask'],
                method_name,
                self.image_name
            )

    # def export_all_results(self):
    #     """导出所有算法的结果"""
    #     if self.parent_app:
    #         for method_name, result_data in self.results.items():
    #             self.parent_app.export_segmentation_high_dpi(
    #                 result_data['original'],
    #                 result_data['segmented'],
    #                 result_data['mask'],
    #                 method_name,
    #                 self.image_name
    #             )

    def export_all_results(self):
        """导出所有算法的结果（带进度条）"""
        if not self.parent_app:
            return

        # 创建进度对话框
        progress_dialog = QProgressDialog("正在导出所有结果，请稍候...", "取消", 0, len(self.results), self)
        progress_dialog.setWindowTitle("批量导出")
        progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
        progress_dialog.setMinimumDuration(0)
        progress_dialog.setAutoClose(True)
        progress_dialog.setAutoReset(True)
        progress_dialog.setValue(0)

        all_exported_files = []

        try:
            for idx, (method_name, result_data) in enumerate(self.results.items(), 1):
                progress_dialog.setLabelText(f"正在导出: {method_name} ({idx}/{len(self.results)})")
                progress_dialog.setValue(idx - 1)
                QApplication.processEvents()

                # 选择保存目录与导出格式（第一次询问，后续使用相同设置）
                if idx == 1:
                    last_dir = self.parent_app.get_last_directory()
                    output_dir = QFileDialog.getExistingDirectory(
                        self,
                        "选择保存目录",
                        last_dir
                    )
                    if not output_dir:
                        progress_dialog.close()
                        return
                    export_fmt = self.parent_app._ask_export_format()
                    if export_fmt is None:
                        progress_dialog.close()
                        return
                else:
                    pass

                exported = self.parent_app.export_single_segmentation_result(
                    result_data['original'],
                    result_data['segmented'],
                    result_data['mask'],
                    method_name,
                    self.image_name,
                    output_dir,
                    export_fmt=export_fmt
                )

                if exported:
                    all_exported_files.extend(exported)

            progress_dialog.setValue(len(self.results))
            progress_dialog.close()

            # 显示统一的导出成功对话框
            if all_exported_files:
                self.parent_app.show_export_success_dialog(
                    all_exported_files,
                    title="批量导出成功",
                    message_prefix=f"已成功导出 {len(all_exported_files)} 个文件"
                )

        except Exception as e:
            progress_dialog.close()
            QMessageBox.critical(self, "导出错误", f"导出过程中发生错误:\n{str(e)}")


# ========== Segmentation Result Preview Window ==========
class SegmentationPreviewWindow(QMainWindow):
    """Segmentation result preview window with export buttons"""

    def __init__(self, original, segmented, subject_transparent, subject_black_bg, mask, method, image_name,
                 parent=None):
        super().__init__(parent)
        self.original = original
        self.segmented = segmented
        self.subject_transparent = subject_transparent  # Subject with transparent background
        self.subject_black_bg = subject_black_bg  # Subject with black background
        self.mask = mask
        self.method = method
        self.image_name = image_name
        self.parent_app = parent

        self.setWindowTitle(f"Segmentation Preview - {os.path.basename(image_name)}")
        self.setWindowIcon(QIcon(r"resources\assets\images\button\segmentation.png"))
        self.resize(900, 600)

        self.init_ui()

    def init_ui(self):
        """Initialize UI"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # Title bar
        title_frame = QFrame()
        title_frame.setStyleSheet("background-color: #2c3e50; border-radius: 5px; padding: 10px;")
        title_layout = QHBoxLayout(title_frame)

        title_label = QLabel(f"Method: {self.method}")
        title_label.setStyleSheet("color: white; font-weight: bold; font-size: 14px;")
        title_layout.addWidget(title_label)

        image_name_label = QLabel(f"Image: {os.path.basename(self.image_name)}")
        image_name_label.setStyleSheet("color: white; font-size: 12px;")
        title_layout.addStretch()
        title_layout.addWidget(image_name_label)

        layout.addWidget(title_frame)

        # Preview area (using scroll area)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("border: 1px solid #dcdfe6; border-radius: 5px;")

        preview_widget = QWidget()
        preview_layout = QHBoxLayout(preview_widget)
        preview_layout.setContentsMargins(5, 5, 5, 5)
        preview_layout.setSpacing(10)

        # Original image preview
        orig_label = QLabel()
        orig_pixmap = self.cv2_to_pixmap(self.original)
        orig_label.setPixmap(orig_pixmap.scaled(300, 300, Qt.AspectRatioMode.KeepAspectRatio,
                                                Qt.TransformationMode.SmoothTransformation))
        orig_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        orig_label.setStyleSheet("border: 2px solid #3498db; border-radius: 5px; padding: 5px;")

        orig_group = QGroupBox("Original Image")
        orig_layout = QVBoxLayout()
        orig_layout.addWidget(orig_label)
        orig_group.setLayout(orig_layout)

        # Segmentation result preview (using transparent background effect)
        seg_label = QLabel()
        if self.subject_transparent is not None:
            seg_pixmap = self.cv2_to_pixmap(self.subject_transparent)
        else:
            seg_pixmap = self.cv2_to_pixmap(self.segmented)
        seg_label.setPixmap(
            seg_pixmap.scaled(300, 300, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
        seg_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        seg_label.setStyleSheet("border: 2px solid #2ecc71; border-radius: 5px; padding: 5px;")

        seg_group = QGroupBox("Segmentation Result")
        seg_layout = QVBoxLayout()
        seg_layout.addWidget(seg_label)
        seg_group.setLayout(seg_layout)

        # Mask preview
        mask_label = QLabel()
        if len(self.mask.shape) == 2:
            mask_rgb = cv2.cvtColor(self.mask, cv2.COLOR_GRAY2RGB)
        else:
            mask_rgb = self.mask
        mask_pixmap = self.cv2_to_pixmap(mask_rgb)
        mask_label.setPixmap(mask_pixmap.scaled(300, 300, Qt.AspectRatioMode.KeepAspectRatio,
                                                Qt.TransformationMode.SmoothTransformation))
        mask_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        mask_label.setStyleSheet("border: 2px solid #e74c3c; border-radius: 5px; padding: 5px;")

        mask_group = QGroupBox("Segmentation Mask")
        mask_layout = QVBoxLayout()
        mask_layout.addWidget(mask_label)
        mask_group.setLayout(mask_layout)

        # Remove transparent background preview box, segmentation result already shows subject_transparent image

        # Subject with black background preview
        if self.subject_black_bg is not None:
            black_bg_label = QLabel()
            black_bg_pixmap = self.cv2_to_pixmap(self.subject_black_bg)
            black_bg_label.setPixmap(black_bg_pixmap.scaled(300, 300, Qt.AspectRatioMode.KeepAspectRatio,
                                                            Qt.TransformationMode.SmoothTransformation))
            black_bg_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            black_bg_label.setStyleSheet("border: 2px solid #f39c12; border-radius: 5px; padding: 5px;")

            black_bg_group = QGroupBox("Black Background")
            black_bg_layout = QVBoxLayout()
            black_bg_layout.addWidget(black_bg_label)
            black_bg_group.setLayout(black_bg_layout)
            preview_layout.addWidget(black_bg_group)

        # Add original three preview boxes
        preview_layout.addWidget(orig_group)
        preview_layout.addWidget(seg_group)
        preview_layout.addWidget(mask_group)

        scroll_area.setWidget(preview_widget)
        layout.addWidget(scroll_area)

        # Button area
        btn_frame = QFrame()
        btn_layout = QHBoxLayout(btn_frame)
        btn_layout.addStretch()

        # Export button
        export_btn = QPushButton("Export High DPI Images")
        export_btn.setIcon(QIcon(r"resources\assets\images\button\save.png"))
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)
        export_btn.clicked.connect(self.export_images)
        btn_layout.addWidget(export_btn)

        # Close button
        close_btn = QPushButton("Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        close_btn.clicked.connect(self.close)
        btn_layout.addWidget(close_btn)

        layout.addWidget(btn_frame)

        # Show current DPI setting info
        dpi_info = QLabel(
            f"Current DPI: {self.parent_app.dpi} DPI | Size: {self.parent_app.fig_width}×{self.parent_app.fig_height} inches")
        dpi_info.setStyleSheet("color: #7f8c8d; font-size: 11px; padding: 5px;")
        dpi_info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(dpi_info)

    def cv2_to_pixmap(self, cv_img):
        """Convert OpenCV image to QPixmap"""
        if len(cv_img.shape) == 2:
            # Grayscale image
            height, width = cv_img.shape
            q_img = QImage(cv_img.data, width, height, width, QImage.Format.Format_Grayscale8)
        else:
            # Color or RGBA image
            height, width, channel = cv_img.shape
            if channel == 4:
                # RGBA image
                bytes_per_line = 4 * width
                # Already RGBA format, use directly
                q_img = QImage(cv_img.data, width, height, bytes_per_line, QImage.Format.Format_RGBA8888)
            else:
                # Regular color image
                bytes_per_line = 3 * width
                img_rgb = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
                q_img = QImage(img_rgb.data, width, height, bytes_per_line, QImage.Format.Format_RGB888)
        return QPixmap.fromImage(q_img)

    def export_images(self):
        """Export images"""
        if self.parent_app:
            self.parent_app.export_segmentation_high_dpi(
                self.original, self.segmented, self.mask,
                self.method, self.image_name
            )


# ========== 裁剪窗口 ==========
class EnhancedCropWindow(QMainWindow):
    """增强版裁剪窗口（完整实现）"""
    cropConfirmed = pyqtSignal(QRect)
    cropCancelled = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("图像裁剪 - 选择目标区域")

        # 添加确认状态标志
        self.confirmed = False

        # 图像数据
        self.image = None
        self.qimage = None
        self.original_cv_image = None  # 存储原始OpenCV图像

        # 窗口尺寸相关
        self.screen_size = QApplication.primaryScreen().size()
        self.min_window_size = QSize(800, 600)  # 最小窗口大小
        self.max_window_size = QSize(1920, 1080)  # 最大窗口大小
        self.margin = 20  # 图片四周留白

        # 窗口尺寸比例因子
        self.width_factor = 0.8  # 窗口宽度占屏幕的比例
        self.height_factor = 0.8  # 窗口高度占屏幕的比例

        # 添加窗口图标
        try:
            self.setWindowIcon(QIcon(r"resources\assets\images\button\crop.png"))
        except:
            pass

        # 初始化UI
        self.init_ui()

        # 设置窗口样式
        self.setup_window_style()

    def keyPressEvent(self, event):
        """键盘快捷键：主键盘 Enter 和小键盘 Enter 都触发确认裁剪，Esc 取消"""
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            self.confirm_crop()
        elif event.key() == Qt.Key.Key_Escape:
            self.cancel_crop()
        else:
            super().keyPressEvent(event)

    def init_ui(self):
        """初始化UI"""
        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 主布局
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        # 1. 工具栏
        toolbar_layout = QHBoxLayout()

        # 缩放控制
        self.zoom_label = QLabel("缩放:")
        self.zoom_slider = QSlider(Qt.Orientation.Horizontal)
        self.zoom_slider.setRange(5, 500)
        self.zoom_slider.setValue(100)
        self.zoom_slider.setTickPosition(QSlider.TickPosition.TicksBelow)
        self.zoom_slider.setTickInterval(50)
        self.zoom_slider.valueChanged.connect(self.on_zoom_changed)

        self.zoom_value_label = QLabel("100%")
        self.zoom_value_label.setFixedWidth(50)

        self.fit_button = QPushButton("适应窗口")
        self.fit_button.clicked.connect(self.fit_to_window)
        self.fit_button.setFixedWidth(80)

        self.reset_button = QPushButton("重置")
        self.reset_button.clicked.connect(self.reset_view)
        self.reset_button.setFixedWidth(60)

        # 提示标签
        self.hint_label = QLabel("提示: 按住Ctrl/Shift拖动图像 | 滚轮缩放 | 右键清除选区")
        self.hint_label.setStyleSheet("color: #666; font-size: 12px; font-family: 'Microsoft YaHei';")

        toolbar_layout.addWidget(self.zoom_label)
        toolbar_layout.addWidget(self.zoom_slider)
        toolbar_layout.addWidget(self.zoom_value_label)
        toolbar_layout.addStretch()
        toolbar_layout.addWidget(self.hint_label)
        toolbar_layout.addStretch()
        toolbar_layout.addWidget(self.fit_button)
        toolbar_layout.addWidget(self.reset_button)

        main_layout.addLayout(toolbar_layout)

        # 2. 创建滚动区域
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: 1px solid #dcdfe6;
                border-radius: 4px;
                background-color: #2c2c2c;
            }
            QScrollBar:vertical, QScrollBar:horizontal {
                border: none;
                background: #f1f1f1;
                width: 10px;
                height: 10px;
                margin: 0px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical, QScrollBar::handle:horizontal {
                background: #c1c1c1;
                min-height: 20px;
                min-width: 20px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical:hover, QScrollBar::handle:horizontal:hover {
                background: #a8a8a8;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical,
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                border: none;
                background: none;
                height: 0px;
                width: 0px;
            }
        """)

        # 创建裁剪控件
        self.crop_widget = CropRectWidget()

        # 设置裁剪控件的背景色
        self.crop_widget.setStyleSheet("background-color: #2c2c2c;")

        # 设置裁剪控件的大小策略
        self.crop_widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # 将裁剪控件放入滚动区域
        scroll_area.setWidget(self.crop_widget)

        main_layout.addWidget(scroll_area, stretch=1)

        # 3. 信息面板
        info_group = QGroupBox("选区信息")
        info_layout = QHBoxLayout()

        self.pos_label = QLabel("位置: (0, 0)")
        self.size_label = QLabel("尺寸: 0 × 0 px")
        self.area_label = QLabel("面积: 0 px²")

        for label in [self.pos_label, self.size_label, self.area_label]:
            label.setStyleSheet("font-family: 'Consolas'; font-size: 12px;")

        info_layout.addWidget(self.pos_label)
        info_layout.addWidget(self.size_label)
        info_layout.addWidget(self.area_label)
        info_layout.addStretch()

        info_group.setLayout(info_layout)
        main_layout.addWidget(info_group)

        # 4. 控制按钮
        button_layout = QHBoxLayout()

        self.clear_button = QPushButton("清除选区")
        self.clear_button.clicked.connect(self.clear_selection)
        self.clear_button.setFixedWidth(100)
        self.clear_button.setStyleSheet("""
            QPushButton {
                background-color: #f0f0f0;
                border: 1px solid #dcdcdc;
                border-radius: 4px;
                padding: 8px 16px;
                font-family: "Microsoft YaHei";
            }
            QPushButton:hover {
                background-color: #ffebee;
                border-color: #f44336;
                color: #f44336;
            }
        """)

        self.cancel_button = QPushButton("取消")
        self.cancel_button.clicked.connect(self.cancel_crop)
        self.cancel_button.setFixedWidth(100)
        self.cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #f0f0f0;
                border: 1px solid #dcdcdc;
                border-radius: 4px;
                padding: 8px 16px;
                font-family: "Microsoft YaHei";
            }
            QPushButton:hover {
                background-color: #fff3e0;
                border-color: #ff9800;
                color: #ff9800;
            }
        """)

        self.confirm_button = QPushButton("确认裁剪")
        self.confirm_button.clicked.connect(self.confirm_crop)
        self.confirm_button.setFixedWidth(100)
        self.confirm_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: 1px solid #45a049;
                border-radius: 4px;
                padding: 8px 16px;
                font-family: "Microsoft YaHei";
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
                border-color: #3d8b40;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
        """)

        button_layout.addWidget(self.clear_button)
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_button)
        button_layout.addWidget(self.confirm_button)

        main_layout.addLayout(button_layout)

        # 连接信号
        self.crop_widget.rectChanged.connect(self.update_info)

        # 设置快捷键
        self.setup_shortcuts()

    def setup_shortcuts(self):
        """设置快捷键"""
        self.cancel_button.setShortcut("Esc")
        self.confirm_button.setShortcut("Return")
        self.fit_button.setShortcut("Space")
        self.clear_button.setShortcut("Delete")

        # 添加最大化/最小化快捷键
        self.minimize_action = QAction(self)
        self.minimize_action.setShortcut("Ctrl+M")
        self.minimize_action.triggered.connect(self.showMinimized)
        self.addAction(self.minimize_action)

        self.maximize_action = QAction(self)
        self.maximize_action.setShortcut("Ctrl+Shift+M")
        self.maximize_action.triggered.connect(self.toggle_maximize)
        self.addAction(self.maximize_action)

    def toggle_maximize(self):
        """切换最大化状态"""
        if self.isMaximized():
            self.showNormal()
        else:
            self.showMaximized()

    def setup_window_style(self):
        """设置窗口样式"""
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f6f7;
            }
            QGroupBox {
                border: 1px solid #dcdfe6;
                border-radius: 6px;
                margin-top: 10px;
                background-color: #ffffff;
                font-family: "Microsoft YaHei";
                font-weight: bold;
                color: #2c3e50;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 5px;
                left: 10px;
            }
            QLabel {
                font-family: "Microsoft YaHei";
            }
        """)

    def calculate_optimal_window_size(self, img_width, img_height):
        """计算最优窗口大小"""
        # 工具栏、信息面板、按钮的大致高度
        toolbar_height = 40
        info_height = 60
        button_height = 50
        margins = self.margin * 2  # 上下左右都有边距

        # 计算显示图片所需的最小窗口尺寸
        min_content_width = img_width + margins
        min_content_height = img_height + margins

        # 加上其他UI元素的高度
        min_window_width = min_content_width
        min_window_height = min_content_height + toolbar_height + info_height + button_height

        # 限制最小窗口大小
        window_width = max(min_window_width, self.min_window_size.width())
        window_height = max(min_window_height, self.min_window_size.height())

        # 限制最大窗口大小（不超过屏幕的80%）
        max_allowed_width = int(self.screen_size.width() * self.width_factor)
        max_allowed_height = int(self.screen_size.height() * self.height_factor)

        window_width = min(window_width, max_allowed_width)
        window_height = min(window_height, max_allowed_height)

        return QSize(window_width, window_height)

    def set_image(self, cv_image):
        """设置OpenCV图像"""
        if cv_image is None:
            return

        try:
            self.original_cv_image = cv_image.copy()
            height, width, channel = cv_image.shape

            if height <= 0 or width <= 0:
                raise Exception("图像尺寸无效")

            bytes_per_line = 3 * width
            img_rgb = cv2.cvtColor(cv_image, cv2.COLOR_BGR2RGB)
            self.qimage = QImage(img_rgb.data, width, height,
                                 bytes_per_line, QImage.Format.Format_RGB888).copy()

            if self.qimage.isNull():
                raise Exception("QImage创建失败")

            # 设置到裁剪控件
            self.crop_widget.set_image(self.qimage)
            self.crop_widget.original_size = QPoint(width, height)

            # 计算最优窗口大小
            optimal_size = self.calculate_optimal_window_size(width, height)

            # 调整窗口大小
            self.resize(optimal_size)

            # 居中显示
            self.center_on_screen()

            # 适应窗口显示
            self.fit_to_window()

            # 更新缩放滑块
            current_scale = int(self.crop_widget.scale_factor * 100)
            self.zoom_slider.setValue(current_scale)
            self.zoom_value_label.setText(f"{current_scale}%")

        except Exception as e:
            print(f"设置图像错误: {e}")
            QMessageBox.critical(self, "错误", f"设置图像失败: {str(e)}")

    def center_on_screen(self):
        """居中显示在屏幕上"""
        screen_geometry = QApplication.primaryScreen().availableGeometry()
        window_geometry = self.frameGeometry()
        center_point = screen_geometry.center()
        window_geometry.moveCenter(center_point)
        self.move(window_geometry.topLeft())

    def on_zoom_changed(self, value):
        """缩放滑块改变"""
        self.crop_widget.scale_factor = value / 100.0
        self.crop_widget.offset = QPoint(0, 0)
        self.crop_widget.update()
        self.zoom_value_label.setText(f"{value}%")

    def fit_to_window(self):
        """适应窗口大小"""
        self.crop_widget.fit_to_view()
        current_scale = int(self.crop_widget.scale_factor * 100)
        self.zoom_slider.setValue(current_scale)
        self.zoom_value_label.setText(f"{current_scale}%")
        self.crop_widget.update()

    def reset_view(self):
        """重置视图"""
        self.crop_widget.reset_view()
        current_scale = int(self.crop_widget.scale_factor * 100)
        self.zoom_slider.setValue(current_scale)
        self.zoom_value_label.setText(f"{current_scale}%")

    def clear_selection(self):
        """清除选区"""
        self.crop_widget.clear_selection()
        self.update_info(QRect())

    def update_info(self, rect):
        """更新信息显示"""
        if rect.isNull():
            self.pos_label.setText("位置: (0, 0)")
            self.size_label.setText("尺寸: 0 × 0 px")
            self.area_label.setText("面积: 0 px²")
        else:
            self.pos_label.setText(f"位置: ({rect.x()}, {rect.y()})")
            self.size_label.setText(f"尺寸: {rect.width()} × {rect.height()} px")
            self.area_label.setText(f"面积: {rect.width() * rect.height()} px²")

    def cancel_crop(self):
        """取消裁剪"""
        self.cropCancelled.emit()
        self.confirmed = True
        self.close()

    def confirm_crop(self):
        """确认裁剪"""
        rect = self.crop_widget.crop_rect
        if rect.isNull():
            reply = QMessageBox.question(
                self, "确认",
                "没有选择任何区域，是否取消裁剪？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.cropCancelled.emit()
                self.confirmed = True
                self.close()
            return

        self.cropConfirmed.emit(rect)
        self.confirmed = True
        self.close()

    def closeEvent(self, event):
        """关闭事件"""
        if self.confirmed:
            event.accept()
            return

        if not self.crop_widget.crop_rect.isNull():
            reply = QMessageBox.question(
                self, "确认",
                "您有未保存的选区，是否取消裁剪并关闭窗口？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.cropCancelled.emit()
                event.accept()
            else:
                event.ignore()
        else:
            self.cropCancelled.emit()
            event.accept()


# ========== 批量分割Worker ==========
class BatchSegmentationWorker(QThread):
    """批量分割工作线程"""
    progress_updated = pyqtSignal(int, str)  # 进度, 当前文件
    file_completed = pyqtSignal(str, dict)  # 文件路径, 结果
    finished_signal = pyqtSignal()
    error_occurred = pyqtSignal(str)

    def __init__(self, image_files, output_dir, selected_methods, parent_app, export_fmt='png'):
        super().__init__()
        self.image_files = image_files
        self.output_dir = output_dir
        self.selected_methods = selected_methods
        self.parent_app = parent_app
        self.export_fmt = export_fmt or 'png'
        self.is_running = True

    def run(self):
        """执行批量分割"""
        try:
            segmenter = RockSegmenter(log_callback=self.log_message)
            if self.parent_app and hasattr(self.parent_app, 'seg_model_path'):
                segmenter.set_dl_config(
                    getattr(self.parent_app, 'seg_model_path', '') or '',
                    getattr(self.parent_app, 'seg_model_use_gpu', True)
                )
            total = len(self.image_files)

            for idx, image_path in enumerate(self.image_files):
                if not self.is_running:
                    break

                self.progress_updated.emit(int((idx / total) * 100), os.path.basename(image_path))

                try:
                    # 读取图像（支持中文路径）
                    img = cv2_imread(image_path)
                    if img is None:
                        self.log_message(f"Failed to load: {image_path}")
                        continue

                    # 执行分割
                    results = segmenter.segment_by_methods(
                        img,
                        os.path.basename(image_path),
                        self.selected_methods,
                        log_callback=self.log_message
                    )

                    if results:
                        # 保存结果
                        self.save_results(image_path, results)
                        self.file_completed.emit(image_path, results)

                except Exception as e:
                    self.error_occurred.emit(f"Error processing {image_path}: {str(e)}")
                    self.log_message(f"Error: {str(e)}")

            self.progress_updated.emit(100, "Completed")
            self.finished_signal.emit()

        except Exception as e:
            self.error_occurred.emit(f"Batch segmentation failed: {str(e)}")

    def log_message(self, message):
        """记录日志"""
        if self.parent_app:
            self.parent_app.log(message)

    def save_results(self, image_path, results):
        """保存分割结果 - 完全复用单个图像分割的导出逻辑"""
        base_name = os.path.splitext(os.path.basename(image_path))[0]
        dataset_name = os.path.basename(os.path.dirname(image_path))

        # 为每个方法创建输出目录
        for method_name, result_data in results.items():
            method_clean = method_name.replace(' ', '_').replace('/', '_')

            # 创建样本文件夹: 日期_样本名
            date_str = datetime.datetime.now().strftime("%Y%m%d")
            sample_dir = os.path.join(
                self.output_dir,
                f"{dataset_name}_分割图_{method_clean}",
                f"{date_str}_{base_name}"
            )
            os.makedirs(sample_dir, exist_ok=True)

            # 获取分割结果
            original = result_data['original']
            segmented = result_data['segmented']
            mask = result_data['mask']

            # 直接调用 parent_app 的 export_single_segmentation_result 方法
            # 这样可以确保批量分割导出与单个图像分割导出完全一致
            try:
                if self.parent_app and hasattr(self.parent_app, 'export_single_segmentation_result'):
                    exported_files = self.parent_app.export_single_segmentation_result(
                        original, segmented, mask, method_name, image_path, sample_dir, export_fmt=self.export_fmt
                    )
                    if exported_files:
                        self.log_message(f"成功导出 {len(exported_files)} 个文件到 {sample_dir}")
                    else:
                        self.log_message(f"警告: 导出文件列表为空")
                else:
                    # 如果 parent_app 不可用，使用 HighDPIExporter 的 save_comprehensive_export 方法
                    self.log_message("parent_app 不可用，使用 HighDPIExporter.save_comprehensive_export")
                    exporter = HighDPIExporter(show_title=False)
                    dpi = 300
                    if self.parent_app:
                        dpi = getattr(self.parent_app, 'dpi', 300)
                    exported_files = exporter.save_comprehensive_export(
                        original, segmented, mask, image_path, sample_dir,
                        export_formats=[self.export_fmt], dpi=dpi, include_individual=True
                    )
            except Exception as e:
                self.log_message(f"导出失败: {str(e)}")
                import traceback
                self.log_message(traceback.format_exc())

    def stop(self):
        """停止处理"""
        self.is_running = False


# ========== 分割模型训练 ==========
class SegModelTrainingWorker(QThread):
    log_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, int)  # ok, exit_code (0 或子进程退出码)

    def __init__(self, img_dir, mask_dir, save_path, epochs, batch_size, parent=None, use_cpu=False):
        super().__init__(parent)
        self.img_dir = img_dir
        self.mask_dir = mask_dir
        self.save_path = save_path
        self.epochs = epochs
        self.batch_size = batch_size
        self.use_cpu = use_cpu

    def run(self):
        ok = False
        script_dir = os.path.dirname(os.path.abspath(__file__))
        run_script = os.path.join(script_dir, "run_unet_training.py")
        main_app = getattr(self.parent(), "parent_app", None) if self.parent() else self.parent()
        crash_log_dir = getattr(main_app, "log_save_path", None) if main_app else None
        if not crash_log_dir or not os.path.isdir(crash_log_dir):
            crash_log_dir = os.path.join(script_dir, "logs")
        os.makedirs(crash_log_dir, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        crash_log_path = os.path.join(crash_log_dir, f"unet_training_{ts}.log")
        if not os.path.isfile(run_script):
            try:
                self.log_signal.emit(f"未找到训练脚本: {run_script}，改用进程内训练")
            except Exception:
                pass
            try:
                from rock_seg_model import train_seg_model
                ok = train_seg_model(
                    self.img_dir, self.mask_dir, self.save_path,
                    epochs=self.epochs, batch_size=self.batch_size,
                    log_fn=lambda s: self.log_signal.emit(s),
                    use_cpu=self.use_cpu,
                )
            except BaseException as e:
                try:
                    self.log_signal.emit(f"训练异常: {e}")
                    self.log_signal.emit(traceback.format_exc())
                except Exception:
                    pass
            finally:
                try:
                    self.finished_signal.emit(ok, 0)
                except Exception:
                    pass
            return
        python_exe = sys.executable
        main_app = getattr(self.parent(), "parent_app", None) if self.parent() else self.parent()
        if main_app and getattr(main_app, "unet_python_path", ""):
            p = main_app.unet_python_path.strip()
            if p and os.path.isfile(p):
                python_exe = p
        cmd = [
            python_exe, run_script,
            "--img_dir", self.img_dir,
            "--mask_dir", self.mask_dir,
            "--save_path", self.save_path,
            "--epochs", str(self.epochs),
            "--batch_size", str(self.batch_size),
        ]
        if self.use_cpu:
            cmd.append("--cpu")
        exit_code = 0
        try:
            with open(crash_log_path, "w", encoding="utf-8") as crash_file:
                crash_file.write(f"U-Net 训练子进程日志 {ts}\n")
                crash_file.write(f"命令: {' '.join(cmd)}\n")
                crash_file.flush()
                env = dict(os.environ)
                env["PYTHONIOENCODING"] = "utf-8"
                if self.use_cpu:
                    env["CUDA_VISIBLE_DEVICES"] = ""
                    env["OMP_NUM_THREADS"] = "1"
                    env["MKL_NUM_THREADS"] = "1"
                    env["MKL_THREADING_LAYER"] = "SEQ"
                    env["KMP_DUPLICATE_LIB_OK"] = "TRUE"
                    env["KMP_WARNINGS"] = "0"
                proc = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    cwd=script_dir,
                    env=env,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    bufsize=1,
                )
                for line in iter(proc.stdout.readline, ""):
                    line = line.rstrip()
                    if not line:
                        continue
                    crash_file.write(line + "\n")
                    crash_file.flush()
                    try:
                        self.log_signal.emit(line)
                    except Exception:
                        pass
                proc.wait()
                ok = proc.returncode == 0
                exit_code = proc.returncode
                crash_file.write(f"\n进程退出码: {proc.returncode}\n")
                if proc.returncode == 3221225477:
                    crash_file.write(
                        "\n说明: 退出码 3221225477 (0xC0000005) 通常为显卡/CUDA 与当前 PyTorch 或驱动不兼容。\n")
                    crash_file.write(
                        "若已勾选「训练时强制使用 CPU」仍闪退，说明当前安装的 PyTorch 为带 CUDA 的版本，加载时仍会触发崩溃。\n")
                    crash_file.write("解决: 在本机命令行执行以下命令，安装纯 CPU 版 PyTorch 后重试：\n")
                    crash_file.write("  pip install torch --index-url https://download.pytorch.org/whl/cpu\n")
                    crash_file.write("(安装完成后无需改回，推理时仍可使用 GPU 版 PyTorch 的其他环境)\n")
                crash_file.flush()
        except Exception as e:
            exit_code = -1
            err_msg = f"子进程启动失败: {e}\n{traceback.format_exc()}"
            try:
                self.log_signal.emit(err_msg)
            except Exception:
                pass
            try:
                with open(crash_log_path, "a", encoding="utf-8") as f:
                    f.write(err_msg)
            except Exception:
                pass
            ok = False
        try:
            self.log_signal.emit(f"训练日志已保存: {crash_log_path}")
            self.finished_signal.emit(ok, exit_code)
        except Exception:
            pass


class SegModelTrainingDialog(QDialog):
    """分割模型训练对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_app = parent
        self.worker = None
        self.setWindowTitle("训练分割模型（U-Net）")
        self.setModal(True)
        self.resize(520, 520)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)

        info = QLabel("准备数据：图像与掩码需同名。掩码为二值图（黑背景白前景），支持 name.png 或 name_mask.png。")
        info.setStyleSheet("color: #7f8c8d; font-size: 11px;")
        info.setWordWrap(True)
        layout.addWidget(info)

        def add_row(label, edit, browse_slot):
            r = QHBoxLayout()
            r.addWidget(QLabel(label))
            r.addWidget(edit)
            b = QPushButton("浏览...")
            b.clicked.connect(browse_slot)
            r.addWidget(b)
            layout.addLayout(r)

        self.img_dir_edit = QLineEdit()
        self.img_dir_edit.setPlaceholderText("图像目录")
        add_row("图像目录:", self.img_dir_edit, self.browse_img_dir)

        self.mask_dir_edit = QLineEdit()
        self.mask_dir_edit.setPlaceholderText("掩码目录（可与图像同目录）")
        add_row("掩码目录:", self.mask_dir_edit, self.browse_mask_dir)

        self.save_edit = QLineEdit()
        self.save_edit.setPlaceholderText("默认: U-Net训练结果.pth")
        add_row("保存路径:", self.save_edit, self.browse_save)

        restore_row = QHBoxLayout()
        restore_btn = QPushButton("一键使用上次路径")
        restore_btn.clicked.connect(self._restore_last_unet_paths)
        restore_row.addWidget(restore_btn)
        restore_row.addStretch()
        layout.addLayout(restore_row)

        hp = QHBoxLayout()
        hp.addWidget(QLabel("训练轮数:"))
        self.epochs_spin = QSpinBox()
        self.epochs_spin.setRange(5, 200)
        self.epochs_spin.setValue(30)
        hp.addWidget(self.epochs_spin)
        hp.addWidget(QLabel("批大小:"))
        self.batch_spin = QSpinBox()
        self.batch_spin.setRange(2, 32)
        self.batch_spin.setValue(8)
        hp.addWidget(self.batch_spin)
        hp.addStretch()
        layout.addLayout(hp)

        self.unet_use_cpu_check = QCheckBox("训练时强制使用 CPU（若 GPU 训练闪退/退出码 0xC0000005 可勾选）")
        self.unet_use_cpu_check.setToolTip("勾选后使用 CPU 训练，速度较慢但可避免显卡驱动导致的崩溃")
        settings_u = QSettings("RockAnalysisTool", "Segmentation")
        self.unet_use_cpu_check.setChecked(settings_u.value("unet_train_use_cpu", False, type=bool))
        layout.addWidget(self.unet_use_cpu_check)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(100)
        layout.addWidget(self.log_text)

        # 训练曲线（训练完成后自动加载 _history.json 并绘制）
        curve_group = QGroupBox("训练曲线")
        curve_layout = QVBoxLayout(curve_group)
        self.curve_fig = Figure(figsize=(5, 2.2), dpi=80)
        self.curve_canvas = FigureCanvasQTAgg(self.curve_fig)
        self.curve_canvas.setMinimumHeight(160)
        curve_layout.addWidget(self.curve_canvas)
        layout.addWidget(curve_group)

        self.train_btn = QPushButton("开始训练")
        self.train_btn.clicked.connect(self.start_training)
        layout.addWidget(self.train_btn)
        self._plot_training_curve([])

    def browse_img_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择图像目录",
                                             self.parent_app.get_last_directory() if self.parent_app else "")
        if d:
            self.img_dir_edit.setText(d)
            if self.parent_app:
                self.parent_app.set_last_directory(d)

    def browse_mask_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择掩码目录", self.img_dir_edit.text() or (
            self.parent_app.get_last_directory() if self.parent_app else ""))
        if d:
            self.mask_dir_edit.setText(d)

    def _restore_last_unet_paths(self):
        """一键恢复上次 U-Net 训练使用的图像目录、掩码目录、保存路径"""
        settings = QSettings("RockAnalysisTool", "Segmentation")
        img = settings.value("unet_train_img_dir", "", type=str)
        mask = settings.value("unet_train_mask_dir", "", type=str)
        save = settings.value("unet_train_save_path", "", type=str)
        if img:
            self.img_dir_edit.setText(img)
        if mask:
            self.mask_dir_edit.setText(mask)
        if save:
            self.save_edit.setText(save)
        if not any((img, mask, save)):
            QMessageBox.information(self, "提示", "暂无上次路径记录，请先选择目录并训练一次。")

    def showEvent(self, event):
        """打开对话框时若输入框为空则从上次路径恢复"""
        super().showEvent(event)
        settings = QSettings("RockAnalysisTool", "Segmentation")
        if not self.img_dir_edit.text().strip():
            v = settings.value("unet_train_img_dir", "", type=str)
            if v:
                self.img_dir_edit.setText(v)
        if not self.mask_dir_edit.text().strip():
            v = settings.value("unet_train_mask_dir", "", type=str)
            if v:
                self.mask_dir_edit.setText(v)
        if not self.save_edit.text().strip():
            v = settings.value("unet_train_save_path", "", type=str)
            if v:
                self.save_edit.setText(v)
            else:
                # 无历史时给默认保存名
                if self.img_dir_edit.text().strip():
                    d = os.path.dirname(self.img_dir_edit.text().strip())
                    self.save_edit.setText(os.path.join(d, "U-Net训练结果.pth"))

    def browse_save(self):
        default_name = "U-Net训练结果.pth"
        start_dir = self.parent_app.get_last_directory() if self.parent_app else ""
        if self.img_dir_edit.text().strip():
            start_dir = os.path.dirname(self.img_dir_edit.text().strip()) or start_dir
        initial = os.path.join(start_dir, default_name)
        p, _ = QFileDialog.getSaveFileName(self, "保存模型", initial, "PyTorch (*.pt *.pth);;All (*)")
        if p:
            if not (p.lower().endswith('.pth') or p.lower().endswith('.pt')):
                p = p + ".pth"
            self.save_edit.setText(p)
            if self.parent_app:
                self.parent_app.set_last_directory(p)

    def start_training(self):
        img_dir = self.img_dir_edit.text().strip()
        mask_dir = self.mask_dir_edit.text().strip() or img_dir
        save_path = self.save_edit.text().strip()
        if not img_dir or not os.path.isdir(img_dir):
            QMessageBox.warning(self, "提示", "请选择有效的图像目录")
            return
        if not mask_dir or not os.path.isdir(mask_dir):
            QMessageBox.warning(self, "提示", "请选择有效的掩码目录")
            return
        if not save_path:
            save_path = os.path.join(os.path.dirname(img_dir), "U-Net训练结果.pth")
            self.save_edit.setText(save_path)
        if not save_path.lower().endswith(('.pt', '.pth')):
            save_path = save_path.rstrip('.') + ".pth"
            self.save_edit.setText(save_path)
        # 保存本次路径供下次一键恢复
        settings = QSettings("RockAnalysisTool", "Segmentation")
        settings.setValue("unet_train_img_dir", img_dir)
        settings.setValue("unet_train_mask_dir", mask_dir)
        settings.setValue("unet_train_save_path", save_path)
        use_cpu = self.unet_use_cpu_check.isChecked()
        settings.setValue("unet_train_use_cpu", use_cpu)
        # 在主线程预初始化 PyTorch/CUDA，避免子线程中首次初始化导致闪退
        try:
            import torch
            _ = torch.ones(1)
            if torch.cuda.is_available():
                _ = torch.ones(1, device="cuda")
        except Exception as e:
            self.log_text.append(f"PyTorch 预检查: {e}，将尝试继续训练（可能使用 CPU）")
        self.train_btn.setEnabled(False)
        self.log_text.clear()
        self.log_text.append("开始训练...")
        self.worker = SegModelTrainingWorker(
            img_dir, mask_dir, save_path,
            self.epochs_spin.value(), self.batch_spin.value(),
            self,
            use_cpu=use_cpu,
        )
        self.worker.log_signal.connect(self.log_text.append)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.start()

    def on_finished(self, ok, exit_code=0):
        self.train_btn.setEnabled(True)
        if ok:
            # 若存在训练曲线 JSON，则绘制
            save_path = self.save_edit.text().strip()
            if save_path:
                import json
                base = save_path.rsplit(".", 1)[0]
                history_path = base + "_history.json"
                if os.path.isfile(history_path):
                    try:
                        with open(history_path, "r", encoding="utf-8") as f:
                            data = json.load(f)
                        self._plot_training_curve(data.get("train_loss", []))
                    except Exception as e:
                        self.log_text.append(f"加载训练曲线失败: {e}")
            QMessageBox.information(self, "完成", "模型训练完成。请在 设置→分割模型 中指定该模型路径。")
        else:
            if exit_code == 3221225477:
                QMessageBox.warning(
                    self, "训练异常退出",
                    "训练子进程闪退（退出码 3221225477）。\n\n"
                    "若未勾选 CPU：多为显卡与当前 PyTorch/CUDA 或驱动不兼容，可先勾选「训练时强制使用 CPU」再试。\n\n"
                    "若已勾选 CPU 仍闪退：当前 PyTorch 为带 CUDA 的版本，加载时仍会崩溃。请在本机命令行执行：\n"
                    "  pip install torch --index-url https://download.pytorch.org/whl/cpu\n"
                    "安装纯 CPU 版 PyTorch 后重试。\n\n"
                    "日志已保存到程序目录下 logs 文件夹。"
                )
            else:
                QMessageBox.warning(self, "失败", "训练未成功，请查看日志。")

    def _plot_training_curve(self, train_loss):
        """绘制 U-Net 训练损失曲线（论文可用）"""
        self.curve_fig.clear()
        if not train_loss:
            ax = self.curve_fig.add_subplot(111)
            ax.set_xticks([])
            ax.set_yticks([])
            ax.text(0.5, 0.5, "无曲线数据", ha="center", va="center", transform=ax.transAxes)
            self.curve_canvas.draw()
            return
        ax = self.curve_fig.add_subplot(111)
        epochs = list(range(1, len(train_loss) + 1))
        ax.plot(epochs, train_loss, "b-", label="Train Loss", linewidth=1.5)
        ax.set_xlabel("Epoch")
        ax.set_ylabel("Loss")
        ax.legend()
        ax.grid(True, alpha=0.3)
        self.curve_fig.tight_layout()
        self.curve_canvas.draw()


# ========== 批量分割对话框 ==========
class BatchSegmentationDialog(QDialog):
    """批量分割对话框"""

    def __init__(self, image_files, dataset_path, parent=None):
        super().__init__(parent)
        self.image_files = image_files
        self.dataset_path = dataset_path
        self.parent_app = parent
        self.worker = None

        self.setWindowTitle("Batch Segmentation")
        self.setModal(True)
        self.resize(600, 500)

        self.init_ui()

    def init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # 标题
        title_label = QLabel("Batch Image Segmentation")
        title_label.setStyleSheet("""
            font-weight: bold;
            font-size: 16px;
            color: #2c3e50;
            padding: 10px;
        """)
        layout.addWidget(title_label)

        # 信息显示
        info_group = QGroupBox("Dataset Information")
        info_layout = QVBoxLayout()
        info_layout.addWidget(QLabel(f"Dataset: {os.path.basename(self.dataset_path)}"))
        info_layout.addWidget(QLabel(f"Total Images: {len(self.image_files)}"))
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)

        # 算法选择
        method_group = QGroupBox("Select Segmentation Methods")
        method_layout = QVBoxLayout()

        # 使用现有的SegmentationMethodDialog的逻辑
        self.method_checks = {}
        methods = [
            ('GrabCut智能分割', '使用GrabCut算法，适合复杂背景'),
            ('颜色阈值分割', '基于颜色范围，适合颜色明显的岩石'),
            ('边缘检测分割', '基于边缘检测，适合边界清晰的岩石'),
            ('自适应阈值分割', '自适应阈值，适合光照不均的图像'),
            ('分水岭分割', '分水岭算法，适合重叠区域'),
            ('K-means聚类分割', 'K-means聚类，适合颜色分布明显的图像'),
            ('深度学习分割', 'U-Net语义分割，需先训练模型')
        ]

        for method_name, description in methods:
            check = QCheckBox(method_name)
            check.setToolTip(description)
            check.setChecked(True)  # 默认全选
            method_layout.addWidget(check)
            self.method_checks[method_name] = check

        # 操作按钮
        btn_layout = QHBoxLayout()
        select_all_btn = QPushButton("Select All")
        select_all_btn.clicked.connect(lambda: [c.setChecked(True) for c in self.method_checks.values()])
        select_none_btn = QPushButton("Select None")
        select_none_btn.clicked.connect(lambda: [c.setChecked(False) for c in self.method_checks.values()])
        btn_layout.addWidget(select_all_btn)
        btn_layout.addWidget(select_none_btn)
        btn_layout.addStretch()
        method_layout.addLayout(btn_layout)

        method_group.setLayout(method_layout)
        layout.addWidget(method_group)

        # 输出目录与导出格式（论文用图：PNG/JPEG/TIFF/PDF）
        output_group = QGroupBox("Output Directory")
        output_layout = QVBoxLayout()

        output_layout_widget = QHBoxLayout()
        self.output_dir_edit = QLineEdit()
        self.output_dir_edit.setPlaceholderText("Select output directory...")
        output_layout_widget.addWidget(self.output_dir_edit)
        browse_btn = QPushButton("Browse...")
        browse_btn.clicked.connect(self.browse_output_dir)
        output_layout_widget.addWidget(browse_btn)
        output_layout.addLayout(output_layout_widget)

        fmt_layout = QHBoxLayout()
        fmt_layout.addWidget(QLabel("Export format:"))
        self.export_format_combo = QComboBox()
        self.export_format_combo.addItems(["PNG (无损)", "JPEG (高画质)", "TIFF (印刷)", "PDF (矢量/高清)"])
        fmt_layout.addWidget(self.export_format_combo)
        fmt_layout.addStretch()
        output_layout.addLayout(fmt_layout)

        output_group.setLayout(output_layout)
        layout.addWidget(output_group)

        # 进度显示
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)

        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        layout.addWidget(self.status_label)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        self.start_btn = QPushButton("Start")
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 8px 20px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        self.start_btn.clicked.connect(self.start_segmentation)
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(self.start_btn)
        btn_layout.addWidget(self.cancel_btn)
        layout.addLayout(btn_layout)

    def browse_output_dir(self):
        """浏览输出目录"""
        last_dir = self.output_dir_edit.text() if self.output_dir_edit.text() else self.dataset_path
        dir_path = QFileDialog.getExistingDirectory(
            self,
            "Select Output Directory",
            last_dir
        )
        if dir_path:
            self.output_dir_edit.setText(dir_path)

    def start_segmentation(self):
        """开始分割"""
        # 检查选中的方法
        selected_methods = [name for name, check in self.method_checks.items() if check.isChecked()]
        if not selected_methods:
            QMessageBox.warning(self, "No Method Selected", "Please select at least one segmentation method.")
            return

        # 检查输出目录
        output_dir = self.output_dir_edit.text().strip()
        if not output_dir:
            QMessageBox.warning(self, "No Output Directory", "Please select an output directory.")
            return

        # 创建输出目录
        os.makedirs(output_dir, exist_ok=True)

        # 禁用开始按钮
        self.start_btn.setEnabled(False)
        self.status_label.setText("Processing...")

        fmt_map = {"PNG (无损)": "png", "JPEG (高画质)": "jpg", "TIFF (印刷)": "tif", "PDF (矢量/高清)": "pdf"}
        export_fmt = fmt_map.get(self.export_format_combo.currentText(), "png")

        # 创建工作线程
        self.worker = BatchSegmentationWorker(
            self.image_files,
            output_dir,
            selected_methods,
            self.parent_app,
            export_fmt=export_fmt,
        )
        self.worker.progress_updated.connect(self.update_progress)
        self.worker.file_completed.connect(self.on_file_completed)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.error_occurred.connect(self.on_error)

        self.worker.start()

    def update_progress(self, value, filename):
        """更新进度"""
        self.progress_bar.setValue(value)
        self.status_label.setText(f"Processing: {filename}")
        QApplication.processEvents()

    def on_file_completed(self, file_path, results):
        """文件处理完成"""
        if self.parent_app:
            self.parent_app.log(f"Completed: {os.path.basename(file_path)}")

    def on_finished(self):
        """处理完成"""
        self.progress_bar.setValue(100)
        self.status_label.setText("Completed!")
        self.start_btn.setEnabled(True)
        QMessageBox.information(
            self,
            "Batch Segmentation Complete",
            f"Successfully processed {len(self.image_files)} images."
        )

    def on_error(self, error_msg):
        """处理错误"""
        if self.parent_app:
            self.parent_app.log(f"Error: {error_msg}")
        QMessageBox.warning(self, "Error", error_msg)

    def closeEvent(self, event):
        """关闭事件"""
        if self.worker and self.worker.isRunning():
            reply = QMessageBox.question(
                self,
                "Segmentation in Progress",
                "Segmentation is still running. Do you want to stop it?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.worker.stop()
                self.worker.wait()
            else:
                event.ignore()
                return
        event.accept()


# ========== 自动标注对话框 ==========
class AutoLabelingDialog(QDialog):
    """机器学习前自动识别/标注数据集"""

    def __init__(self, image_files, dataset_path, parent=None):
        super().__init__(parent)
        self.image_files = image_files
        self.dataset_path = dataset_path
        self.parent_app = parent
        self.labels = {}  # img_path -> label

        self.setWindowTitle("自动标注 - 数据集标注")
        self.setModal(True)
        self.resize(650, 520)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)

        header = QLabel("自动标注方式（可多选前两种，手动可编辑）")
        header.setStyleSheet("font-weight: bold; font-size: 13px;")
        layout.addWidget(header)

        # 识别方式
        method_group = QGroupBox("识别方式")
        method_layout = QVBoxLayout()
        self.radio_folder = QRadioButton("按文件夹名：使用子文件夹名作为标签（推荐）")
        self.radio_folder.setChecked(True)
        self.radio_folder.setToolTip("若图片在 dataset/红色/ 下，则标签为「红色」")
        method_layout.addWidget(self.radio_folder)

        self.radio_color = QRadioButton("颜色识别：对每张图进行颜色分析，取主色为标签")
        self.radio_color.setToolTip("调用岩石颜色识别，将识别结果中占比最高的颜色作为标签")
        method_layout.addWidget(self.radio_color)

        method_layout.addWidget(QLabel("手动标注：在下方表格中直接编辑标签"))
        method_group.setLayout(method_layout)
        layout.addWidget(method_group)

        # 预览表格
        table_group = QGroupBox("标注预览（可编辑）")
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["图片", "当前标签", "操作"])
        self.table.horizontalHeader().setStretchLastSection(True)
        table_group_layout = QVBoxLayout()
        table_group_layout.addWidget(self.table)
        layout.addWidget(table_group)

        # 按钮
        btn_layout = QHBoxLayout()
        gen_btn = QPushButton("生成/更新标签")
        gen_btn.clicked.connect(self.generate_labels)
        export_btn = QPushButton("导出标注CSV")
        export_btn.clicked.connect(self.export_labels)
        btn_layout.addWidget(gen_btn)
        btn_layout.addWidget(export_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        self.generate_labels()

    def generate_labels(self):
        """根据选择的方式生成标签"""
        self.table.setRowCount(0)
        display_count = min(200, len(self.image_files))
        for i, img_path in enumerate(self.image_files):
            label = ""
            if self.radio_folder.isChecked():
                rel = os.path.relpath(os.path.dirname(img_path), self.dataset_path)
                label = rel if rel != "." else os.path.basename(self.dataset_path)
            elif self.radio_color.isChecked():
                label = self._recognize_color(img_path)
            self.labels[img_path] = label
            if i < display_count:
                row = self.table.rowCount()
                self.table.insertRow(row)
                self.table.setItem(row, 0, QTableWidgetItem(os.path.basename(img_path)))
                item = QTableWidgetItem(label)
                self.table.setItem(row, 1, item)
        if len(self.image_files) > 200 and self.parent_app:
            self.parent_app.log(f"自动标注表格显示前200条，共{len(self.image_files)}张已生成标签")

    def _recognize_color(self, img_path):
        """对单张图进行颜色识别，返回主色标签（简化版：中心区域与标准色最近）"""
        try:
            img = cv2_imread(img_path)
            if img is None:
                return "unknown"
            if not self.parent_app or not hasattr(self.parent_app,
                                                  'standard_vectors') or self.parent_app.standard_vectors is None:
                return "unknown"
            self.parent_app.load_csv_data()
            std = self.parent_app.standard_vectors
            names = self.parent_app.color_names
            if std is None or not len(names):
                return "unknown"
            h, w = img.shape[:2]
            cx, cy = w // 2, h // 2
            roi = img[max(0, cy - 20):cy + 20, max(0, cx - 20):cx + 20]
            mean_bgr = np.mean(roi.reshape(-1, 3), axis=0)
            mean_rgb = mean_bgr[::-1]  # BGR -> RGB
            min_dist, best = float('inf'), "unknown"
            for j, vec in enumerate(std):
                d = np.linalg.norm(mean_rgb - np.array(vec[:3]))
                if d < min_dist:
                    min_dist, best = d, str(names[j])
            return best
        except Exception:
            return "unknown"

    def export_labels(self):
        """导出标注为CSV"""
        for row in range(self.table.rowCount()):
            path_item = self.table.item(row, 0)
            label_item = self.table.item(row, 1)
            if path_item and label_item:
                for p in self.image_files:
                    if os.path.basename(p) == path_item.text():
                        self.labels[p] = label_item.text()
                        break
        out_path, _ = QFileDialog.getSaveFileName(self, "保存标注CSV", "", "CSV (*.csv)")
        if not out_path:
            return
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write("img_name,岩石颜色\n")
            for p in self.image_files:
                f.write(f"{os.path.basename(p)},{self.labels.get(p, '')}\n")
        QMessageBox.information(self, "完成", f"已保存至 {out_path}")


# ========== 训练数据准备对话框 ==========
class TrainingDataPreparationDialog(QDialog):
    """训练数据准备对话框"""

    def __init__(self, image_files, dataset_path, parent=None):
        super().__init__(parent)
        self.image_files = image_files
        self.dataset_path = dataset_path
        self.parent_app = parent

        self.setWindowTitle("Training Data Preparation")
        self.setModal(True)
        self.resize(700, 600)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # 标题
        title_label = QLabel("Prepare Training / Validation / Test Datasets")
        title_label.setStyleSheet("""
            font-weight: bold;
            font-size: 16px;
            color: #2c3e50;
            padding: 10px;
        """)
        layout.addWidget(title_label)

        # 数据集信息
        info_group = QGroupBox("Dataset Information")
        info_layout = QVBoxLayout()
        info_layout.addWidget(QLabel(f"Dataset: {os.path.basename(self.dataset_path)}"))
        info_layout.addWidget(QLabel(f"Total Images: {len(self.image_files)}"))
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)

        # 标签策略
        label_group = QGroupBox("Labeling Strategy")
        label_layout = QVBoxLayout()

        self.radio_folder_label = QRadioButton("Use first-level subfolder name as class label (recommended)")
        self.radio_folder_label.setChecked(True)
        self.radio_single_label = QRadioButton("Use same label for all images")

        label_layout.addWidget(self.radio_folder_label)
        single_label_layout = QHBoxLayout()
        single_label_layout.addWidget(self.radio_single_label)
        single_label_layout.addWidget(QLabel("Label:"))
        self.single_label_edit = QLineEdit()
        self.single_label_edit.setPlaceholderText("e.g., Rock")
        self.single_label_edit.setEnabled(False)
        single_label_layout.addWidget(self.single_label_edit)
        single_label_layout.addStretch()
        label_layout.addLayout(single_label_layout)

        # 切换单标签输入可用状态
        self.radio_single_label.toggled.connect(self.single_label_edit.setEnabled)

        label_group.setLayout(label_layout)
        layout.addWidget(label_group)

        # 数据集划分
        split_group = QGroupBox("Dataset Split Ratio (%)")
        split_layout = QHBoxLayout()

        split_layout.addWidget(QLabel("Train:"))
        self.train_ratio_spin = QDoubleSpinBox()
        self.train_ratio_spin.setRange(0, 100)
        self.train_ratio_spin.setValue(70.0)
        self.train_ratio_spin.setDecimals(1)
        split_layout.addWidget(self.train_ratio_spin)

        split_layout.addWidget(QLabel("Val:"))
        self.val_ratio_spin = QDoubleSpinBox()
        self.val_ratio_spin.setRange(0, 100)
        self.val_ratio_spin.setValue(20.0)
        self.val_ratio_spin.setDecimals(1)
        split_layout.addWidget(self.val_ratio_spin)

        split_layout.addWidget(QLabel("Test:"))
        self.test_ratio_spin = QDoubleSpinBox()
        self.test_ratio_spin.setRange(0, 100)
        self.test_ratio_spin.setValue(10.0)
        self.test_ratio_spin.setDecimals(1)
        split_layout.addWidget(self.test_ratio_spin)

        split_layout.addStretch()
        split_group.setLayout(split_layout)
        layout.addWidget(split_group)

        # 输出目录
        output_group = QGroupBox("Output Root Directory")
        output_layout = QVBoxLayout()
        h = QHBoxLayout()
        self.output_root_edit = QLineEdit()
        default_root = os.path.join(self.dataset_path, "training_data")
        self.output_root_edit.setText(default_root)
        h.addWidget(self.output_root_edit)
        browse_btn = QPushButton("Browse...")
        browse_btn.clicked.connect(self.browse_output_root)
        h.addWidget(browse_btn)
        output_layout.addLayout(h)
        output_group.setLayout(output_layout)
        layout.addWidget(output_group)

        # 统计/预览
        stats_group = QGroupBox("Preview & Statistics")
        stats_layout = QVBoxLayout()
        self.stats_text = QTextEdit()
        self.stats_text.setReadOnly(True)
        self.stats_text.setStyleSheet("""
            QTextEdit {
                border: 1px solid #dcdfe6;
                border-radius: 4px;
                background-color: #f8f9fa;
                font-family: Consolas, monospace;
                font-size: 11px;
            }
        """)
        stats_layout.addWidget(self.stats_text)

        preview_btn = QPushButton("Preview Split")
        preview_btn.clicked.connect(self.preview_split)
        stats_layout.addWidget(preview_btn, alignment=Qt.AlignmentFlag.AlignRight)

        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        ok_btn = QPushButton("Start")
        ok_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 8px 20px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        ok_btn.clicked.connect(self.start_preparation)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

    def browse_output_root(self):
        """浏览输出根目录"""
        last_dir = self.output_root_edit.text() or self.dataset_path
        dir_path = QFileDialog.getExistingDirectory(
            self,
            "Select Output Root Directory",
            last_dir
        )
        if dir_path:
            self.output_root_edit.setText(dir_path)

    def compute_label_mapping(self):
        """根据选择的策略生成 label -> [image_paths] 映射"""
        label_to_images = {}

        if self.radio_folder_label.isChecked():
            for img_path in self.image_files:
                rel = os.path.relpath(os.path.dirname(img_path), self.dataset_path)
                if rel == ".":
                    label = "unknown"
                else:
                    # 使用第一级子目录作为标签
                    label = rel.split(os.sep)[0]
                label_to_images.setdefault(label, []).append(img_path)
        else:
            label = self.single_label_edit.text().strip() or "Rock"
            label_to_images[label] = list(self.image_files)

        return label_to_images

    def preview_split(self):
        """预览划分结果"""
        label_to_images = self.compute_label_mapping()
        train_r = self.train_ratio_spin.value()
        val_r = self.val_ratio_spin.value()
        test_r = self.test_ratio_spin.value()
        total_ratio = train_r + val_r + test_r
        if total_ratio <= 0:
            QMessageBox.warning(self, "Invalid Ratio", "Total ratio must be greater than 0.")
            return

        train_r /= total_ratio
        val_r /= total_ratio
        test_r /= total_ratio

        lines = []
        total_images = len(self.image_files)
        lines.append(f"Total images: {total_images}")
        lines.append(f"Labels: {len(label_to_images)} -> {', '.join(label_to_images.keys())}")
        lines.append(
            f"Split ratio (normalized): Train {train_r * 100:.1f}%, Val {val_r * 100:.1f}%, Test {test_r * 100:.1f}%")
        lines.append("")
        lines.append("Per-label split preview:")

        for label, imgs in label_to_images.items():
            n = len(imgs)
            n_train = int(n * train_r)
            n_val = int(n * val_r)
            n_test = n - n_train - n_val
            lines.append(f"  {label}: {n} -> train {n_train}, val {n_val}, test {n_test}")

        self.stats_text.setPlainText("\n".join(lines))

    def start_preparation(self):
        """开始实际的数据划分和拷贝"""
        label_to_images = self.compute_label_mapping()
        if not label_to_images:
            QMessageBox.warning(self, "No Data", "No images to process.")
            return

        output_root = self.output_root_edit.text().strip()
        if not output_root:
            QMessageBox.warning(self, "No Output Directory", "Please select output root directory.")
            return

        train_r = self.train_ratio_spin.value()
        val_r = self.val_ratio_spin.value()
        test_r = self.test_ratio_spin.value()
        total_ratio = train_r + val_r + test_r
        if total_ratio <= 0:
            QMessageBox.warning(self, "Invalid Ratio", "Total ratio must be greater than 0.")
            return

        train_r /= total_ratio
        val_r /= total_ratio
        test_r /= total_ratio

        # 目标结构: output_root/train/label/, val/label/, test/label/
        train_dir = os.path.join(output_root, "train")
        val_dir = os.path.join(output_root, "val")
        test_dir = os.path.join(output_root, "test")
        os.makedirs(train_dir, exist_ok=True)
        os.makedirs(val_dir, exist_ok=True)
        os.makedirs(test_dir, exist_ok=True)

        train_rows = []
        val_rows = []
        test_rows = []

        for label, imgs in label_to_images.items():
            imgs_copy = list(imgs)
            random.shuffle(imgs_copy)
            n = len(imgs_copy)
            n_train = int(n * train_r)
            n_val = int(n * val_r)
            n_test = n - n_train - n_val

            train_imgs = imgs_copy[:n_train]
            val_imgs = imgs_copy[n_train:n_train + n_val]
            test_imgs = imgs_copy[n_train + n_val:]

            # 创建标签子文件夹
            label_train_dir = os.path.join(train_dir, label)
            label_val_dir = os.path.join(val_dir, label)
            label_test_dir = os.path.join(test_dir, label)
            os.makedirs(label_train_dir, exist_ok=True)
            os.makedirs(label_val_dir, exist_ok=True)
            os.makedirs(label_test_dir, exist_ok=True)

            # 拷贝并记录 CSV 行
            for src in train_imgs:
                dst = os.path.join(label_train_dir, os.path.basename(src))
                shutil.copy2(src, dst)
                train_rows.append({"img_name": os.path.basename(src), "岩石颜色": label})

            for src in val_imgs:
                dst = os.path.join(label_val_dir, os.path.basename(src))
                shutil.copy2(src, dst)
                val_rows.append({"img_name": os.path.basename(src), "岩石颜色": label})

            for src in test_imgs:
                dst = os.path.join(label_test_dir, os.path.basename(src))
                shutil.copy2(src, dst)
                test_rows.append({"img_name": os.path.basename(src), "岩石颜色": label})

        # 保存 CSV
        data_dir = output_root
        train_csv_path = os.path.join(data_dir, "train_labels.csv")
        val_csv_path = os.path.join(data_dir, "val_labels.csv")
        test_csv_path = os.path.join(data_dir, "test_labels.csv")

        if train_rows:
            pd.DataFrame(train_rows).to_csv(train_csv_path, index=False, encoding="utf-8-sig")
        if val_rows:
            pd.DataFrame(val_rows).to_csv(val_csv_path, index=False, encoding="utf-8-sig")
        if test_rows:
            pd.DataFrame(test_rows).to_csv(test_csv_path, index=False, encoding="utf-8-sig")

        if self.parent_app:
            self.parent_app.log(f"Training data prepared at: {output_root}")
            self.parent_app.log(f"Train CSV: {train_csv_path}")
            self.parent_app.log(f"Val CSV: {val_csv_path}")
            self.parent_app.log(f"Test CSV: {test_csv_path}")

        QMessageBox.information(
            self,
            "Success",
            "Training / validation / test datasets have been prepared successfully."
        )
        self.accept()


# ========== 数据集信息窗口 ==========
class DatasetInfoWindow(QMainWindow):
    """（已弃用）旧版数据集信息窗口（仅保留代码，不再使用）"""

    def __init__(self, dataset_path, parent=None):
        super().__init__(parent)
        self.dataset_path = dataset_path
        self.parent_app = parent
        self.image_files = []
        # 预览分页相关
        self.preview_page_size = 20
        self.current_preview_page = 0
        self.show_all_preview = False

        self.setWindowTitle(f"Dataset Info - {os.path.basename(dataset_path)}")
        self.setWindowIcon(QIcon(r"resources\assets\images\button\addFolder.png"))
        self.resize(1000, 700)

        self.init_ui()
        self.scan_dataset()

    def init_ui(self):
        """初始化UI"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)

        # 标题栏
        title_frame = QFrame()
        title_frame.setStyleSheet("background-color: #2c3e50; border-radius: 5px; padding: 15px;")
        title_layout = QHBoxLayout(title_frame)

        title_label = QLabel(f"Dataset: {os.path.basename(self.dataset_path)}")
        title_label.setStyleSheet("color: white; font-weight: bold; font-size: 16px;")
        title_layout.addWidget(title_label)
        title_layout.addStretch()

        # 操作按钮
        batch_seg_btn = QPushButton("Batch Segmentation")
        batch_seg_btn.setIcon(QIcon(r"resources\assets\images\button\segmentation.png"))
        batch_seg_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        batch_seg_btn.clicked.connect(self.start_batch_segmentation)
        title_layout.addWidget(batch_seg_btn)

        prep_data_btn = QPushButton("Prepare Training Data")
        prep_data_btn.setIcon(QIcon(r"resources\assets\images\button\save.png"))
        prep_data_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        prep_data_btn.clicked.connect(self.prepare_training_data)
        title_layout.addWidget(prep_data_btn)

        layout.addWidget(title_frame)

        # 信息显示区域（使用标签页）
        info_tabs = QTabWidget()
        info_tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #ddd;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #ecf0f1;
                color: #34495e;
                padding: 8px 20px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background-color: #3498db;
                color: white;
            }
        """)

        # 统计信息标签页
        stats_tab = self.create_stats_tab()
        info_tabs.addTab(stats_tab, "Statistics")

        # 文件夹结构标签页
        structure_tab = self.create_structure_tab()
        info_tabs.addTab(structure_tab, "Folder Structure")

        # 图片预览标签页
        preview_tab = self.create_preview_tab()
        info_tabs.addTab(preview_tab, "Image Preview")

        layout.addWidget(info_tabs)

    def create_stats_tab(self):
        """创建统计信息标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        # 基本信息表格
        self.stats_table = QTableWidget()
        self.stats_table.setColumnCount(2)
        self.stats_table.setHorizontalHeaderLabels(["Property", "Value"])
        self.stats_table.horizontalHeader().setStretchLastSection(True)
        self.stats_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.stats_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 5px;
                gridline-color: #e0e0e0;
            }
            QTableWidget::item {
                padding: 5px;
            }
        """)

        layout.addWidget(self.stats_table)
        return widget

    def create_structure_tab(self):
        """创建文件夹结构标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        self.structure_tree = QTreeWidget()
        self.structure_tree.setHeaderLabel("Folder Structure")
        self.structure_tree.setStyleSheet("""
            QTreeWidget {
                border: 1px solid #ddd;
                border-radius: 5px;
            }
        """)

        layout.addWidget(self.structure_tree)
        return widget

    def create_preview_tab(self):
        """创建图片预览标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("border: 1px solid #ddd; border-radius: 5px;")

        preview_widget = QWidget()
        self.preview_layout = QGridLayout(preview_widget)
        self.preview_layout.setSpacing(10)

        scroll_area.setWidget(preview_widget)
        layout.addWidget(scroll_area)

        # 分页控制区域
        ctrl_layout = QHBoxLayout()
        self.preview_page_label = QLabel("第 1 页 / 共 1 页")
        self.preview_page_label.setStyleSheet("color: #34495e;")
        ctrl_layout.addWidget(self.preview_page_label)
        ctrl_layout.addStretch()

        self.prev_preview_btn = QPushButton("上一页")
        self.prev_preview_btn.clicked.connect(self.prev_preview_page)
        ctrl_layout.addWidget(self.prev_preview_btn)

        self.next_preview_btn = QPushButton("下一页")
        self.next_preview_btn.clicked.connect(self.next_preview_page)
        ctrl_layout.addWidget(self.next_preview_btn)

        load_all_btn = QPushButton("加载全部")
        load_all_btn.clicked.connect(self.load_all_previews)
        ctrl_layout.addWidget(load_all_btn)

        layout.addLayout(ctrl_layout)

        return widget

    def scan_dataset(self):
        """扫描数据集"""
        supported_formats = ('.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.tif', '.webp')
        self.image_files = []

        for root, dirs, files in os.walk(self.dataset_path):
            for file in files:
                if file.lower().endswith(supported_formats):
                    self.image_files.append(os.path.join(root, file))

        self.update_statistics()
        self.update_structure()
        self.update_preview()

    def update_statistics(self):
        """更新统计信息"""
        # 统计信息
        total_images = len(self.image_files)
        formats = {}
        folders = {}

        for img_path in self.image_files:
            ext = os.path.splitext(img_path)[1].lower()
            formats[ext] = formats.get(ext, 0) + 1

            folder = os.path.dirname(img_path)
            if folder == self.dataset_path:
                folder_name = "Root"
            else:
                folder_name = os.path.relpath(folder, self.dataset_path)
            folders[folder_name] = folders.get(folder_name, 0) + 1

        # 更新表格
        self.stats_table.setRowCount(6 + len(formats) + len(folders))
        row = 0

        self.stats_table.setItem(row, 0, QTableWidgetItem("Dataset Path"))
        self.stats_table.setItem(row, 1, QTableWidgetItem(self.dataset_path))
        row += 1

        self.stats_table.setItem(row, 0, QTableWidgetItem("Total Images"))
        self.stats_table.setItem(row, 1, QTableWidgetItem(str(total_images)))
        row += 1

        self.stats_table.setItem(row, 0, QTableWidgetItem("Image Formats"))
        self.stats_table.setItem(row, 1, QTableWidgetItem(", ".join(formats.keys())))
        row += 1

        self.stats_table.setItem(row, 0, QTableWidgetItem("Format Distribution"))
        format_str = ", ".join([f"{k}: {v}" for k, v in formats.items()])
        self.stats_table.setItem(row, 1, QTableWidgetItem(format_str))
        row += 1

        self.stats_table.setItem(row, 0, QTableWidgetItem("Number of Folders"))
        self.stats_table.setItem(row, 1, QTableWidgetItem(str(len(folders))))
        row += 1

        self.stats_table.setItem(row, 0, QTableWidgetItem("Folder Distribution"))
        folder_str = ", ".join([f"{k}: {v}" for k, v in sorted(folders.items(), key=lambda x: x[1], reverse=True)[:10]])
        if len(folders) > 10:
            folder_str += f" ... (and {len(folders) - 10} more)"
        self.stats_table.setItem(row, 1, QTableWidgetItem(folder_str))

    def update_structure(self):
        """更新文件夹结构"""
        self.structure_tree.clear()

        root_item = QTreeWidgetItem(self.structure_tree)
        root_item.setText(0, os.path.basename(self.dataset_path))
        root_item.setExpanded(True)

        # 构建树形结构
        folder_dict = {}
        for img_path in self.image_files:
            folder = os.path.dirname(img_path)
            if folder not in folder_dict:
                folder_dict[folder] = []
            folder_dict[folder].append(os.path.basename(img_path))

        # 添加文件夹和文件（显示全部文件，不再只显示前20个）
        for folder, files in sorted(folder_dict.items()):
            rel_path = os.path.relpath(folder, self.dataset_path)
            if rel_path == '.':
                parent = root_item
            else:
                parts = rel_path.split(os.sep)
                parent = root_item
                for part in parts:
                    found = False
                    for i in range(parent.childCount()):
                        if parent.child(i).text(0) == part:
                            parent = parent.child(i)
                            found = True
                            break
                    if not found:
                        new_item = QTreeWidgetItem(parent)
                        new_item.setText(0, part)
                        parent = new_item

            # 添加文件（全部显示）
            for file in files:
                file_item = QTreeWidgetItem(parent)
                file_item.setText(0, file)

    def update_preview(self):
        """更新图片预览"""
        # 清除现有预览
        for i in reversed(range(self.preview_layout.count())):
            self.preview_layout.itemAt(i).widget().setParent(None)

        total = len(self.image_files)
        if total == 0:
            # 更新分页标签
            if hasattr(self, "preview_page_label"):
                self.preview_page_label.setText("第 0 页 / 共 0 页 (共 0 张)")
            return

        # 计算分页
        if self.show_all_preview:
            start_index = 0
            end_index = total
            current_page = 1
            total_pages = 1
        else:
            page_size = self.preview_page_size
            total_pages = (total + page_size - 1) // page_size
            if self.current_preview_page >= total_pages:
                self.current_preview_page = max(0, total_pages - 1)
            current_page = self.current_preview_page + 1
            start_index = self.current_preview_page * page_size
            end_index = min(start_index + page_size, total)

        # 更新分页标签
        if hasattr(self, "preview_page_label"):
            self.preview_page_label.setText(
                f"第 {current_page} 页 / 共 {total_pages} 页 (共 {total} 张)"
            )

        # 根据分页状态更新按钮可用性
        if hasattr(self, "prev_preview_btn") and hasattr(self, "next_preview_btn"):
            if self.show_all_preview or total_pages <= 1:
                self.prev_preview_btn.setEnabled(False)
                self.next_preview_btn.setEnabled(False)
            else:
                self.prev_preview_btn.setEnabled(current_page > 1)
                self.next_preview_btn.setEnabled(current_page < total_pages)

        cols = 4

        for idx, img_path in enumerate(self.image_files[start_index:end_index]):
            try:
                # 使用 PIL 加载图片（更好地支持中文路径，且避免 OpenCV 控制台警告）
                pil_img = Image.open(img_path)
                if pil_img.mode != "RGB":
                    pil_img = pil_img.convert("RGB")

                # 创建缩略图
                thumb_size = 150
                pil_img.thumbnail((thumb_size, thumb_size))
                img_array = np.array(pil_img)  # RGB

                # 转换为QPixmap
                h, w, ch = img_array.shape
                bytes_per_line = ch * w
                q_image = QImage(img_array.data, w, h, bytes_per_line, QImage.Format.Format_RGB888)
                pixmap = QPixmap.fromImage(q_image)

                # 创建标签
                label = QLabel()
                label.setPixmap(pixmap)
                label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                label.setStyleSheet("""
                    QLabel {
                        border: 2px solid #ddd;
                        border-radius: 5px;
                        padding: 5px;
                        background-color: white;
                    }
                    QLabel:hover {
                        border: 2px solid #3498db;
                    }
                """)
                label.setToolTip(os.path.basename(img_path))

                row = idx // cols
                col = idx % cols
                self.preview_layout.addWidget(label, row, col)
            except Exception as e:
                self.parent_app.log(f"Error loading preview for {img_path}: {str(e)}")

    def prev_preview_page(self):
        """上一页预览"""
        if self.show_all_preview:
            # 退出“加载全部”模式，回到分页模式的最后一页
            self.show_all_preview = False
        if self.current_preview_page > 0:
            self.current_preview_page -= 1
        self.update_preview()

    def next_preview_page(self):
        """下一页预览"""
        if self.show_all_preview:
            # 退出“加载全部”模式，回到分页模式第一页
            self.show_all_preview = False
            self.current_preview_page = 0
        total = len(self.image_files)
        if total > 0:
            total_pages = (total + self.preview_page_size - 1) // self.preview_page_size
            if self.current_preview_page < total_pages - 1:
                self.current_preview_page += 1
        self.update_preview()

    def load_all_previews(self):
        """加载全部预览"""
        self.show_all_preview = True
        self.current_preview_page = 0
        self.update_preview()

    def start_batch_segmentation(self):
        """启动批量分割"""
        if not self.image_files:
            QMessageBox.warning(self, "No Images", "No images found in the dataset.")
            return

        # 调用批量分割对话框
        dialog = BatchSegmentationDialog(self.image_files, self.dataset_path, self.parent_app)
        dialog.exec()

    def prepare_training_data(self):
        """准备训练数据"""
        if not self.image_files:
            QMessageBox.warning(self, "No Images", "No images found in the dataset.")
            return

        # 调用训练数据准备对话框
        dialog = TrainingDataPreparationDialog(self.image_files, self.dataset_path, self.parent_app)
        dialog.exec()


# ========== 模型训练后台线程 ==========
class ModelTrainingWorker(QThread):
    """模型训练后台线程，调用 2模型训练.py 中的 ModelTrainer"""
    log_message = pyqtSignal(str)
    finished_signal = pyqtSignal(dict)
    error_signal = pyqtSignal(str)

    def __init__(self, train_csv, train_dir, val_csv, val_dir,
                 model_type, epochs, batch_size, patience,
                 enable_tensorboard, tensorboard_port, tensorboard_log_dir,
                 parent=None):
        super().__init__(parent)
        self.train_csv = train_csv
        self.train_dir = train_dir
        self.val_csv = val_csv
        self.val_dir = val_dir
        self.model_type = model_type
        self.epochs = epochs
        self.batch_size = batch_size
        self.patience = patience
        self.enable_tensorboard = enable_tensorboard
        self.tensorboard_port = tensorboard_port
        self.tensorboard_log_dir = tensorboard_log_dir

    def run(self):
        try:
            self.log_message.emit("Loading training module (model_train.py)...")

            # 重定向 stdout/stderr 到训练日志
            class StdoutRedirector:
                def __init__(self, emit_fn):
                    self.emit_fn = emit_fn

                def write(self, text):
                    if text and text.strip():
                        for line in text.rstrip().split('\n'):
                            if line.strip():
                                self.emit_fn(line)

                def flush(self):
                    pass

            old_stdout, old_stderr = sys.stdout, sys.stderr
            sys.stdout = StdoutRedirector(self.log_message.emit)
            sys.stderr = StdoutRedirector(self.log_message.emit)

            try:
                base_dir = os.path.dirname(os.path.abspath(__file__))
                ml_path = os.path.join(base_dir, "model_train.py")
                if not os.path.exists(ml_path):
                    raise FileNotFoundError(f"model_train.py not found at {ml_path}")

                spec = importlib.util.spec_from_file_location("rock_ml", ml_path)
                ml = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(ml)

                # 覆盖 Config 的相关目录，使其更贴近当前工程
                try:
                    ml.Config.DATA_DIR = os.path.dirname(self.train_csv)
                    ml.Config.TRAIN_DIR = self.train_dir
                    ml.Config.VAL_DIR = self.val_dir
                    if self.tensorboard_log_dir:
                        ml.Config.TENSORBOARD_DIR = self.tensorboard_log_dir
                except Exception:
                    pass

                trainer = ml.ModelTrainer()

                self.log_message.emit(
                    f"Start training model: {self.model_type}, epochs={self.epochs}, "
                    f"batch_size={self.batch_size}, patience={self.patience}"
                )
                self.log_message.emit(f"Train CSV: {self.train_csv}")
                self.log_message.emit(f"Val CSV: {self.val_csv}")
                self.log_message.emit(f"Train Dir: {self.train_dir}")
                self.log_message.emit(f"Val Dir: {self.val_dir}")

                # 调用训练函数
                results = trainer.train_model(
                    self.train_csv,
                    self.train_dir,
                    self.val_csv,
                    self.val_dir,
                    model_type=self.model_type,
                    epochs=self.epochs,
                    batch_size=self.batch_size,
                    patience=self.patience,
                )

                if not results:
                    self.error_signal.emit("Training did not return results. Please check logs.")
                    return

                self.log_message.emit(f"Training finished. Best val acc: {results.get('best_val_acc', 0):.2f}%")
                self.finished_signal.emit(results)
            finally:
                sys.stdout, sys.stderr = old_stdout, old_stderr
        except Exception as e:
            tb = traceback.format_exc()
            self.error_signal.emit(f"Training failed: {e}\n{tb}")


# ========== 模型训练窗口 ==========
class ModelTrainingWindow(QMainWindow):
    """模型训练窗口：调用 2模型训练.py，并在界面中可视化训练曲线"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_app = parent
        self.worker = None
        self.results = None

        self.setWindowTitle("模型训练 - 岩石颜色识别分类")
        self.setWindowIcon(QIcon(r"resources\assets\images\button\train.png"))
        self.resize(760, 630)

        self.init_ui()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        # 左侧：配置 + 日志
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(10)

        # 配置组
        config_group = QGroupBox("Training Configuration")
        config_layout = QVBoxLayout(config_group)

        # 路径配置
        self.train_csv_edit = QLineEdit()
        self.val_csv_edit = QLineEdit()
        self.train_dir_edit = QLineEdit()
        self.val_dir_edit = QLineEdit()

        # 尝试从最近的 training_data 推断默认路径
        default_root = ""
        if self.parent_app and hasattr(self.parent_app, "segmentation_folder") and self.parent_app.segmentation_folder:
            possible_root = os.path.join(self.parent_app.segmentation_folder, "training_data")
            if os.path.exists(possible_root):
                default_root = possible_root
        if default_root:
            self.train_csv_edit.setText(os.path.join(default_root, "train_labels.csv"))
            self.val_csv_edit.setText(os.path.join(default_root, "val_labels.csv"))
            self.train_dir_edit.setText(os.path.join(default_root, "train"))
            self.val_dir_edit.setText(os.path.join(default_root, "val"))
        else:
            self._load_last_training_paths()

        # 模型与超参数（含用途说明）
        model_row = QHBoxLayout()
        model_row.addWidget(QLabel("模型选择:"))
        self.model_combo = QComboBox()
        self.MODEL_DESCRIPTIONS = {
            "resnet": "ResNet：经典深度残差网络，适合图像分类。优点：结构成熟、泛化好。",
            "efficientnet": "EfficientNet：轻量高效，精度与速度平衡。优点：参数量少、推理快。",
            "mobilenet": "MobileNet：移动端优化，极轻量。优点：速度最快、适合部署。",
            "custom": "Custom CNN：自定义卷积网络。优点：可针对岩石数据定制。",
            "unet": "U-Net分割：语义分割模型，用于岩石区域分割。优点：边界精细、适合地质图像。",
        }
        self.model_combo.addItem("ResNet", "resnet")
        self.model_combo.addItem("EfficientNet", "efficientnet")
        self.model_combo.addItem("MobileNet", "mobilenet")
        self.model_combo.addItem("Custom CNN", "custom")
        self.model_combo.addItem("U-Net 分割", "unet")
        self.model_combo.currentIndexChanged.connect(self._on_model_type_changed)
        model_row.addWidget(self.model_combo)
        model_row.addStretch()
        config_layout.addLayout(model_row)
        self.model_desc_label = QLabel()
        self.model_desc_label.setStyleSheet("color: #7f8c8d; font-size: 11px;")
        self.model_desc_label.setWordWrap(True)
        config_layout.addWidget(self.model_desc_label)

        # 分类模型专用配置（ResNet/EfficientNet/MobileNet/Custom）
        self.class_config_widget = QWidget()
        class_config_layout = QVBoxLayout(self.class_config_widget)
        class_config_layout.setContentsMargins(0, 5, 0, 0)

        def add_path_row(label_text, line_edit, browse_slot):
            row = QHBoxLayout()
            row.addWidget(QLabel(label_text))
            row.addWidget(line_edit)
            btn = QPushButton("Browse")
            btn.clicked.connect(browse_slot)
            row.addWidget(btn)
            class_config_layout.addLayout(row)

        last_use_row = QHBoxLayout()
        last_use_btn = QPushButton("一键选择上次选择的文件夹")
        last_use_btn.clicked.connect(self.restore_last_training_paths)
        last_use_row.addWidget(last_use_btn)
        last_use_row.addStretch()
        class_config_layout.addLayout(last_use_row)

        add_path_row("Train数据集标签:", self.train_csv_edit, self.browse_train_csv)
        add_path_row("Val数据集标签:", self.val_csv_edit, self.browse_val_csv)
        add_path_row("Train训练数据目录:", self.train_dir_edit, self.browse_train_dir)
        add_path_row("Val数据目录:", self.val_dir_edit, self.browse_val_dir)

        hp_row1 = QHBoxLayout()
        hp_row1.addWidget(QLabel("早停耐心值:"))
        self.epochs_spin = QDoubleSpinBox()
        self.epochs_spin.setRange(1, 500)
        self.epochs_spin.setDecimals(0)
        self.epochs_spin.setValue(50)
        hp_row1.addWidget(self.epochs_spin)

        hp_row1.addWidget(QLabel("批处理大小:"))
        self.batch_spin = QDoubleSpinBox()
        self.batch_spin.setRange(1, 512)
        self.batch_spin.setDecimals(0)
        self.batch_spin.setValue(32)
        hp_row1.addWidget(self.batch_spin)
        hp_row1.addStretch()
        class_config_layout.addLayout(hp_row1)

        hp_row2 = QHBoxLayout()
        hp_row2.addWidget(QLabel("训练轮数:"))
        self.patience_spin = QDoubleSpinBox()
        self.patience_spin.setRange(1, 100)
        self.patience_spin.setDecimals(0)
        self.patience_spin.setValue(15)
        hp_row2.addWidget(self.patience_spin)
        hp_row2.addStretch()
        class_config_layout.addLayout(hp_row2)

        # TensorBoard 选项（读取主窗口设置）
        tb_group = QGroupBox("TensorBoard (可选)")
        tb_layout = QVBoxLayout(tb_group)
        self.tb_enable_check = QCheckBox("启用 TensorBoard 服务器")
        if self.parent_app:
            self.tb_enable_check.setChecked(self.parent_app.enable_tensorboard)
        tb_layout.addWidget(self.tb_enable_check)
        hint = QLabel("注：如果启用了 TensorBoard 功能，那么在训练过程中它将会自动启动。\n"
                      "此选项仅表明您计划自行运行 `tensorboard --logdir=...` 命令。")
        hint.setStyleSheet("color: #7f8c8d; font-size: 11px;")
        tb_layout.addWidget(hint)

        class_config_layout.addWidget(tb_group)
        config_layout.addWidget(self.class_config_widget)

        # U-Net 专用配置（默认隐藏）
        self.unet_config_widget = QWidget()
        unet_layout = QVBoxLayout(self.unet_config_widget)
        unet_layout.setContentsMargins(0, 5, 0, 0)

        def add_unet_row(lbl, edit, browse_slot):
            r = QHBoxLayout()
            r.addWidget(QLabel(lbl))
            r.addWidget(edit)
            b = QPushButton("浏览...")
            b.clicked.connect(browse_slot)
            r.addWidget(b)
            unet_layout.addLayout(r)

        self.unet_img_dir_edit = QLineEdit()
        self.unet_mask_dir_edit = QLineEdit()
        self.unet_save_edit = QLineEdit()
        self.unet_epochs_spin = QSpinBox()
        self.unet_epochs_spin.setRange(5, 200)
        self.unet_epochs_spin.setValue(30)
        add_unet_row("图像目录:", self.unet_img_dir_edit, lambda: self._browse_unet_dir(self.unet_img_dir_edit))
        add_unet_row("掩码目录:", self.unet_mask_dir_edit, lambda: self._browse_unet_dir(self.unet_mask_dir_edit))
        add_unet_row("模型保存:", self.unet_save_edit, self._browse_unet_save)
        self.unet_save_edit.setPlaceholderText("默认: U-Net训练结果.pth")
        unet_restore_btn = QPushButton("一键使用上次路径")
        unet_restore_btn.clicked.connect(self._restore_last_unet_paths_ml)
        unet_restore_row = QHBoxLayout()
        unet_restore_row.addWidget(unet_restore_btn)
        unet_restore_row.addStretch()
        unet_layout.addLayout(unet_restore_row)
        self.unet_use_cpu_check_ml = QCheckBox("训练时强制使用 CPU（若 GPU 训练闪退可勾选）")
        self.unet_use_cpu_check_ml.setToolTip("勾选后使用 CPU 训练，避免退出码 0xC0000005 等显卡相关崩溃")
        seg_settings = QSettings("RockAnalysisTool", "Segmentation")
        self.unet_use_cpu_check_ml.setChecked(seg_settings.value("unet_train_use_cpu", False, type=bool))
        unet_layout.addWidget(self.unet_use_cpu_check_ml)
        eh = QHBoxLayout()
        eh.addWidget(QLabel("训练轮数:"))
        eh.addWidget(self.unet_epochs_spin)
        eh.addStretch()
        unet_layout.addLayout(eh)
        config_layout.addWidget(self.unet_config_widget)

        self._on_model_type_changed()
        left_layout.addWidget(config_group)

        # 日志组
        log_group = QGroupBox("训练日志")
        log_layout = QVBoxLayout(log_group)
        self.train_log_text = QTextEdit()
        self.train_log_text.setReadOnly(True)
        self.train_log_text.setFont(QFont("Consolas", 10))
        log_layout.addWidget(self.train_log_text)
        left_layout.addWidget(log_group, stretch=1)

        # 控制按钮
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        if self.parent_app:
            screenshot_btn = QPushButton("界面捕获")
            screenshot_btn.setIcon(QIcon(r"resources\assets\images\button\screenshot-fill.png"))
            screenshot_btn.clicked.connect(lambda: self.parent_app.capture_widget_screenshot(self, "ModelTraining"))
            btn_row.addWidget(screenshot_btn)
        self.start_btn = QPushButton("开启训练")
        # self.start_btn.setStyleSheet("""
        #     QPushButton {
        #         background-color: #3498db;
        #         color: white;
        #         padding: 8px 20px;
        #         border-radius: 4px;
        #         font-weight: bold;
        #     }
        #     QPushButton:hover {
        #         background-color: #2980b9;
        #     }
        # """)
        self.start_btn.clicked.connect(self.start_training)
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.close)
        btn_row.addWidget(self.start_btn)
        btn_row.addWidget(close_btn)
        left_layout.addLayout(btn_row)

        main_layout.addWidget(left_panel, stretch=4)

        # 右侧：训练曲线图
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(10)

        plot_group = QGroupBox("训练曲线")
        plot_layout = QVBoxLayout(plot_group)

        self.fig = Figure(figsize=(5, 3), dpi=80)
        self.canvas = FigureCanvasQTAgg(self.fig)
        plot_layout.addWidget(self.canvas)

        right_layout.addWidget(plot_group)
        main_layout.addWidget(right_panel, stretch=5)

    # 路径选择槽函数
    def browse_train_csv(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Train CSV", "", "CSV Files (*.csv)")
        if path:
            self.train_csv_edit.setText(path)

    def browse_val_csv(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Val CSV", "", "CSV Files (*.csv)")
        if path:
            self.val_csv_edit.setText(path)

    def browse_train_dir(self):
        path = QFileDialog.getExistingDirectory(self, "Select Train Image Directory", "")
        if path:
            self.train_dir_edit.setText(path)

    def browse_val_dir(self):
        path = QFileDialog.getExistingDirectory(self, "Select Val Image Directory", "")
        if path:
            self.val_dir_edit.setText(path)

    def _load_last_training_paths(self):
        """加载上次使用的训练路径"""
        settings = QSettings("RoCAS", "ImageEditor")
        for key, edit in [("train_csv", self.train_csv_edit), ("val_csv", self.val_csv_edit),
                          ("train_dir", self.train_dir_edit), ("val_dir", self.val_dir_edit)]:
            val = settings.value(f"training_{key}", "")
            if val and os.path.exists(val if key.endswith("_dir") else val):
                edit.setText(val)

    def restore_last_training_paths(self):
        """一键恢复上次选择的训练路径"""
        settings = QSettings("RoCAS", "ImageEditor")
        changed = False
        for key, edit in [("train_csv", self.train_csv_edit), ("val_csv", self.val_csv_edit),
                          ("train_dir", self.train_dir_edit), ("val_dir", self.val_dir_edit)]:
            val = settings.value(f"training_{key}", "")
            if val:
                edit.setText(val)
                changed = True
        if not changed:
            QMessageBox.information(self, "提示", "暂无上次选择的路径记录。")

    def _on_model_type_changed(self):
        """根据模型类型切换显示分类配置或 U-Net 配置"""
        key = self.model_combo.currentData() or "resnet"
        self.model_desc_label.setText(self.MODEL_DESCRIPTIONS.get(key, ""))
        is_unet = (key == "unet")
        self.class_config_widget.setVisible(not is_unet)
        self.unet_config_widget.setVisible(is_unet)
        if is_unet:
            self._restore_last_unet_paths_ml(fill_only_empty=True)

    def _restore_last_unet_paths_ml(self, fill_only_empty=False):
        """一键恢复上次 U-Net 路径（机器学习面板）"""
        settings = QSettings("RockAnalysisTool", "Segmentation")
        img = settings.value("unet_train_img_dir", "", type=str)
        mask = settings.value("unet_train_mask_dir", "", type=str)
        save = settings.value("unet_train_save_path", "", type=str)
        if fill_only_empty:
            if not self.unet_img_dir_edit.text().strip() and img:
                self.unet_img_dir_edit.setText(img)
            if not self.unet_mask_dir_edit.text().strip() and mask:
                self.unet_mask_dir_edit.setText(mask)
            if not self.unet_save_edit.text().strip():
                if save:
                    self.unet_save_edit.setText(save)
                elif self.unet_img_dir_edit.text().strip():
                    d = os.path.dirname(self.unet_img_dir_edit.text().strip())
                    self.unet_save_edit.setText(os.path.join(d, "U-Net训练结果.pth"))
        else:
            if img:
                self.unet_img_dir_edit.setText(img)
            if mask:
                self.unet_mask_dir_edit.setText(mask)
            if save:
                self.unet_save_edit.setText(save)
            if not any((img, mask, save)):
                QMessageBox.information(self, "提示", "暂无上次路径记录，请先选择目录并训练一次。")

    def _browse_unet_dir(self, line_edit):
        last = line_edit.text() or (self.unet_img_dir_edit.text() if line_edit != self.unet_img_dir_edit else "")
        if self.parent_app and hasattr(self.parent_app, "get_last_directory"):
            last = last or self.parent_app.get_last_directory()
        path = QFileDialog.getExistingDirectory(self, "选择目录", last)
        if path:
            line_edit.setText(path)

    def _browse_unet_save(self):
        default_name = "U-Net训练结果.pth"
        start_dir = ""
        if self.unet_img_dir_edit.text().strip():
            start_dir = os.path.dirname(self.unet_img_dir_edit.text().strip())
        if self.parent_app and hasattr(self.parent_app, "get_last_directory"):
            start_dir = start_dir or self.parent_app.get_last_directory()
        initial = os.path.join(start_dir, default_name) if start_dir else default_name
        path, _ = QFileDialog.getSaveFileName(self, "保存 U-Net 模型", initial, "PyTorch (*.pt *.pth);;All (*)")
        if path:
            if not (path.lower().endswith('.pth') or path.lower().endswith('.pt')):
                path = path + ".pth"
            self.unet_save_edit.setText(path)

    def append_log(self, text):
        time_str = datetime.datetime.now().strftime("%H:%M:%S")
        self.train_log_text.append(f"[{time_str}] {text}")
        self.train_log_text.ensureCursorVisible()
        if self.parent_app:
            self.parent_app.log(text, source="training")

    def start_training(self):
        """启动训练线程"""
        model_key = self.model_combo.currentData() or "resnet"

        if model_key == "unet":
            # U-Net 分割模型训练
            img_dir = self.unet_img_dir_edit.text().strip()
            mask_dir = self.unet_mask_dir_edit.text().strip()
            save_path = self.unet_save_edit.text().strip()
            if not img_dir or not os.path.isdir(img_dir):
                QMessageBox.warning(self, "路径无效", "请选择有效的图像目录")
                return
            if not mask_dir or not os.path.isdir(mask_dir):
                QMessageBox.warning(self, "路径无效", "请选择有效的掩码目录")
                return
            if not save_path:
                save_path = os.path.join(os.path.dirname(img_dir), "U-Net训练结果.pth")
                self.unet_save_edit.setText(save_path)
            if save_path and not save_path.lower().endswith(('.pt', '.pth')):
                save_path = save_path.rstrip('.') + ".pth"
                self.unet_save_edit.setText(save_path)
            settings = QSettings("RockAnalysisTool", "Segmentation")
            settings.setValue("unet_train_img_dir", img_dir)
            settings.setValue("unet_train_mask_dir", mask_dir)
            settings.setValue("unet_train_save_path", save_path)
            use_cpu_ml = self.unet_use_cpu_check_ml.isChecked()
            settings.setValue("unet_train_use_cpu", use_cpu_ml)
            # 主线程预初始化 PyTorch，避免子线程首次初始化导致闪退
            try:
                import torch
                _ = torch.ones(1)
                if torch.cuda.is_available():
                    _ = torch.ones(1, device="cuda")
            except Exception as e:
                self.append_log(f"PyTorch 预检查: {e}，将尝试继续（可能使用 CPU）")
            epochs = self.unet_epochs_spin.value()
            batch_size = 8
            self.start_btn.setEnabled(False)
            self.append_log("开始 U-Net 分割模型训练...")
            self.worker = SegModelTrainingWorker(
                img_dir, mask_dir, save_path, epochs, batch_size, self, use_cpu=use_cpu_ml
            )
            self.worker.log_signal.connect(self.append_log)
            self.worker.finished_signal.connect(self._on_unet_training_finished)
            self.worker.start()
            return

        # 分类模型训练（ResNet/EfficientNet/MobileNet/Custom）
        train_csv = self.train_csv_edit.text().strip()
        val_csv = self.val_csv_edit.text().strip()
        train_dir = self.train_dir_edit.text().strip()
        val_dir = self.val_dir_edit.text().strip()

        if not (os.path.exists(train_csv) and os.path.exists(val_csv) and
                os.path.isdir(train_dir) and os.path.isdir(val_dir)):
            QMessageBox.warning(self, "Invalid Paths", "Please check train/val CSV and directories.")
            return

        model_type = model_key  # resnet, efficientnet, mobilenet, custom
        epochs = int(self.epochs_spin.value())
        batch_size = int(self.batch_spin.value())
        patience = int(self.patience_spin.value())

        enable_tb = self.tb_enable_check.isChecked()
        tb_port = self.parent_app.tensorboard_port if self.parent_app else 6006
        tb_log_dir = self.parent_app.tensorboard_log_dir if (
                self.parent_app and self.parent_app.tensorboard_log_dir) else ""

        # 保存本次路径供下次一键恢复
        settings = QSettings("RoCAS", "ImageEditor")
        settings.setValue("training_train_csv", train_csv)
        settings.setValue("training_val_csv", val_csv)
        settings.setValue("training_train_dir", train_dir)
        settings.setValue("training_val_dir", val_dir)

        self.start_btn.setEnabled(False)
        self.append_log("Starting training in background thread...")

        self.worker = ModelTrainingWorker(
            train_csv,
            train_dir,
            val_csv,
            val_dir,
            model_type,
            epochs,
            batch_size,
            patience,
            enable_tb,
            tb_port,
            tb_log_dir,
            self
        )
        self.worker.log_message.connect(self.append_log)
        self.worker.finished_signal.connect(self.on_training_finished)
        self.worker.error_signal.connect(self.on_training_error)
        self.worker.start()

    def _on_unet_training_finished(self, ok, exit_code=0):
        """U-Net 训练完成"""
        self.start_btn.setEnabled(True)
        if ok:
            self.append_log("U-Net 模型训练完成。请在 设置→分割模型 中指定该模型路径。")
            QMessageBox.information(self, "完成", "U-Net 模型训练完成。请在 设置→分割模型 中指定该模型路径。")
        else:
            self.append_log("U-Net 训练未成功，请查看日志。")
            if exit_code == 3221225477:
                QMessageBox.warning(
                    self, "训练异常退出",
                    "训练子进程闪退（3221225477）。已勾选 CPU 仍闪退时，请安装纯 CPU 版 PyTorch 后重试：\n"
                    "pip install torch --index-url https://download.pytorch.org/whl/cpu"
                )

    def on_training_finished(self, results):
        """训练完成"""
        self.results = results
        self.start_btn.setEnabled(True)
        self.append_log("Training completed. Updating plots...")
        try:
            history = results.get("history", {})
            self.update_plots(history)
        except Exception as e:
            self.append_log(f"Failed to update plots: {e}")
        # 若生成了评估报告，提示用户并可选打开模型评估窗口
        eval_path = results.get("eval_report_path")
        if eval_path:
            eval_path = os.path.abspath(eval_path) if not os.path.isabs(eval_path) else eval_path
        if eval_path and os.path.isdir(eval_path):
            report_png = os.path.join(eval_path, "evaluation_report.png")
            if os.path.isfile(report_png):
                self.append_log(f"评估报告已生成: {eval_path}")
                if self.parent_app:
                    self.parent_app.log(f"评估报告目录: {eval_path}")
                open_eval = QMessageBox.question(
                    self,
                    "训练完成",
                    "颜色分类模型训练已完成，评估报告已生成。\n\n是否在「模型评估」窗口中打开该报告？",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes
                )
                if open_eval == QMessageBox.StandardButton.Yes and self.parent_app:
                    self.parent_app.model_evaluation_window = ModelEvaluationWindow(self.parent_app)
                    self.parent_app.model_evaluation_window.load_report_directory(eval_path)
                    self.parent_app.model_evaluation_window.show()
        elif eval_path:
            self.append_log(f"评估报告路径（请手动在模型评估中选择）: {eval_path}")

    def on_training_error(self, msg):
        """训练出错"""
        self.start_btn.setEnabled(True)
        self.append_log(msg)
        QMessageBox.critical(self, "Training Error", msg)

    def update_plots(self, history):
        """根据 history 绘制 Loss / Accuracy 曲线"""
        self.fig.clear()
        ax1 = self.fig.add_subplot(2, 1, 1)
        ax2 = self.fig.add_subplot(2, 1, 2)

        epochs = list(range(1, len(history.get("train_loss", [])) + 1))

        if epochs:
            ax1.plot(epochs, history.get("train_loss", []), label="Train Loss")
            ax1.plot(epochs, history.get("val_loss", []), label="Val Loss")
            ax1.set_xlabel("Epoch")
            ax1.set_ylabel("Loss")
            ax1.legend()
            ax1.grid(True, alpha=0.3)

            ax2.plot(epochs, history.get("train_acc", []), label="Train Acc")
            ax2.plot(epochs, history.get("val_acc", []), label="Val Acc")
            ax2.set_xlabel("Epoch")
            ax2.set_ylabel("Accuracy (%)")
            ax2.legend()
            ax2.grid(True, alpha=0.3)

        self.fig.tight_layout()
        self.canvas.draw()


# ========== 模型评估窗口 ==========
class ModelEvaluationWindow(QMainWindow):
    """模型评估窗口，显示评估结果和后处理可视化"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_app = parent
        self.eval_report_path = None
        self.eval_data = {}

        self.setWindowTitle("模型评估参数设定")
        self.setWindowIcon(QIcon(r"resources\assets\images\button\evaluation.png"))
        self.resize(700, 500)

        self.init_ui()

    def init_ui(self):
        """初始化UI"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)

        # 顶部：标题 + 使用说明 + 目录选择按钮
        header_layout = QHBoxLayout()
        title_label = QLabel("模型评估结果")
        title_label.setStyleSheet("font-weight: bold; color: #34495e; font-size: 13px;")
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        help_label = QLabel(
            "说明：本窗口用于<b>查看</b>已生成的评估报告，不执行评估。\n"
            "请点击下方按钮选择<b>包含 evaluation_report.png 的目录</b>。\n"
            "该目录由「模型训练」中完成<b>颜色识别分类模型</b>训练后自动生成（通常在 所有结果/eval_report_时间戳 下）。\n"
            "若仅做了「用模型识别」而未在训练流程中生成报告，请先在「模型训练」中训练颜色分类模型，训练结束时会自动生成评估报告目录。"
        )
        help_label.setStyleSheet("color: #7f8c8d; font-size: 11px; padding: 6px;")
        help_label.setWordWrap(True)
        help_label.setToolTip("评估报告 = 在验证集上跑分类后生成的混淆矩阵、准确率等；识别单张图不会生成此报告。")
        layout.addWidget(help_label)

        select_btn = QPushButton("选择评估保存目录")
        select_btn.setIcon(QIcon(r"resources\assets\images\button\addFolder.png"))
        select_btn.clicked.connect(self.select_eval_directory)
        header_layout.addWidget(select_btn)
        if self.parent_app:
            screenshot_btn = QPushButton("界面捕获")
            screenshot_btn.setIcon(QIcon(r"resources\assets\images\button\screenshot-fill.png"))
            screenshot_btn.clicked.connect(lambda: self.parent_app.capture_widget_screenshot(self, "ModelEvaluation"))
            header_layout.addWidget(screenshot_btn)

        layout.addLayout(header_layout)

        # 主内容区域（使用标签页）
        self.content_tabs = QTabWidget()
        self.content_tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #ddd;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #ecf0f1;
                color: #34495e;
                padding: 8px 20px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background-color: #3498db;
                color: white;
            }
        """)

        # 评估报告标签页
        report_tab = self.create_report_tab()
        self.content_tabs.addTab(report_tab, "评估报告")

        # 错误分类样本标签页
        errors_tab = self.create_errors_tab()
        self.content_tabs.addTab(errors_tab, "错误分类样本")

        # 详细指标标签页
        metrics_tab = self.create_metrics_tab()
        self.content_tabs.addTab(metrics_tab, "详细指标")

        layout.addWidget(self.content_tabs)

        # 底部按钮
        btn_frame = QFrame()
        btn_layout = QHBoxLayout(btn_frame)
        btn_layout.addStretch()

        self.export_btn = QPushButton("导出结果图")
        # self.export_btn.setIcon(QIcon(r"resources\assets\images\button\save.png"))
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_figure)
        btn_layout.addWidget(self.export_btn)

        open_folder_btn = QPushButton("从文件夹打开")
        open_folder_btn.setEnabled(False)
        open_folder_btn.clicked.connect(self.open_eval_folder)
        btn_layout.addWidget(open_folder_btn)

        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.close)
        btn_layout.addWidget(close_btn)

        layout.addWidget(btn_frame)

        self.open_folder_btn = open_folder_btn

    def create_report_tab(self):
        """创建评估报告标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        # 评估报告图片显示区域
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("border: 1px solid #ddd; border-radius: 5px;")

        self.report_label = QLabel(
            "请点击「选择评估保存目录」，选择包含 evaluation_report.png 的文件夹（通常由模型训练完成后自动生成）。")
        self.report_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.report_label.setStyleSheet("""
            QLabel {
                border: 2px dashed #bdc3c7;
                border-radius: 10px;
                background-color: #f8f9fa;
                color: #7f8c8d;
                padding: 50px;
                font-size: 14px;
            }
        """)
        scroll_area.setWidget(self.report_label)

        layout.addWidget(scroll_area)
        return widget

    def create_errors_tab(self):
        """创建错误分类样本标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        # 错误分类样本表格
        self.errors_table = QTableWidget()
        self.errors_table.setColumnCount(4)
        self.errors_table.setHorizontalHeaderLabels(["Image Path", "True Label", "Predicted Label", "Confidence"])
        self.errors_table.horizontalHeader().setStretchLastSection(True)
        self.errors_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.errors_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 5px;
                gridline-color: #e0e0e0;
            }
            QTableWidget::item {
                padding: 5px;
            }
        """)

        layout.addWidget(self.errors_table)
        return widget

    def create_metrics_tab(self):
        """创建详细指标标签页"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        # 详细指标表格
        self.metrics_table = QTableWidget()
        self.metrics_table.setColumnCount(2)
        self.metrics_table.setHorizontalHeaderLabels(["Metric", "Value"])
        self.metrics_table.horizontalHeader().setStretchLastSection(True)
        self.metrics_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.metrics_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 5px;
                gridline-color: #e0e0e0;
            }
            QTableWidget::item {
                padding: 5px;
            }
        """)

        layout.addWidget(self.metrics_table)
        return widget

    def select_eval_directory(self):
        """选择评估结果目录"""
        last_dir = self.eval_report_path if self.eval_report_path else os.getcwd()
        dir_path = QFileDialog.getExistingDirectory(
            self,
            "Select Evaluation Report Directory",
            last_dir
        )

        if not dir_path:
            return

        self.eval_report_path = dir_path
        self.load_evaluation_results()

    def load_report_directory(self, dir_path):
        """直接加载指定目录的评估报告（供训练完成后自动打开时调用）"""
        if not dir_path or not os.path.isdir(dir_path):
            return
        self.eval_report_path = os.path.abspath(dir_path)
        self.load_evaluation_results()

    def load_evaluation_results(self):
        """加载评估结果"""
        try:
            # 查找evaluation_report.png
            report_img_path = os.path.join(self.eval_report_path, "evaluation_report.png")
            if not os.path.exists(report_img_path):
                QMessageBox.warning(
                    self,
                    "未找到评估报告",
                    f"该目录下没有 evaluation_report.png：\n{self.eval_report_path}\n\n"
                    "请选择由「模型训练」中颜色分类训练完成后生成的评估目录（形如 所有结果/eval_report_时间戳），"
                    "或先完成一次颜色识别模型训练以自动生成报告。"
                )
                return

            # 加载评估报告图片
            pixmap = QPixmap(report_img_path)
            if pixmap.isNull():
                raise Exception("Failed to load evaluation report image")

            # 显示图片（缩放以适应窗口）
            self.report_label.setPixmap(pixmap.scaled(
                pixmap.size(),
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation
            ))
            self.report_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

            # 加载错误分类样本CSV
            errors_csv_path = os.path.join(self.eval_report_path, "misclassified_samples.csv")
            if os.path.exists(errors_csv_path):
                self.load_errors_csv(errors_csv_path)

            # 加载详细报告CSV
            detailed_csv_path = os.path.join(self.eval_report_path, "detailed_report.csv")
            if os.path.exists(detailed_csv_path):
                self.load_metrics_csv(detailed_csv_path)

            # 启用按钮
            self.export_btn.setEnabled(True)
            self.open_folder_btn.setEnabled(True)

            if self.parent_app:
                self.parent_app.log(f"Evaluation results loaded from: {self.eval_report_path}")

        except Exception as e:
            QMessageBox.critical(
                self,
                "Error Loading Results",
                f"Failed to load evaluation results:\n{str(e)}"
            )
            if self.parent_app:
                self.parent_app.log(f"Error loading evaluation results: {str(e)}")

    def load_errors_csv(self, csv_path):
        """加载错误分类样本CSV"""
        try:
            df = pd.read_csv(csv_path)
            self.errors_table.setRowCount(len(df))

            for row_idx, (_, row) in enumerate(df.iterrows()):
                self.errors_table.setItem(row_idx, 0, QTableWidgetItem(str(row.get('image_path', ''))))
                self.errors_table.setItem(row_idx, 1, QTableWidgetItem(str(row.get('true', ''))))
                self.errors_table.setItem(row_idx, 2, QTableWidgetItem(str(row.get('predicted', ''))))
                self.errors_table.setItem(row_idx, 3, QTableWidgetItem(f"{row.get('confidence', 0):.4f}"))
        except Exception as e:
            if self.parent_app:
                self.parent_app.log(f"Error loading errors CSV: {str(e)}")

    def load_metrics_csv(self, csv_path):
        """加载详细指标CSV"""
        try:
            df = pd.read_csv(csv_path, index_col=0)
            # 提取总体指标
            metrics = []

            if 'accuracy' in df.index:
                metrics.append(("Overall Accuracy", f"{df.loc['accuracy', 'f1-score']:.4f}"))

            if 'macro avg' in df.index:
                metrics.append(("Macro Precision", f"{df.loc['macro avg', 'precision']:.4f}"))
                metrics.append(("Macro Recall", f"{df.loc['macro avg', 'recall']:.4f}"))
                metrics.append(("Macro F1-Score", f"{df.loc['macro avg', 'f1-score']:.4f}"))

            if 'weighted avg' in df.index:
                metrics.append(("Weighted Precision", f"{df.loc['weighted avg', 'precision']:.4f}"))
                metrics.append(("Weighted Recall", f"{df.loc['weighted avg', 'recall']:.4f}"))
                metrics.append(("Weighted F1-Score", f"{df.loc['weighted avg', 'f1-score']:.4f}"))

            # 添加每个类别的指标
            for idx in df.index:
                if idx not in ['accuracy', 'macro avg', 'weighted avg']:
                    metrics.append((f"{idx} - Precision", f"{df.loc[idx, 'precision']:.4f}"))
                    metrics.append((f"{idx} - Recall", f"{df.loc[idx, 'recall']:.4f}"))
                    metrics.append((f"{idx} - F1-Score", f"{df.loc[idx, 'f1-score']:.4f}"))
                    metrics.append((f"{idx} - Support", f"{int(df.loc[idx, 'support'])}"))

            self.metrics_table.setRowCount(len(metrics))
            for row_idx, (metric_name, metric_value) in enumerate(metrics):
                self.metrics_table.setItem(row_idx, 0, QTableWidgetItem(metric_name))
                self.metrics_table.setItem(row_idx, 1, QTableWidgetItem(metric_value))
        except Exception as e:
            if self.parent_app:
                self.parent_app.log(f"Error loading metrics CSV: {str(e)}")

    def export_figure(self):
        """导出评估报告图片：支持 PNG/JPEG/TIFF/PDF 多格式与高 DPI（论文用图）"""
        if not self.eval_report_path:
            return

        report_img_path = os.path.join(self.eval_report_path, "evaluation_report.png")
        if not os.path.exists(report_img_path):
            QMessageBox.warning(self, "File Not Found", "evaluation_report.png not found.")
            return

        dpi = self.parent_app.dpi if self.parent_app else 300
        default_name = f"evaluation_report_dpi{dpi}.png"
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出评估报告图（可选格式与高DPI）",
            default_name,
            IMAGE_EXPORT_FILTER,
        )
        if not save_path:
            return

        try:
            fmt, _ = get_export_format_from_path(save_path)
            pil_img = Image.open(report_img_path)

            if fmt == 'pdf':
                fig, ax = plt.subplots(dpi=dpi)
                ax.imshow(np.array(pil_img))
                ax.axis('off')
                plt.tight_layout(pad=0)
                fig.savefig(save_path, dpi=dpi, format='pdf', bbox_inches='tight', pad_inches=0)
                plt.close(fig)
            elif fmt == 'tif':
                pil_img.save(save_path, format='TIFF', dpi=(float(dpi), float(dpi)), compression="tiff_lzw")
            elif fmt == 'jpg':
                if pil_img.mode in ('RGBA', 'P'):
                    pil_img = pil_img.convert('RGB')
                pil_img.save(save_path, format='JPEG', quality=95, dpi=(float(dpi), float(dpi)))
            else:
                pil_img.save(save_path, format='PNG', dpi=(float(dpi), float(dpi)))
            QMessageBox.information(
                self,
                "Export Successful",
                f"Evaluation report exported to:\n{save_path}"
            )
            if self.parent_app:
                self.parent_app.log(f"Exported evaluation report to: {save_path}")
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"Failed to export figure:\n{str(e)}"
            )

    def open_eval_folder(self):
        """打开评估结果文件夹"""
        if not self.eval_report_path:
            return

        if platform.system() == "Windows":
            subprocess.run(['explorer', os.path.normpath(self.eval_report_path)])
        else:
            QDesktopServices.openUrl(QUrl.fromLocalFile(self.eval_report_path))


# ========== 机器学习流程可视化面板 ==========
class MLPipelineVisualizationPanel(QWidget):
    """机器学习流程可视化面板，显示完整的ML流程"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_app = parent
        self.pipeline_data = {}

        self.init_ui()

    def init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)

        # 标题
        title_label = QLabel("Machine Learning Pipeline Overview")
        title_label.setStyleSheet("""
            font-weight: bold;
            font-size: 16px;
            color: #2c3e50;
            padding: 10px;
        """)
        layout.addWidget(title_label)

        # 流程步骤卡片
        steps_layout = QVBoxLayout()
        steps_layout.setSpacing(10)

        # 步骤1：数据预处理
        step1_card = self.create_step_card(
            "Step 1: Data Preprocessing",
            "Image segmentation, color analysis, and label generation",
            "#3498db"
        )
        steps_layout.addWidget(step1_card)

        # 步骤2：数据准备
        step2_card = self.create_step_card(
            "Step 2: Data Preparation",
            "Dataset splitting (train/val/test) and CSV generation",
            "#2ecc71"
        )
        steps_layout.addWidget(step2_card)

        # 步骤3：模型训练
        step3_card = self.create_step_card(
            "Step 3: Model Training",
            "Deep learning model training with data augmentation",
            "#e74c3c"
        )
        steps_layout.addWidget(step3_card)

        # 步骤5：后处理
        step5_card = self.create_step_card(
            "Step 5: Post-processing",
            "Result visualization and paper-ready figure generation",
            "#9b59b6"
        )
        steps_layout.addWidget(step5_card)

        layout.addLayout(steps_layout)

        # 统计信息区域
        stats_group = QGroupBox("Pipeline Statistics")
        stats_layout = QGridLayout()

        self.stats_labels = {}
        stats_items = [
            ("Total Images", "total_images"),
            ("Classes", "num_classes"),
            ("Train Samples", "train_samples"),
            ("Val Samples", "val_samples"),
            ("Test Samples", "test_samples"),
            ("Model Type", "model_type"),
            ("Best Accuracy", "best_accuracy"),
            ("Training Epochs", "training_epochs")
        ]

        row = 0
        col = 0
        for label_text, key in stats_items:
            label = QLabel(f"{label_text}:")
            label.setStyleSheet("font-weight: bold; color: #34495e;")
            value_label = QLabel("N/A")
            value_label.setStyleSheet("color: #7f8c8d;")
            self.stats_labels[key] = value_label

            stats_layout.addWidget(label, row, col * 2)
            stats_layout.addWidget(value_label, row, col * 2 + 1)

            col += 1
            if col >= 2:
                col = 0
                row += 1

        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)

        layout.addStretch()

    def create_step_card(self, title, description, color):
        """创建流程步骤卡片"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: white;
                border: 2px solid {color};
                border-radius: 8px;
                padding: 15px;
            }}
        """)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(10, 10, 10, 10)

        title_label = QLabel(title)
        title_label.setStyleSheet(f"""
            font-weight: bold;
            font-size: 14px;
            color: {color};
        """)
        layout.addWidget(title_label)

        desc_label = QLabel(description)
        desc_label.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        layout.addWidget(desc_label)

        return card

    def update_statistics(self, data):
        """更新统计信息"""
        self.pipeline_data = data

        for key, label in self.stats_labels.items():
            value = data.get(key, "N/A")
            if isinstance(value, float):
                label.setText(f"{value:.2f}%")
            else:
                label.setText(str(value))


if __name__ == '__main__':
    app = QApplication(sys.argv)
    # 避免 Qt 使用有问题的 Fixedsys 字体导致 CreateFontFaceFromHDC 报错
    app.setFont(QFont("Microsoft YaHei", 9))

    # 解决 Matplotlib 中文乱码
    matplotlib.use('QtAgg')  # 声明使用 Qt 后端
    matplotlib.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']
    matplotlib.rcParams['axes.unicode_minus'] = False
    # 降低嵌入式图表内存占用，避免 copy_from_bbox Out of memory
    matplotlib.rcParams['figure.dpi'] = 80
    matplotlib.rcParams['figure.max_open_warning'] = 0

    window = ImageEditorApp()
    window.show()
    sys.exit(app.exec())
